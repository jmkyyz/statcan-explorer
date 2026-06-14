#!/usr/bin/env python3
"""
StatCan WDS Proxy Server
========================
Sits between your browser and the Statistics Canada WDS API, handling
CORS and translating raw WDS responses into a clean JSON format the
frontend can consume directly.

Usage:
    pip install flask flask-cors requests
    python proxy.py

The server will listen on http://localhost:5001
"""

import base64
import csv
import gzip
import io
import json
import os
import re
import shutil
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, date
from email.utils import formatdate, parsedate_to_datetime

import requests
from dotenv import load_dotenv
load_dotenv()
from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS

app = Flask(__name__)
CORS(app)  # Allow all origins (restrict in production)

# ---------------------------------------------------------------------------
# Route: GET /  →  serve the frontend
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    here = os.path.dirname(os.path.abspath(__file__))
    return send_from_directory(here, "statcan-explorer.html")

@app.route("/vectors-template.xlsx")
def vectors_template():
    here = os.path.dirname(os.path.abspath(__file__))
    return send_from_directory(here, "vectors-template.xlsx")

@app.route("/vectors.xlsx")
def vectors():
    here = os.path.dirname(os.path.abspath(__file__))
    return send_from_directory(here, "Vectors.xlsx")

@app.route("/lab")
def lab():
    here = os.path.dirname(os.path.abspath(__file__))
    return send_from_directory(here, "statcan-explorer-lab.html")

@app.route("/tax")
def tax_tracker():
    here = os.path.dirname(os.path.abspath(__file__))
    return send_from_directory(here, "tax-dollar-tracker.html")

EV_DIST = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ev-dashboard", "dist")

@app.route("/ev")
@app.route("/ev/")
def ev_dashboard():
    return send_from_directory(EV_DIST, "index.html")

@app.route("/ev/<path:filename>")
def ev_static(filename):
    target = os.path.join(EV_DIST, filename)
    if os.path.isfile(target):
        return send_from_directory(EV_DIST, filename)
    return send_from_directory(EV_DIST, "index.html")

@app.route("/ckan/<path:ckan_path>")
def ckan_proxy(ckan_path):
    """Proxy for open.canada.ca CKAN API — strips WAF-triggering headers."""
    url = f"https://open.canada.ca/data/en/api/3/action/{ckan_path}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-CA,en-US;q=0.9,en;q=0.8",
    }
    try:
        resp = requests.get(url, params=request.args, headers=headers, timeout=30)
        from flask import Response
        return Response(resp.content, status=resp.status_code,
                        content_type=resp.headers.get("Content-Type", "application/json"))
    except requests.exceptions.RequestException as exc:
        return jsonify({"error": str(exc)}), 502

STATCAN_BASE   = "https://www150.statcan.gc.ca/t1/wds/rest"
BOC_BASE       = "https://www.bankofcanada.ca/valet"
CIMT_REF_BASE  = "https://www150.statcan.gc.ca/n1/pub/71-607-x/2021004"
CIMT_ZIP_BASE  = "https://www150.statcan.gc.ca/n1/pub/71-607-x/2021004/zip"
CIMT_CACHE_DIR = "/tmp/cimt_cache"

# CIMT province numeric ID → 2-letter CSV code  (1 = Canada total, no filter)
PROVINCE_CODES: dict[str, str | None] = {
    "1":  None,
    "10": "NL", "11": "PE", "12": "NS", "13": "NB",
    "24": "QC", "35": "ON", "46": "MB", "47": "SK",
    "48": "AB", "59": "BC", "60": "YT", "61": "NT", "62": "NU",
}

# ---------------------------------------------------------------------------
# Scalar factor multipliers (from StatCan codeset)
# ---------------------------------------------------------------------------
SCALAR = {
    0: 1,           # units
    1: 10,
    2: 100,
    3: 1_000,
    4: 10_000,
    5: 100_000,
    6: 1_000_000,
    7: 10_000_000,
    8: 100_000_000,
    9: 1_000_000_000,
}

# ---------------------------------------------------------------------------
# Unit of Measure (UOM) codes from StatCan codeset.
# Each entry: (base_unit_label, multiplier_to_convert_to_base_unit)
# The multiplier is applied ON TOP of the scalar factor so that the frontend
# always receives values in the stated base unit (dollars, persons, etc.).
# ---------------------------------------------------------------------------
UOM_INFO = {
    0:   ("",                    1),
    9:   ("number",              1),
    14:  ("persons",         1_000),      # reported as thousands of persons
    17:  ("index",               1),      # index (e.g. CPI 2002=100)
    18:  ("percent",             1),
    20:  ("index",               1),
    21:  ("index",               1),
    39:  ("persons",             1),
    47:  ("hours",               1),
    48:  ("hours",               1),
    56:  ("dollars/hour",        1),
    81:  ("dollars",             1),
    115: ("",                1_000),      # generic thousands
    224: ("dollars",         1_000),      # thousands of dollars → dollars
    229: ("dollars",     1_000_000),      # millions of dollars  → dollars
    246: ("dollars", 1_000_000_000),      # billions of dollars  → dollars
    239: ("percent",             1),      # "Percent"
    242: ("percent",             1),      # "Percentage"
    300: ("units",               1),      # "Units" (e.g. vehicle registrations)
    301: ("vehicle-km",          1),      # "Vehicle-kilometres"
    302: ("vehicles",            1),      # "Vehicles"
    396: ("ppts",                1),      # "Percentage point"
    428: ("persons",             1),      # persons (scalar handles scale)
}

# Frequency code -> human label
FREQ_LABEL = {
    1: "Daily",
    2: "Weekly",
    4: "Biweekly",
    6: "Monthly",
    9: "Quarterly",
    11: "Semi-annual",
    12: "Annual",
}

# ---------------------------------------------------------------------------
# Helper: convert a refPer date string + frequencyCode into a display label
# ---------------------------------------------------------------------------
def period_label(ref_per: str, freq_code: int) -> str:
    """
    StatCan returns refPer as YYYY-MM-DD always.
    We map it to a friendly label based on frequency:
      Monthly  -> "2023-01"
      Quarterly-> "2023 Q1"
      Annual   -> "2023"
    """
    try:
        d = datetime.strptime(ref_per[:10], "%Y-%m-%d")
    except ValueError:
        return ref_per

    if freq_code == 12:                    # Annual
        return str(d.year)
    if freq_code == 9:                     # Quarterly
        q = (d.month - 1) // 3 + 1
        return f"{d.year} Q{q}"
    if freq_code == 6:                     # Monthly
        return f"{d.year}-{d.month:02d}"
    if freq_code == 2:                     # Weekly
        return f"{d.year}-W{d.isocalendar()[1]:02d}"
    # Default: return ISO date
    return ref_per[:10]


# ---------------------------------------------------------------------------
# Route: GET /api/series
# Query params:
#   vectors  – comma-separated vector IDs, e.g. "41690973,2062809"
#   fromDate – ISO date string YYYY-MM-DD (preferred)
#   toDate   – ISO date string YYYY-MM-DD (preferred)
#   from     – start year fallback, e.g. "2010"
#   to       – end year fallback, e.g. "2024"
#   periods  – how many latest periods to request from StatCan (default 360)
# ---------------------------------------------------------------------------
@app.route("/api/series")
def get_series():
    raw_vectors   = request.args.get("vectors",  "")
    from_date_str = request.args.get("fromDate", "")
    to_date_str   = request.args.get("toDate",   "")
    from_year     = request.args.get("from",     type=int)
    to_year       = request.args.get("to",       type=int)
    n_periods     = request.args.get("periods",  default=360, type=int)
    agg           = request.args.get("agg",      "")

    if not raw_vectors:
        return jsonify({"error": "No vectors specified"}), 400

    vector_ids = [v.strip().lstrip("vV") for v in raw_vectors.split(",") if v.strip()]
    if not vector_ids:
        return jsonify({"error": "No valid vector IDs"}), 400

    # Clamp n_periods to something reasonable
    # Daily series can require up to 3650+ periods for a 10-year range
    n_periods = min(max(n_periods, 1), 4000)

    # ------------------------------------------------------------------
    # Steps 1+2: Fire both StatCan requests in parallel.
    #   • getSeriesInfoFromVector     → memberUomCode + scalarFactorCode
    #   • getDataFromVectorsAndLatestNPeriods → actual data points
    # The data endpoint returns scalarFactorCode=None at the series level
    # (it lives per data-point there), so we must rely on the info endpoint.
    # ------------------------------------------------------------------
    info_payload = [{"vectorId": int(v)} for v in vector_ids]
    data_payload = [{"vectorId": int(v), "latestN": n_periods} for v in vector_ids]

    info_result = {}   # will hold the raw JSON from getSeriesInfoFromVector
    data_result = {}   # will hold the raw JSON from getDataFromVectorsAndLatestNPeriods
    data_error  = None

    def _fetch_info():
        r = requests.post(
            f"{STATCAN_BASE}/getSeriesInfoFromVector",
            json=info_payload, timeout=45,
            headers={"Content-Type": "application/json"},
        )
        r.raise_for_status()
        return r.json()

    def _fetch_data():
        r = requests.post(
            f"{STATCAN_BASE}/getDataFromVectorsAndLatestNPeriods",
            json=data_payload, timeout=60,
            headers={"Content-Type": "application/json"},
        )
        r.raise_for_status()
        return r.json()

    with ThreadPoolExecutor(max_workers=2) as pool:
        fut_info = pool.submit(_fetch_info)
        fut_data = pool.submit(_fetch_data)
        # Data fetch is mandatory; info fetch is best-effort
        try:
            data_result = fut_data.result()
        except requests.exceptions.Timeout:
            return jsonify({"error": "StatCan API timed out – try fewer periods or try again later"}), 504
        except requests.exceptions.RequestException as exc:
            return jsonify({"error": f"StatCan API error: {exc}"}), 502
        try:
            info_result = fut_info.result()
        except Exception:
            info_result = []   # fallback: no UOM conversion

    # Build lookup dicts from series info
    uom_by_vector: dict[int, int] = {}
    scalar_by_vector: dict[int, int] = {}
    for info_item in (info_result or []):
        if info_item.get("status") == "SUCCESS":
            io = info_item["object"]
            vid = io.get("vectorId")
            if vid is not None:
                uom_by_vector[int(vid)]    = int(io.get("memberUomCode",    0) or 0)
                scalar_by_vector[int(vid)] = int(io.get("scalarFactorCode", 0) or 0)

    raw = data_result

    # ------------------------------------------------------------------
    # Step 3: Resolve date boundaries
    # ------------------------------------------------------------------
    start_date = None
    end_date   = None

    if from_date_str:
        try:
            start_date = datetime.strptime(from_date_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    elif from_year:
        start_date = date(from_year, 1, 1)

    if to_date_str:
        try:
            end_date = datetime.strptime(to_date_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    elif to_year:
        end_date = date(to_year, 12, 31)

    # ------------------------------------------------------------------
    # Step 4: Parse response into clean series objects
    # ------------------------------------------------------------------
    results = []

    for item in raw:
        if item.get("status") != "SUCCESS":
            results.append({
                "vectorId": None,
                "error": item.get("object", "Unknown error from StatCan"),
            })
            continue

        obj = item["object"]
        vector_id    = obj.get("vectorId")
        freq_code    = None

        # Use scalarFactorCode and memberUomCode from the pre-fetched series
        # info (getSeriesInfoFromVector).  The data endpoint returns
        # scalarFactorCode = None at the series level (it's per data-point
        # there), so we must rely on the info endpoint for correctness.
        vid_int      = int(vector_id) if vector_id is not None else None
        scalar_code  = scalar_by_vector.get(vid_int, 0) if vid_int else 0
        multiplier   = SCALAR.get(scalar_code, 1)

        uom_code     = uom_by_vector.get(vid_int, 0) if vid_int else 0
        uom_label, uom_mult = UOM_INFO.get(uom_code, ("", 1))
        total_mult   = multiplier * uom_mult   # scalar × UOM conversion

        data_points  = []

        for dp in obj.get("vectorDataPoint", []):
            ref_per   = dp.get("refPer", "")
            raw_value = dp.get("value")
            freq_code = dp.get("frequencyCode", freq_code)

            # Skip suppressed / unavailable data points
            if raw_value is None or dp.get("statusCode") in (1, 8, 9):
                continue

            # Date range filter
            if start_date or end_date:
                try:
                    dp_date = datetime.strptime(ref_per[:10], "%Y-%m-%d").date()
                    if start_date and dp_date < start_date:
                        continue
                    if end_date and dp_date > end_date:
                        continue
                except ValueError:
                    pass

            # Apply scalar × UOM multiplier to convert to base unit
            try:
                value = float(raw_value) * total_mult
            except (TypeError, ValueError):
                continue

            label = period_label(ref_per, freq_code or 6)
            data_points.append({"label": label, "date": ref_per[:10], "value": value})

        # Aggregate daily → monthly sums when requested
        if agg == "monthly_sum" and freq_code == 1:
            monthly: dict[str, float] = {}
            for dp in data_points:
                mk = dp["date"][:7]          # "YYYY-MM"
                monthly[mk] = monthly.get(mk, 0.0) + dp["value"]
            data_points = [
                {"label": mk, "date": mk + "-01", "value": v}
                for mk, v in sorted(monthly.items())
            ]
            freq_code = 6                    # treat as monthly going forward

        results.append({
            "vectorId":        vector_id,
            "frequency":       FREQ_LABEL.get(freq_code, "Unknown") if freq_code else "Unknown",
            "frequencyCode":   freq_code,
            "scalarFactorCode": scalar_code,
            "uomCode":         uom_code,
            "uom":             uom_label,   # base unit after conversion (e.g. "dollars")
            "multiplier":      total_mult,  # scalar × UOM multiplier actually applied
            "data":            data_points,
        })

    return jsonify({"series": results})


# ---------------------------------------------------------------------------
# Route: GET /api/metadata
# Query params:
#   vectors – comma-separated vector IDs
# ---------------------------------------------------------------------------
@app.route("/api/metadata")
def get_metadata():
    raw_vectors = request.args.get("vectors", "")
    if not raw_vectors:
        return jsonify({"error": "No vectors specified"}), 400

    vector_ids = [v.strip().lstrip("vV") for v in raw_vectors.split(",") if v.strip()]
    payload = [{"vectorId": int(v)} for v in vector_ids]

    try:
        resp = requests.post(
            f"{STATCAN_BASE}/getSeriesInfoFromVector",
            json=payload,
            timeout=20,
            headers={"Content-Type": "application/json"},
        )
        resp.raise_for_status()
    except requests.exceptions.RequestException as exc:
        return jsonify({"error": f"StatCan API error: {exc}"}), 502

    raw = resp.json()
    results = []
    for item in raw:
        if item.get("status") != "SUCCESS":
            results.append({"error": item.get("object", "Error")})
            continue
        obj = item["object"]
        results.append({
            "vectorId":     obj.get("vectorId"),
            "productId":    obj.get("productId"),
            "coordinate":   obj.get("coordinate"),
            "titleEn":      obj.get("SeriesTitleEn", ""),
            "titleFr":      obj.get("SeriesTitleFr", ""),
            "frequencyCode": obj.get("frequencyCode"),
            "frequency":    FREQ_LABEL.get(obj.get("frequencyCode"), "Unknown"),
            "scalarFactorCode": obj.get("scalarFactorCode", 0),
            "terminated":   obj.get("terminated", 0),
        })

    return jsonify({"metadata": results})


# ---------------------------------------------------------------------------
# Route: GET /api/vector-health
# Query params:
#   vectors – comma-separated vector IDs (small set: the user's active series)
# Lightweight liveness probe so the UI can flag dead series the moment they're
# added.  Reports two conditions per vector:
#   • terminated – StatCan has discontinued the series
#   • empty      – the vector returns no usable observations
# Also returns the latest reference period (handy for the "last data" tooltip).
# ---------------------------------------------------------------------------
def _is_terminated(val):
    # StatCan's `terminated` field is inconsistent across vectors (0/1, "0"/"1",
    # null, or a termination date string).  Treat any "non-zero / non-empty"
    # value as terminated; verified against live vectors.
    return val not in (None, "", 0, "0", False)


@app.route("/api/vector-health")
def vector_health():
    raw_vectors = request.args.get("vectors", "")
    if not raw_vectors:
        return jsonify({"error": "No vectors specified"}), 400

    vector_ids = [v.strip().lstrip("vV") for v in raw_vectors.split(",") if v.strip()]
    if not vector_ids:
        return jsonify({"error": "No valid vector IDs"}), 400

    info_payload = [{"vectorId": int(v)} for v in vector_ids]
    data_payload = [{"vectorId": int(v), "latestN": 1} for v in vector_ids]

    def _post(endpoint, payload):
        r = requests.post(
            f"{STATCAN_BASE}/{endpoint}",
            json=payload, timeout=30,
            headers={"Content-Type": "application/json"},
        )
        r.raise_for_status()
        return r.json()

    try:
        with ThreadPoolExecutor(max_workers=2) as pool:
            fut_info = pool.submit(_post, "getSeriesInfoFromVector", info_payload)
            fut_data = pool.submit(_post, "getDataFromVectorsAndLatestNPeriods", data_payload)
            info_raw = fut_info.result()
            data_raw = fut_data.result()
    except requests.exceptions.RequestException as exc:
        return jsonify({"error": f"StatCan API error: {exc}"}), 502

    # terminated flag + title from the series-info endpoint
    terminated_by, title_by = {}, {}
    for item in (info_raw or []):
        if item.get("status") == "SUCCESS":
            o = item["object"]
            vid = o.get("vectorId")
            if vid is not None:
                terminated_by[int(vid)] = _is_terminated(o.get("terminated"))
                title_by[int(vid)] = o.get("SeriesTitleEn", "")

    # empty? + latest reference period from the data endpoint (latestN=1)
    has_data_by, last_ref_by = {}, {}
    for item in (data_raw or []):
        if item.get("status") == "SUCCESS":
            o = item["object"]
            vid = o.get("vectorId")
            dps = [dp for dp in o.get("vectorDataPoint", [])
                   if dp.get("value") is not None and dp.get("statusCode") not in (1, 8, 9)]
            if vid is not None:
                has_data_by[int(vid)] = len(dps) > 0
                if dps:
                    last_ref_by[int(vid)] = max(dp.get("refPer", "")[:10] for dp in dps)

    results = []
    for v in vector_ids:
        vid = int(v)
        # If a fetch failed for a vector, default to "healthy" so we never
        # false-flag a working series on a transient API hiccup.
        results.append({
            "vectorId":   vid,
            "terminated": terminated_by.get(vid, False),
            "empty":      not has_data_by.get(vid, True),
            "lastRefPer": last_ref_by.get(vid),
            "title":      title_by.get(vid, ""),
        })

    return jsonify({"health": results})


# ---------------------------------------------------------------------------
# Route: GET /api/table-metadata
# Query params:
#   pid – product/table ID, e.g. "36100104"  (digits only, no dashes)
# ---------------------------------------------------------------------------
@app.route("/api/table-metadata")
def get_table_metadata():
    pid_raw = request.args.get("pid", "")
    if not pid_raw:
        return jsonify({"error": "No pid specified"}), 400

    pid = re.sub(r"\D", "", pid_raw)   # strip dashes/spaces
    payload = [{"productId": int(pid)}]

    try:
        resp = requests.post(
            f"{STATCAN_BASE}/getCubeMetadata",
            json=payload,
            timeout=20,
            headers={"Content-Type": "application/json"},
        )
        resp.raise_for_status()
    except requests.exceptions.RequestException as exc:
        return jsonify({"error": f"StatCan API error: {exc}"}), 502

    raw = resp.json()
    if not raw or raw[0].get("status") != "SUCCESS":
        return jsonify({"error": "Table not found or StatCan error"}), 404

    obj = raw[0]["object"]
    return jsonify({
        "productId":    obj.get("productId"),
        "cansimId":     obj.get("cansimId"),
        "titleEn":      obj.get("cubeTitleEn"),
        "titleFr":      obj.get("cubeTitleFr"),
        "startDate":    obj.get("cubeStartDate"),
        "endDate":      obj.get("cubeEndDate"),
        "frequency":    FREQ_LABEL.get(obj.get("frequencyCode"), "Unknown"),
        "frequencyCode": obj.get("frequencyCode"),
        "releaseTime":  obj.get("releaseTime"),
        "dimensions":   obj.get("dimension", []),
    })


# ---------------------------------------------------------------------------
# Route: GET /api/boc
# Query params:
#   series   – comma-separated BoC V-codes, e.g. "V39079,V39078"
#   fromDate – ISO date string YYYY-MM-DD
#   toDate   – ISO date string YYYY-MM-DD
#
# Fetches from the Bank of Canada Valet API and converts daily observations
# to monthly (last value per calendar month) so the frontend treats BoC
# series identically to StatCan monthly series.
# ---------------------------------------------------------------------------
@app.route("/api/boc")
def get_boc_series():
    raw_series = request.args.get("series",   "")
    from_date  = request.args.get("fromDate", "")
    to_date    = request.args.get("toDate",   "")

    if not raw_series:
        return jsonify({"error": "No series specified"}), 400

    series_codes = [s.strip() for s in raw_series.split(",") if s.strip()]
    results = []

    for code in series_codes:
        # The frontend strips a leading v/V via replace(/^[vV]/, ''), turning
        # "V39079" → "39079".  Re-add the prefix only for pure numeric codes.
        # Named codes (FXUSDCAD, BD.CDN.10YR.DQ.YLD, W.BCPI, etc.) are passed
        # through unchanged because they contain non-digit characters.
        boc_code = f"V{code}" if code.isdigit() else code

        params = {}
        if from_date:
            params["start_date"] = from_date
        if to_date:
            params["end_date"] = to_date

        try:
            resp = requests.get(
                f"{BOC_BASE}/observations/{boc_code}/json",
                params=params,
                timeout=25,
            )
            resp.raise_for_status()
        except requests.exceptions.RequestException as exc:
            results.append({"vectorId": code, "error": str(exc)})
            continue

        data         = resp.json()
        observations = data.get("observations", [])

        # Convert to monthly: keep the last daily value for each calendar month
        monthly = {}   # "YYYY-MM" → {"date": "YYYY-MM-DD", "value": float}
        for obs in observations:
            d = obs.get("d", "")
            if not d:
                continue
            val_obj = obs.get(boc_code, {})   # BoC payload key uses V-prefix
            val = val_obj.get("v")
            if val is None or val == "":
                continue
            try:
                val_f = float(val)
            except (ValueError, TypeError):
                continue
            month_key = d[:7]                           # "YYYY-MM"
            # Later dates overwrite earlier ones → end-of-month wins
            if month_key not in monthly or d > monthly[month_key]["date"]:
                monthly[month_key] = {"date": d, "value": val_f}

        data_points = [
            {"label": mk, "date": monthly[mk]["date"], "value": monthly[mk]["value"]}
            for mk in sorted(monthly)
        ]

        results.append({
            "vectorId":      code,
            "frequency":     "Monthly",
            "frequencyCode": 6,
            "uom":           "percent",
            "data":          data_points,
        })

    return jsonify({"series": results})


# ---------------------------------------------------------------------------
# Route: GET /api/fred
# Query params:
#   series   – comma-separated FRED series IDs, e.g. "UNRATE,PAYEMS"
#   fromDate – ISO date string YYYY-MM-DD
#   toDate   – ISO date string YYYY-MM-DD
# ---------------------------------------------------------------------------
@app.route("/api/fred")
def get_fred():
    series_raw = request.args.get("series", "")
    from_date  = request.args.get("fromDate", "")
    to_date    = request.args.get("toDate",   "")

    api_key = os.environ.get("FRED_API_KEY", "")
    if not api_key:
        return jsonify({"error": "FRED_API_KEY not configured on server"}), 503

    ids = [s.strip() for s in series_raw.split(",") if s.strip()]
    if not ids:
        return jsonify({"error": "No series specified"}), 400

    def _fred_label(date_str, freq):
        d = datetime.strptime(date_str, "%Y-%m-%d")
        if freq == "quarterly":
            q = (d.month - 1) // 3 + 1
            return f"{d.year} Q{q}"
        return f"{d.year}-{d.month:02d}"

    results = []
    for sid in ids:
        params = {"series_id": sid, "api_key": api_key, "file_type": "json"}
        if from_date:
            params["observation_start"] = from_date
        if to_date:
            params["observation_end"] = to_date
        try:
            r = requests.get(
                "https://api.stlouisfed.org/fred/series/observations",
                params=params, timeout=30,
            )
            r.raise_for_status()
            data = r.json()
        except requests.exceptions.Timeout:
            return jsonify({"error": f"FRED timed out fetching {sid}"}), 504
        except requests.exceptions.RequestException as exc:
            return jsonify({"error": f"FRED API error: {exc}"}), 502

        observations = [
            obs for obs in data.get("observations", [])
            if obs.get("value") not in (".", None, "")
        ]

        # Detect frequency from date gap between first two points
        freq = "monthly"
        if len(observations) >= 2:
            d1 = datetime.strptime(observations[0]["date"], "%Y-%m-%d")
            d2 = datetime.strptime(observations[1]["date"], "%Y-%m-%d")
            if (d2.year - d1.year) * 12 + (d2.month - d1.month) >= 3:
                freq = "quarterly"

        results.append({
            "vectorId":      sid,
            "frequency":     "Quarterly" if freq == "quarterly" else "Monthly",
            "frequencyCode": 9 if freq == "quarterly" else 12,
            "uom":           "",
            "data": [
                {"date": obs["date"], "label": _fred_label(obs["date"], freq), "value": float(obs["value"])}
                for obs in observations
            ],
        })

    return jsonify({"series": results})


# ---------------------------------------------------------------------------
# CIMT bulk-CSV helpers
# StatCan removed the /t1/cimt/rest/getReport API; data is now distributed as
# yearly ZIP files on the Open Government Portal.  Each ZIP contains CSV files
# at different HS granularities.  We download, cache, and stream-filter them.
#
# Imports ZIP:  CIMT-CICM_Imp_{year}.zip
#   ODPFN022_*N.csv  HS2  cols: YearMonth, HS2, Country, Province, State, Value
#   ODPFN015_*N.csv  HS6  cols: YearMonth, HS6, Country, Province, State, Value, Qty, UOM
#
# Exports ZIP:  CIMT-CICM_Tot_Exp_{year}.zip
#   ODPFN019_*N.csv  HS6  cols: YearMonth, HS6, Country, State, Value, Qty, UOM
#   (no Province column in exports)
#
# Country IDs used by the frontend are numeric (from countriesF.js); the CSV
# uses ISO 2-letter codes.  id=1000 (World) means no country filter.
# ---------------------------------------------------------------------------

_country_code_map: dict[int, str] | None = None   # CIMT numeric id → ISO code
_cimt_zip_checked: dict[str, float] = {}          # fname → last freshness-check timestamp


def _get_country_code_map() -> dict[int, str]:
    global _country_code_map
    if _country_code_map is not None:
        return _country_code_map
    try:
        r = requests.get(f"{CIMT_REF_BASE}/countriesF.js", timeout=20)
        r.raise_for_status()
        pairs = re.findall(r'"id":\s*(\d+)[^}]*?"c_code":\s*"([^"]+)"', r.text, re.S)
        _country_code_map = {int(cid): code for cid, code in pairs}
    except Exception:
        _country_code_map = {}
    return _country_code_map


def _get_cimt_zip_path(flow: str, year: int) -> str:
    """Return local path to the cached ZIP, downloading only if StatCan has updated it.

    Uses HTTP If-Modified-Since so a full re-download only happens when the file
    actually changes (i.e. when StatCan publishes revised data).  A 1-hour
    in-memory cooldown avoids hitting StatCan on every request.
    """
    os.makedirs(CIMT_CACHE_DIR, exist_ok=True)
    fname = f"CIMT-CICM_{'Tot_Exp' if flow == '0' else 'Imp'}_{year}.zip"
    path  = os.path.join(CIMT_CACHE_DIR, fname)
    url   = f"{CIMT_ZIP_BASE}/{fname}"
    now   = time.time()

    # Skip network check entirely if we verified freshness within the last hour
    if os.path.exists(path) and (now - _cimt_zip_checked.get(fname, 0)) < 3600:
        return path

    headers = {}
    if os.path.exists(path):
        # Send file's mtime as If-Modified-Since; server returns 304 if unchanged
        headers["If-Modified-Since"] = formatdate(os.path.getmtime(path), usegmt=True)

    resp = requests.get(url, headers=headers, timeout=180, stream=True)

    if resp.status_code == 304:
        _cimt_zip_checked[fname] = now
        return path  # cached copy is still current — no download needed

    resp.raise_for_status()
    with open(path, "wb") as fh:
        for chunk in resp.iter_content(chunk_size=65536):
            fh.write(chunk)

    # Set the file's mtime to the server's Last-Modified so the next
    # If-Modified-Since comparison is accurate
    last_mod = resp.headers.get("Last-Modified")
    if last_mod:
        try:
            server_ts = parsedate_to_datetime(last_mod).timestamp()
            os.utime(path, (server_ts, server_ts))
        except Exception:
            pass

    _cimt_zip_checked[fname] = now
    return path


def _find_csv_name(zf: zipfile.ZipFile, flow: str, need_hs6: bool) -> str | None:
    """Return the ZIP member name of the appropriate CSV."""
    if flow == "0":       # exports: only HS6 available
        prefix = "ODPFN019"
    elif need_hs6:        # imports with HS4 filter → need HS6 detail
        prefix = "ODPFN015"
    else:                 # imports, all-HS query → use small HS2 file
        prefix = "ODPFN022"
    for name in zf.namelist():
        if name.endswith(".csv") and os.path.basename(name).startswith(prefix):
            return name
    return None


# ---------------------------------------------------------------------------
# Route: GET /api/cimt
# Query params:
#   flow       – 0=exports, 1=imports
#   province   – CIMT province ID (1=Canada, 10=NL, 11=PEI, 12=NS, 13=NB,
#                24=QC, 35=ON, 46=MB, 47=SK, 48=AB, 59=BC)
#   country    – CIMT country ID (1000=World, 9=US, etc.)
#   hs4        – HS4 heading code (0=all, or 4-digit code like "0101")
#   start_date – YYYY-MM-DD
#   end_date   – YYYY-MM-DD
# ---------------------------------------------------------------------------
@app.route("/api/cimt")
def get_cimt():
    flow       = request.args.get("flow",       "0")
    province   = request.args.get("province",   "1")
    country    = request.args.get("country",    "1000")
    hs4        = request.args.get("hs4",        "0")
    start_date = request.args.get("start_date", "2020-01-01")
    end_date   = request.args.get("end_date",   "2025-01-01")

    try:
        start_dt = datetime.strptime(start_date[:10], "%Y-%m-%d")
        end_dt   = datetime.strptime(end_date[:10],   "%Y-%m-%d")
    except ValueError:
        return jsonify({"error": "Invalid start_date or end_date"}), 400

    years    = list(range(start_dt.year, end_dt.year + 1))
    start_ym = start_dt.strftime("%Y%m")   # "202001"
    end_ym   = end_dt.strftime("%Y%m")     # "202412"

    # Resolve province filter: None means no filter (all of Canada)
    prov_code = PROVINCE_CODES.get(str(province))

    # Resolve country filter: id=1000 (World) means no filter
    iso_code: str | None = None
    if country != "1000":
        iso_code = _get_country_code_map().get(int(country))

    # Exports have no Province column, so always use HS6; imports use HS2 when
    # hs4="0" (all commodities) to avoid streaming the larger HS6 file.
    need_hs6 = (hs4 != "0") or (flow == "0")

    def _process_year(year: int) -> dict[str, float]:
        by_month: dict[str, float] = {}
        try:
            zip_path = _get_cimt_zip_path(flow, year)
        except requests.exceptions.HTTPError as exc:
            if exc.response is not None and exc.response.status_code == 404:
                return by_month   # year not yet published
            raise

        with zipfile.ZipFile(zip_path, "r") as zf:
            csv_name = _find_csv_name(zf, flow, need_hs6)
            if not csv_name:
                return by_month

            with zf.open(csv_name) as raw:
                reader = csv.reader(io.TextIOWrapper(raw, encoding="utf-8", errors="replace"))
                next(reader)  # skip header

                # Column layout differs between flows:
                #   imports HS2/HS6: YM(0), HS(1), Country(2), Province(3), State(4), Value(5)
                #   exports HS6:     YM(0), HS(1), Country(2), State(3),    Value(4)
                if flow == "0":
                    cty_col, prov_col, val_col = 2, None, 4
                else:
                    cty_col, prov_col, val_col = 2, 3, 5

                for row in reader:
                    if len(row) <= val_col:
                        continue
                    ym = row[0].strip()
                    if ym < start_ym or ym > end_ym:
                        continue
                    if iso_code and row[cty_col].strip() != iso_code:
                        continue
                    if prov_col is not None and prov_code and row[prov_col].strip() != prov_code:
                        continue
                    if hs4 != "0" and not row[1].strip().startswith(hs4):
                        continue
                    try:
                        val = float(row[val_col])
                    except (ValueError, IndexError):
                        continue
                    by_month[ym] = by_month.get(ym, 0.0) + val

        return by_month

    by_date: dict[str, float] = {}
    with ThreadPoolExecutor(max_workers=min(4, len(years))) as pool:
        futures = {pool.submit(_process_year, y): y for y in years}
        for fut in as_completed(futures):
            try:
                for ym, val in fut.result().items():
                    by_date[ym] = by_date.get(ym, 0.0) + val
            except Exception:
                pass

    if not by_date:
        return jsonify({"error": "No data returned for this selection"}), 404

    series = [
        {"label": f"{ym[:4]}-{ym[4:6]}", "date": f"{ym[:4]}-{ym[4:6]}-01", "value": v}
        for ym, v in sorted(by_date.items())
    ]
    return jsonify({
        "series":        series,
        "uom":           "dollars",
        "frequency":     "Monthly",
        "frequencyCode": 6,
    })


# ---------------------------------------------------------------------------
# Route: GET /api/cimt-ref
# Returns HS4 headings and country list from the CIMT web app reference files.
# Results are cached for the lifetime of the server process.
# ---------------------------------------------------------------------------
_cimt_ref_cache: dict | None = None

@app.route("/api/cimt-ref")
def get_cimt_ref():
    global _cimt_ref_cache, _country_code_map
    if _cimt_ref_cache is not None:
        return jsonify(_cimt_ref_cache)

    try:
        hs4_resp = requests.get(f"{CIMT_REF_BASE}/hs4F.js",       timeout=20)
        cnt_resp = requests.get(f"{CIMT_REF_BASE}/countriesF.js",  timeout=20)
        hs4_resp.raise_for_status()
        cnt_resp.raise_for_status()
    except requests.exceptions.RequestException as exc:
        return jsonify({"error": f"Failed to fetch CIMT reference data: {exc}"}), 502

    # hs4F.js format: {"HS": "0101", "EN": "Live horses...", "FR": "..."}
    hs4_pairs = re.findall(r'"HS":\s*"(\d+)"[^}]*?"EN":\s*"([^"]+)"', hs4_resp.text, re.S)
    hs4_list  = [{"hs": h, "en": e} for h, e in hs4_pairs]

    # countriesF.js: extract id, c_code (ISO), and English name
    # Also populate _country_code_map so /api/cimt doesn't need a separate fetch.
    cnt_entries = re.findall(
        r'"id":\s*(\d+)[^}]*?"c_code":\s*"([^"]+)"[^}]*?"en":\s*"([^"]+)"',
        cnt_resp.text, re.S,
    )
    seen_ids: set[int] = set()
    country_list = []
    code_map: dict[int, str] = {}
    for raw_id, c_code, en in cnt_entries:
        cid = int(raw_id)
        code_map[cid] = c_code
        if cid == 1 or cid in seen_ids:
            continue
        seen_ids.add(cid)
        country_list.append({"id": cid, "en": en})
    country_list.sort(key=lambda x: x["en"])
    _country_code_map = code_map   # cache for /api/cimt use

    _cimt_ref_cache = {"hs4": hs4_list, "countries": country_list}
    return jsonify(_cimt_ref_cache)


# ═══════════════════════════════════════════════════════════════════════════
# SERIES WIZARD — admin tool for adding categories/series to Vectors.xlsx
# ═══════════════════════════════════════════════════════════════════════════

VECTORS_PATH    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Vectors.xlsx")
BACKUP_DIR      = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backups")
CUBES_CACHE     = "/tmp/statcan_cubes_lite.json"
CUBES_TTL       = 24 * 3600          # refresh table list daily
GITHUB_REPO     = os.environ.get("GITHUB_REPO", "jmkyyz/statcan-explorer")

CATALOG_COLUMNS = [
    "category", "freq", "series_id", "series_name", "table_id",
    "dim1_name", "dim1_value", "dim2_name", "dim2_value",
    "dim3_name", "dim3_value", "dim4_name", "dim4_value",
    "dim5_name", "dim5_value", "vector", "full_label", "short_label",
    "dim1_group",
]

_cubes_cache_mem: list | None = None
_cubes_cache_ts: float = 0


@app.route("/wizard")
def wizard():
    here = os.path.dirname(os.path.abspath(__file__))
    return send_from_directory(here, "wizard.html")


# ---------------------------------------------------------------------------
# Route: GET /api/wizard-info
# Tells the wizard frontend whether an admin key is required and whether
# GitHub persistence is configured on this server.
# ---------------------------------------------------------------------------
@app.route("/api/wizard-info")
def wizard_info():
    return jsonify({
        "adminRequired":    _admin_required(),
        "keyConfigured":    bool(os.environ.get("ADMIN_KEY")),
        "githubConfigured": bool(os.environ.get("GITHUB_TOKEN")),
        "repo":             GITHUB_REPO if os.environ.get("GITHUB_TOKEN") else None,
    })


# ---------------------------------------------------------------------------
# Route: POST /api/verify-key
# Lets the wizard's lock screen validate a key before unlocking the UI.
# ---------------------------------------------------------------------------
@app.route("/api/verify-key", methods=["POST"])
def verify_key():
    return jsonify({"ok": _check_admin()})


def _admin_required() -> bool:
    """Admin auth is in force when a key is configured, or always in
    production (Render sets the RENDER env var) as a fail-safe so the
    catalog can never be modified before ADMIN_KEY is configured."""
    return bool(os.environ.get("ADMIN_KEY")) or bool(os.environ.get("RENDER"))


def _check_admin() -> bool:
    """True if the request is authorized to use protected wizard endpoints."""
    if not _admin_required():
        return True   # local dev, no key configured → open
    admin_key = os.environ.get("ADMIN_KEY", "")
    if not admin_key:
        return False  # production without a configured key → locked
    import hmac
    return hmac.compare_digest(request.headers.get("X-Admin-Key", ""), admin_key)


# ---------------------------------------------------------------------------
# Route: GET /api/cubes-list
# Slimmed list of every StatCan table (English), cached for 24h.
# ---------------------------------------------------------------------------
@app.route("/api/cubes-list")
def cubes_list():
    global _cubes_cache_mem, _cubes_cache_ts
    now = time.time()

    if _cubes_cache_mem is not None and now - _cubes_cache_ts < CUBES_TTL:
        return jsonify({"cubes": _cubes_cache_mem})

    # Try the on-disk cache before hitting StatCan
    if os.path.exists(CUBES_CACHE) and now - os.path.getmtime(CUBES_CACHE) < CUBES_TTL:
        try:
            with open(CUBES_CACHE) as fh:
                _cubes_cache_mem = json.load(fh)
            _cubes_cache_ts = now
            return jsonify({"cubes": _cubes_cache_mem})
        except Exception:
            pass

    try:
        r = requests.get(f"{STATCAN_BASE}/getAllCubesListLite", timeout=90)
        r.raise_for_status()
        raw = r.json()
    except requests.exceptions.RequestException as exc:
        # Fall back to a stale disk cache if the fetch fails
        if os.path.exists(CUBES_CACHE):
            with open(CUBES_CACHE) as fh:
                return jsonify({"cubes": json.load(fh), "stale": True})
        return jsonify({"error": f"StatCan API error: {exc}"}), 502

    slim = [
        {
            "pid":      c.get("productId"),
            "title":    c.get("cubeTitleEn", ""),
            "start":    (c.get("cubeStartDate") or "")[:7],
            "end":      (c.get("cubeEndDate") or "")[:7],
            "freq":     FREQ_LABEL.get(c.get("frequencyCode"), ""),
            "archived": str(c.get("archived", "")) == "1",
        }
        for c in raw
        if c.get("productId") and c.get("cubeTitleEn")
    ]
    _cubes_cache_mem = slim
    _cubes_cache_ts  = now
    try:
        with open(CUBES_CACHE, "w") as fh:
            json.dump(slim, fh)
    except Exception:
        pass
    return jsonify({"cubes": slim})


# ---------------------------------------------------------------------------
# Route: POST /api/resolve-vectors
# Body: {"productId": 18100004, "coordinates": ["2.2.0.0.0.0.0.0.0.0", ...]}
# Resolves dimension-member coordinates to vector IDs via WDS
# getSeriesInfoFromCubePidCoord, in batches.  Combinations that don't exist
# in the cube come back with vectorId null so the frontend can skip them.
# ---------------------------------------------------------------------------
@app.route("/api/resolve-vectors", methods=["POST"])
def resolve_vectors():
    if not _check_admin():
        return jsonify({"error": "Invalid or missing admin key"}), 401

    body = request.get_json(force=True, silent=True) or {}
    pid    = body.get("productId")
    coords = body.get("coordinates", [])

    if not pid or not isinstance(coords, list) or not coords:
        return jsonify({"error": "productId and coordinates are required"}), 400
    if len(coords) > 12000:
        return jsonify({"error": "Too many coordinates in one request (max 12000)"}), 400

    BATCH = 100

    def _resolve_batch(batch: list[str]) -> list[dict]:
        payload = [{"productId": int(pid), "coordinate": c} for c in batch]
        r = requests.post(
            f"{STATCAN_BASE}/getSeriesInfoFromCubePidCoord",
            json=payload, timeout=60,
            headers={"Content-Type": "application/json"},
        )
        r.raise_for_status()
        out = []
        for item in r.json():
            obj = item.get("object", {}) if isinstance(item, dict) else {}
            ok  = (item.get("status") == "SUCCESS"
                   and obj.get("responseStatusCode") == 0
                   and obj.get("vectorId"))
            out.append({
                "coordinate": obj.get("coordinate"),
                "vectorId":   int(obj["vectorId"]) if ok else None,
                "title":      obj.get("SeriesTitleEn") or "",
                "terminated": obj.get("terminated") or 0,
            })
        return out

    batches = [coords[i:i + BATCH] for i in range(0, len(coords), BATCH)]
    results_by_coord: dict[str, dict] = {}
    errors = 0

    with ThreadPoolExecutor(max_workers=4) as pool:
        futures = [pool.submit(_resolve_batch, b) for b in batches]
        for fut in as_completed(futures):
            try:
                for res in fut.result():
                    if res.get("coordinate"):
                        results_by_coord[res["coordinate"]] = res
            except Exception:
                errors += 1

    # Preserve request order; coordinates StatCan never echoed back count as misses
    results = [
        results_by_coord.get(c, {"coordinate": c, "vectorId": None, "title": "", "terminated": 0})
        for c in coords
    ]
    resp = {"results": results}
    if errors:
        resp["warning"] = f"{errors} batch(es) failed — some vectors may be missing"
    return jsonify(resp)


# ---------------------------------------------------------------------------
# Catalog helpers
# ---------------------------------------------------------------------------
# Parsed-catalog cache (mtime, header, rows). Reading the 50-80k-row workbook
# with openpyxl takes ~7s, and the summary / dedup / catalog-rows / reorder
# paths all need it — so we parse once and reuse until the file changes.
# Mutations refresh this in _save_catalog from the rows they just wrote, so a
# wizard add/delete never re-parses the file (only the unavoidable ~5s write).
_parsed_catalog = None   # (mtime, header, rows)


def _read_catalog():
    """Parse Vectors.xlsx → (header, list-of-row-dicts), cached by file mtime."""
    global _parsed_catalog
    import openpyxl
    try:
        mtime = os.path.getmtime(VECTORS_PATH)
    except OSError:
        mtime = None
    if _parsed_catalog is not None and _parsed_catalog[0] == mtime:
        return _parsed_catalog[1], _parsed_catalog[2]

    wb = openpyxl.load_workbook(VECTORS_PATH, read_only=True)
    ws = wb["series"] if "series" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h or "").strip() for h in next(rows_iter)]
    rows = []
    for r in rows_iter:
        d = {header[i]: ("" if r[i] is None else str(r[i]).strip())
             for i in range(min(len(header), len(r)))}
        if d.get("series_id"):
            rows.append(d)
    wb.close()
    # Hold the ~273MB parse cache only off-Render (local dev has RAM to spare and
    # benefits from the speed). On Render's 512MB instance, writes are blocked and
    # serving uses the small gz cache, so we don't keep the big list resident.
    if not os.environ.get("RENDER"):
        _parsed_catalog = (mtime, header, rows)
    return header, rows


def _dim_key(row: dict) -> str:
    return "__".join(row.get(f"dim{i}_value", "") for i in range(1, 6))


# ---------------------------------------------------------------------------
# Route: GET /api/catalog
#   (no params)        → summary: categories + series list
#   ?series_id=cpi_nsa → that series' rows (for the wizard's extend mode)
# ---------------------------------------------------------------------------
@app.route("/api/catalog")
def get_catalog():
    try:
        _, rows = _read_catalog()
    except Exception as exc:
        return jsonify({"error": f"Could not read Vectors.xlsx: {exc}"}), 500

    series_id = request.args.get("series_id", "").strip()
    if series_id:
        s_rows = [r for r in rows if r["series_id"] == series_id]
        if not s_rows:
            return jsonify({"error": f"series_id '{series_id}' not found"}), 404
        first = s_rows[0]
        return jsonify({
            "seriesId":   series_id,
            "seriesName": first.get("series_name", ""),
            "category":   first.get("category", ""),
            "tableId":    first.get("table_id", ""),
            "freq":       first.get("freq", "M"),
            "dimNames":   [first.get(f"dim{i}_name", "") for i in range(1, 6)],
            "rows": [
                {
                    "dimValues": [r.get(f"dim{i}_value", "") for i in range(1, 6)],
                    "vector":     r.get("vector", ""),
                    "fullLabel":  r.get("full_label", ""),
                    "shortLabel": r.get("short_label", ""),
                    "dim1Group":  r.get("dim1_group", ""),
                }
                for r in s_rows
            ],
        })

    categories: list[str] = []
    series: dict[str, dict] = {}
    for r in rows:
        cat = r.get("category", "")
        if cat and cat not in categories:
            categories.append(cat)
        sid = r["series_id"]
        if sid not in series:
            series[sid] = {
                "seriesId":   sid,
                "seriesName": r.get("series_name", ""),
                "category":   cat,
                "tableId":    r.get("table_id", ""),
                "freq":       r.get("freq", "M"),
                "dimNames":   [n for n in (r.get(f"dim{i}_name", "") for i in range(1, 6)) if n],
                "rowCount":   0,
            }
        series[sid]["rowCount"] += 1

    return jsonify({"categories": categories, "series": list(series.values())})


# ---------------------------------------------------------------------------
# Route: GET /api/catalog-rows
# Full catalog as raw rows keyed by column name — the frontend's fast load
# path. Lets the browser skip ~10s of in-browser XLSX decoding on large
# catalogs (JSON.parse is ~100x faster than SheetJS). Cached in memory and
# rebuilt only when Vectors.xlsx changes on disk; gzipped when the client
# accepts it (multi-MB JSON → ~1MB on the wire).
# ---------------------------------------------------------------------------
_catalog_rows_cache = None   # (mtime, gz_bytes) — gz only; raw JSON (~70MB at 100k+
                             # rows) is decompressed on demand for the rare non-gzip client


def _build_catalog_cache():
    """(Re)build the gzipped-JSON catalog cache from Vectors.xlsx if stale.
    Returns (mtime, gz_bytes), or None on failure. Pre-warmed at startup so the
    first visitor after a deploy doesn't pay the spreadsheet parse."""
    global _catalog_rows_cache
    try:
        mtime = os.path.getmtime(VECTORS_PATH)
        if _catalog_rows_cache is not None and _catalog_rows_cache[0] == mtime:
            return _catalog_rows_cache
        _, rows = _read_catalog()
        raw = json.dumps({"rows": rows}, separators=(",", ":")).encode("utf-8")
        gz = gzip.compress(raw, 6)
        del raw                       # don't keep the ~70MB raw JSON resident
        _catalog_rows_cache = (mtime, gz)
        return _catalog_rows_cache
    except Exception:
        return None


@app.route("/api/catalog-rows")
def get_catalog_rows():
    cache = _build_catalog_cache()
    if cache is None:
        return jsonify({"error": "Could not read Vectors.xlsx"}), 500

    _, gz = cache
    use_gzip = "gzip" in request.headers.get("Accept-Encoding", "")
    body = gz if use_gzip else gzip.decompress(gz)   # browsers accept gzip → no decompress
    resp = app.response_class(body, mimetype="application/json")
    if use_gzip:
        resp.headers["Content-Encoding"] = "gzip"
        resp.headers["Vary"] = "Accept-Encoding"
    resp.headers["Content-Length"] = str(len(body))
    return resp


# ---------------------------------------------------------------------------
# GitHub persistence: commit the updated Vectors.xlsx so Render redeploys
# and the change survives ephemeral-disk restarts.
# ---------------------------------------------------------------------------
def _github_commit_vectors(message: str) -> str | None:
    token = os.environ.get("GITHUB_TOKEN", "")
    if not token:
        return None
    api = f"https://api.github.com/repos/{GITHUB_REPO}/contents/Vectors.xlsx"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
    }
    sha = None
    r = requests.get(api, headers=headers, params={"ref": "main"}, timeout=30)
    if r.status_code == 200:
        sha = r.json().get("sha")
    with open(VECTORS_PATH, "rb") as fh:
        content = base64.b64encode(fh.read()).decode()
    payload = {"message": message, "content": content, "branch": "main"}
    if sha:
        payload["sha"] = sha
    r = requests.put(api, headers=headers, json=payload, timeout=60)
    r.raise_for_status()
    return r.json().get("commit", {}).get("sha")


# Safeguard: block catalog-modifying writes on the LIVE server once the catalog
# is large. Rewriting a big workbook risks the 512MB memory limit, and editing
# live also fragments the catalog vs. local work (causing divergence). Big
# rebuilds belong on the local wizard. Local dev has no RENDER env → never blocked.
LIVE_WRITE_MAX_BYTES = 8_000_000   # ~8MB ≈ ~85k rows


def _live_write_blocked():
    if not os.environ.get("RENDER"):
        return None
    try:
        if os.path.getsize(VECTORS_PATH) <= LIVE_WRITE_MAX_BYTES:
            return None
    except OSError:
        return None
    return jsonify({"error": "This catalog is too large to edit on the live site. "
                    "Rebuild on the local wizard (localhost:5001/wizard) and push — "
                    "editing the live site risks crashing it and diverging from your "
                    "local copy."}), 403


# ---------------------------------------------------------------------------
# Route: POST /api/catalog/append
# Headers: X-Admin-Key (required when ADMIN_KEY env var is set)
# Body: {"rows": [{category, freq, series_id, ... vector, full_label, ...}],
#        "dryRun": false}
# Appends new rows to Vectors.xlsx (deduplicated against existing rows),
# after writing a timestamped backup.  Commits to GitHub when configured.
# ---------------------------------------------------------------------------
@app.route("/api/catalog/append", methods=["POST"])
def catalog_append():
    if not _check_admin():
        return jsonify({"error": "Invalid or missing admin key"}), 401
    blocked = _live_write_blocked()
    if blocked:
        return blocked

    body = request.get_json(force=True, silent=True) or {}
    rows = body.get("rows", [])
    dry  = bool(body.get("dryRun"))

    if not isinstance(rows, list) or not rows:
        return jsonify({"error": "No rows provided"}), 400
    if len(rows) > 12000:
        return jsonify({"error": "Too many rows in one request (max 12000)"}), 400

    for i, r in enumerate(rows):
        for req_col in ("category", "freq", "series_id", "vector"):
            if not str(r.get(req_col, "")).strip():
                return jsonify({"error": f"Row {i + 1} is missing '{req_col}'"}), 400

    # Read existing rows (read_only, streaming ~50MB) to dedup against.
    try:
        header, existing = _read_catalog()
    except Exception as exc:
        return jsonify({"error": f"Could not read Vectors.xlsx: {exc}"}), 500

    existing_vec: set[tuple] = set()
    existing_dim: set[tuple] = set()
    for d in existing:
        sid = d.get("series_id", "")
        existing_vec.add((sid, d.get("vector", "").lstrip("vV")))
        existing_dim.add((sid, _dim_key(d)))

    new_rows, skipped = [], 0
    for r in rows:
        sid = str(r["series_id"]).strip()
        vec = str(r["vector"]).strip().lstrip("vV")
        if (sid, vec) in existing_vec or (sid, _dim_key(r)) in existing_dim:
            skipped += 1
            continue
        existing_vec.add((sid, vec))
        existing_dim.add((sid, _dim_key(r)))
        new_rows.append(r)

    result = {"added": len(new_rows), "skipped": skipped, "dryRun": dry}
    if dry or not new_rows:
        return jsonify(result)

    # Extend the header with any new dim*_level columns the incoming rows carry.
    for lvl_col in (f"dim{i}_level" for i in range(1, 6)):
        if lvl_col not in header and any(str(r.get(lvl_col, "")).strip() for r in new_rows):
            header.append(lvl_col)

    sids = sorted({str(r["series_id"]) for r in new_rows})
    msg = f"Wizard: add {len(new_rows)} rows to {', '.join(sids[:5])}"
    if len(sids) > 5:
        msg += f" (+{len(sids) - 5} more)"

    save = _save_catalog(header, existing + new_rows, msg)
    result["backup"] = save.get("backup")
    for k in ("committed", "commitSha", "commitError"):
        if k in save:
            result[k] = save[k]
    return jsonify(result)


# ---------------------------------------------------------------------------
# Route: POST /api/catalog/delete
# Headers: X-Admin-Key (required when ADMIN_KEY env var is set)
# Body: {"seriesId": "cpi_nsa"}  or  {"category": "Prices"}
# Removes all matching rows from Vectors.xlsx after a timestamped backup.
# Commits to GitHub when configured.
# ---------------------------------------------------------------------------
@app.route("/api/catalog/delete", methods=["POST"])
def catalog_delete():
    if not _check_admin():
        return jsonify({"error": "Invalid or missing admin key"}), 401
    blocked = _live_write_blocked()
    if blocked:
        return blocked

    body      = request.get_json(force=True, silent=True) or {}
    series_id = str(body.get("seriesId", "") or "").strip()
    category  = str(body.get("category", "") or "").strip()

    if bool(series_id) == bool(category):
        return jsonify({"error": "Provide exactly one of seriesId or category"}), 400

    # Stream the catalog (read_only ~50MB) and keep everything that doesn't match.
    try:
        header, rows = _read_catalog()
    except Exception as exc:
        return jsonify({"error": f"Could not read Vectors.xlsx: {exc}"}), 500

    key = "series_id" if series_id else "category"
    target_val = series_id or category
    kept = [r for r in rows if r.get(key, "") != target_val]
    deleted = len(rows) - len(kept)
    if not deleted:
        return jsonify({"error": f"No rows found for {'series' if series_id else 'category'} '{target_val}'"}), 404

    what = f"series {series_id}" if series_id else f"category {category}"
    result = _save_catalog(header, kept, f"Wizard: delete {what} ({deleted} rows)")
    result["deleted"] = deleted
    return jsonify(result)


# ---------------------------------------------------------------------------
# Shared writer for every catalog mutation (append / delete / reorder / rename).
# Rewrites Vectors.xlsx from a list of row-dicts after a timestamped backup,
# invalidates the /api/catalog-rows cache, and commits to GitHub when
# configured.
#
# CRITICAL: uses openpyxl write_only mode, which streams rows to disk instead
# of holding the whole sheet in memory. A normal/writable workbook of the 50k+
# row catalog peaks at ~770MB and OOM-kills Render's 512MB instance; write_only
# (paired with read_only reads everywhere else) keeps it well under ~150MB.
# ---------------------------------------------------------------------------
def _save_catalog(header, rows, commit_msg):
    global _catalog_rows_cache, _parsed_catalog
    import openpyxl
    os.makedirs(BACKUP_DIR, exist_ok=True)
    backup_name = f"Vectors-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    shutil.copy2(VECTORS_PATH, os.path.join(BACKUP_DIR, backup_name))

    n = 0
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("series")
    ws.append(header)
    col_of = {name: i for i, name in enumerate(header)}
    vec_idx = col_of.get("vector")
    for r in rows:
        out = [""] * len(header)
        for col, idx in col_of.items():
            out[idx] = str(r.get(col, "") or "")
        if vec_idx is not None:
            out[vec_idx] = out[vec_idx].lstrip("vV")
        ws.append(out)
        n += 1
    wb.save(VECTORS_PATH)
    wb.close()

    # Invalidate caches (don't rebuild a second in-memory copy here). At 100k+
    # rows, building a normalized cache_rows list during the write doubled peak
    # memory and OOM-killed Render's 512MB instance. The next read re-parses
    # once (cached thereafter); a wizard write never holds two full copies.
    _parsed_catalog = None
    _catalog_rows_cache = None

    result = {"rows": n, "backup": f"backups/{backup_name}"}
    if os.environ.get("GITHUB_TOKEN"):
        try:
            result["commitSha"] = _github_commit_vectors(commit_msg)
            result["committed"] = True
        except Exception as exc:
            result["committed"] = False
            result["commitError"] = str(exc)
    return result


# ---------------------------------------------------------------------------
# Route: POST /api/catalog/reorder
# Body: {"seriesOrder": ["sid1", "sid2", ...]}  — the full desired order of
# series_ids. Category order falls out of where each category's series land
# (categories stay contiguous blocks). Each series' rows keep their internal
# order; any series_id omitted from the list is appended in its original spot.
# ---------------------------------------------------------------------------
@app.route("/api/catalog/reorder", methods=["POST"])
def catalog_reorder():
    if not _check_admin():
        return jsonify({"error": "Invalid or missing admin key"}), 401
    blocked = _live_write_blocked()
    if blocked:
        return blocked

    body = request.get_json(force=True, silent=True) or {}
    order = body.get("seriesOrder")
    if not isinstance(order, list) or not order:
        return jsonify({"error": "seriesOrder (list of series_id) is required"}), 400

    header, rows = _read_catalog()
    by_sid: dict[str, list] = {}
    original: list[str] = []
    for r in rows:
        sid = r.get("series_id", "")
        if sid not in by_sid:
            by_sid[sid] = []
            original.append(sid)
        by_sid[sid].append(r)

    new_rows, used = [], set()
    for sid in order:
        if sid in by_sid and sid not in used:
            new_rows.extend(by_sid[sid])
            used.add(sid)
    leftover = 0
    for sid in original:          # safety: keep any series the client didn't list
        if sid not in used:
            new_rows.extend(by_sid[sid])
            leftover += 1

    result = _save_catalog(header, new_rows, "Wizard: reorder categories/series")
    result["leftover"] = leftover
    return jsonify(result)


# ---------------------------------------------------------------------------
# Route: POST /api/catalog/rename
# Body: {"type": "category", "oldName": "...", "newName": "..."}
#   or  {"type": "series",   "seriesId": "...", "newName": "..."}
# Renames in place — series_id stays stable so saved configs/bookmarks survive.
# ---------------------------------------------------------------------------
@app.route("/api/catalog/rename", methods=["POST"])
def catalog_rename():
    if not _check_admin():
        return jsonify({"error": "Invalid or missing admin key"}), 401
    blocked = _live_write_blocked()
    if blocked:
        return blocked

    body = request.get_json(force=True, silent=True) or {}
    kind = str(body.get("type", "")).strip()
    new_name = str(body.get("newName", "")).strip()
    if not new_name:
        return jsonify({"error": "newName is required"}), 400

    header, rows = _read_catalog()
    changed = 0
    if kind == "category":
        old = str(body.get("oldName", "")).strip()
        if not old:
            return jsonify({"error": "oldName is required"}), 400
        for r in rows:
            if r.get("category", "") == old:
                r["category"] = new_name
                changed += 1
        msg = f"Wizard: rename category '{old}' -> '{new_name}'"
    elif kind == "series":
        sid = str(body.get("seriesId", "")).strip()
        if not sid:
            return jsonify({"error": "seriesId is required"}), 400
        for r in rows:
            if r.get("series_id", "") == sid:
                r["series_name"] = new_name
                changed += 1
        msg = f"Wizard: rename series '{sid}' -> '{new_name}'"
    else:
        return jsonify({"error": "type must be 'category' or 'series'"}), 400

    if not changed:
        return jsonify({"error": "No matching rows found"}), 404

    result = _save_catalog(header, rows, msg)
    result["updated"] = changed
    return jsonify(result)


# ---------------------------------------------------------------------------
# Route: GET /api/health
# ---------------------------------------------------------------------------
@app.route("/api/health")
def health():
    return jsonify({"status": "ok", "statcan_base": STATCAN_BASE})


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import threading
    port = int(os.environ.get("PORT", 5001))
    # Network binding. Default 0.0.0.0 (all interfaces) so Render can route to it
    # and so localhost testing on a phone works. Set HOST=127.0.0.1 to restrict
    # the server to THIS machine only (e.g. on a shared office network) — then no
    # one else on the LAN can reach the open-mode wizard.
    host = os.environ.get("HOST", "0.0.0.0")
    reach = "this machine only" if host in ("127.0.0.1", "localhost") else "all network interfaces"
    print("=" * 60)
    print(f"  StatCan WDS Proxy  →  http://localhost:{port}")
    print(f"  bind: {host}  ({reach})")
    print("=" * 60)
    # Pre-warm the catalog cache in the background so the first request after a
    # deploy doesn't pay the spreadsheet-parse cost. Port still binds instantly.
    threading.Thread(target=_build_catalog_cache, daemon=True).start()
    app.run(host=host, port=port, debug=False)
