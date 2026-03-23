import openpyxl
from openpyxl.styles import PatternFill, Font

FILL_GREY = PatternFill(fill_type="solid", fgColor="FFEEEEEE")
FILL_WHITE = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
FONT = Font(name="Arial", size=10)

wb = openpyxl.load_workbook("/Users/jasonkirby/Desktop/StatCanApp/vectors-template.xlsx")
ws = wb.active

start_row = ws.max_row + 1

# category | freq | series_id | series_name | table_id | dim1_name | dim1_value | dim2_name | dim2_value | vector | full_label | short_label
# Columns: 1-12

rows_to_append = []

# ── CATEGORY: Prices (grey EEEEEE, monthly) ──────────────────────────────────
prices_10 = [
    ("prc_cpi_nsa", "CPI, NSA", "18-10-0004-01", "", "", "", "", "41690973", "Consumer Price Index, NSA", "CPI NSA"),
    ("prc_cpi_food_nsa", "CPI food, NSA", "18-10-0004-01", "", "", "", "", "41690974", "CPI: Food, NSA", "Food NSA"),
    ("prc_cpi_foodstores", "CPI food purchased from stores, NSA", "18-10-0004-01", "", "", "", "", "41690975", "CPI: Food purchased from stores, NSA", "Food from stores"),
    ("prc_cpi_exfood_nsa", "CPI excl. food, NSA", "18-10-0004-01", "", "", "", "", "41691232", "CPI excluding food, NSA", "Excl. food NSA"),
    ("prc_cpi_core_nsa", "CPI excl. food and energy, NSA", "18-10-0004-01", "", "", "", "", "41691233", "CPI excluding food and energy, NSA", "Core NSA"),
    ("prc_cpi_sa", "CPI, SA", "18-10-0004-01", "", "", "", "", "41690914", "Consumer Price Index, SA", "CPI SA"),
    ("prc_cpi_exfood_sa", "CPI excl. food, SA", "18-10-0004-01", "", "", "", "", "41690923", "CPI excluding food, SA", "Excl. food SA"),
    ("prc_cpi_core_sa", "CPI excl. food and energy, SA", "18-10-0004-01", "", "", "", "", "41690924", "CPI excluding food and energy, SA", "Core SA"),
    ("prc_cpi_median", "CPI core weighted median", "18-10-0004-01", "", "", "", "", "1481215115", "CPI: Weighted median", "Weighted median"),
    ("prc_cpi_trimmed", "CPI core trimmed mean", "18-10-0004-01", "", "", "", "", "1481215116", "CPI: Trimmed mean", "Trimmed mean"),
]
for sid, sname, tid, d1n, d1v, d2n, d2v, vec, fl, sl in prices_10:
    rows_to_append.append(("Prices", "monthly", sid, sname, tid, d1n, d1v, d2n, d2v, vec, fl, sl, FILL_GREY))

# ── CATEGORY: Sales (white FFFFFF, monthly) ───────────────────────────────────

# ret_nsa: 110 rows (11 geographies × 10 industries)
geo_list = [
    "Canada",
    "Newfoundland and Labrador",
    "Prince Edward Island",
    "Nova Scotia",
    "New Brunswick",
    "Quebec",
    "Ontario",
    "Manitoba",
    "Saskatchewan",
    "Alberta",
    "British Columbia",
]
industry_codes = [
    "Retail trade [44-45]",
    "Motor vehicle and parts dealers [441]",
    "Building material and garden equipment and supplies dealers [444]",
    "Food and beverage retailers [445]",
    "Furniture, home furnishings, electronics and appliances retailers [449]",
    "General merchandise retailers [455]",
    "Health and personal care retailers [456]",
    "Gasoline stations and fuel vendors [457]",
    "Clothing, clothing accessories, shoes, jewellery, luggage and leather goods retailers [458]",
    "Sporting goods, hobby, musical instrument, book, and miscellaneous retailers [459]",
]
industry_short = [
    "Retail trade",
    "Motor vehicles",
    "Building materials",
    "Food & beverage",
    "Furniture & electronics",
    "General merchandise",
    "Health & personal care",
    "Gasoline & fuel",
    "Clothing & accessories",
    "Sporting goods & misc.",
]
geo_vectors_nsa = {
    "Canada": [1446859481, 1446859485, 1446859497, 1446859499, 1446859511, 1446859521, 1446859523, 1446859525, 1446859527, 1446859535],
    "Newfoundland and Labrador": [1446859542, 1446859544, 1446859550, 1446859551, 1446859557, 1446859562, 1446859563, 1446859564, 1446859565, 1446859569],
    "Prince Edward Island": [1446859573, 1446859575, 1446859581, 1446859582, 1446859588, 1446859593, 1446859594, 1446859595, 1446859596, 1446859600],
    "Nova Scotia": [1446859604, 1446859606, 1446859612, 1446859613, 1446859619, 1446859624, 1446859625, 1446859626, 1446859627, 1446859631],
    "New Brunswick": [1446859635, 1446859637, 1446859643, 1446859644, 1446859650, 1446859655, 1446859656, 1446859657, 1446859658, 1446859662],
    "Quebec": [1446859666, 1446859668, 1446859674, 1446859675, 1446859681, 1446859686, 1446859687, 1446859688, 1446859689, 1446859693],
    "Ontario": [1446859788, 1446859790, 1446859796, 1446859797, 1446859803, 1446859808, 1446859809, 1446859810, 1446859811, 1446859815],
    "Manitoba": [1446859880, 1446859882, 1446859888, 1446859889, 1446859895, 1446859900, 1446859901, 1446859902, 1446859903, 1446859907],
    "Saskatchewan": [1446859941, 1446859943, 1446859949, 1446859950, 1446859956, 1446859961, 1446859962, 1446859963, 1446859964, 1446859968],
    "Alberta": [1446859972, 1446859974, 1446859980, 1446859981, 1446859987, 1446859992, 1446859993, 1446859994, 1446859995, 1446859999],
    "British Columbia": [1446860063, 1446860065, 1446860071, 1446860072, 1446860078, 1446860083, 1446860084, 1446860085, 1446860086, 1446860090],
}
for geo in geo_list:
    vecs = geo_vectors_nsa[geo]
    for i, (ind, sh, vec) in enumerate(zip(industry_codes, industry_short, vecs)):
        fl = f"{geo} - {ind}"
        rows_to_append.append(("Sales", "monthly", "ret_nsa", "Retail trade NSA", "20-10-0056-01", "Geography", geo, "Industry", ind, str(vec), fl, sh, FILL_WHITE))

# ret_sa: 20 rows
# Canada: 10 industries
canada_sa = [
    ("Retail trade [44-45]", 1446859483, "Retail trade"),
    ("Motor vehicle and parts dealers [441]", 1446859486, "Motor vehicles"),
    ("Building material and garden equipment and supplies dealers [444]", 1446859498, "Building materials"),
    ("Food and beverage retailers [445]", 1446859500, "Food & beverage"),
    ("Furniture, home furnishings, electronics and appliances retailers [449]", 1446859512, "Furniture & electronics"),
    ("General merchandise retailers [455]", 1446859522, "General merchandise"),
    ("Health and personal care retailers [456]", 1446859524, "Health & personal care"),
    ("Gasoline stations and fuel vendors [457]", 1446859526, "Gasoline & fuel"),
    ("Clothing, clothing accessories, shoes, jewellery, luggage and leather goods retailers [458]", 1446859528, "Clothing & accessories"),
    ("Sporting goods, hobby, musical instrument, book, and miscellaneous retailers [459]", 1446859536, "Sporting goods & misc."),
]
for ind, vec, sh in canada_sa:
    fl = f"Canada - {ind} SA"
    rows_to_append.append(("Sales", "monthly", "ret_sa", "Retail trade SA", "20-10-0056-01", "Geography", "Canada", "Industry", ind, str(vec), fl, sh, FILL_WHITE))

# Provincial totals for ret_sa
prov_sa = [
    ("Newfoundland and Labrador", 1446859543),
    ("Prince Edward Island", 1446859574),
    ("Nova Scotia", 1446859605),
    ("New Brunswick", 1446859636),
    ("Quebec", 1446859667),
    ("Ontario", 1446859789),
    ("Manitoba", 1446859881),
    ("Saskatchewan", 1446859942),
    ("Alberta", 1446859973),
    ("British Columbia", 1446860064),
]
for geo, vec in prov_sa:
    fl = f"{geo} - Retail trade SA"
    rows_to_append.append(("Sales", "monthly", "ret_sa", "Retail trade SA", "20-10-0056-01", "Geography", geo, "Industry", "Retail trade [44-45]", str(vec), fl, "Retail trade", FILL_WHITE))

# ret_vol: 1 row, no dims
rows_to_append.append(("Sales", "monthly", "ret_vol", "Retail trade volume index", "20-10-0067-01", "", "", "", "", "1446870181", "Retail sales chained Fisher volume index (2017=100)", "Volume index", FILL_WHITE))

# mfg_real: 22 rows, dim1_name=Industry, no dim2
mfg_real_rows = [
    ("Total, durable and non-durable goods", 123263908, "Real manufacturing sales: Total", "Total manufacturing"),
    ("Food manufacturing", 123263910, "Real manufacturing sales: Food manufacturing", "Food"),
    ("Beverage and tobacco product manufacturing", 123263911, "Real manufacturing sales: Beverage and tobacco", "Beverage & tobacco"),
    ("Textile mills", 123263912, "Real manufacturing sales: Textile mills", "Textile mills"),
    ("Textile product mills", 123263913, "Real manufacturing sales: Textile product mills", "Textile products"),
    ("Apparel manufacturing", 123263914, "Real manufacturing sales: Apparel", "Apparel"),
    ("Leather and allied product manufacturing", 123263915, "Real manufacturing sales: Leather products", "Leather products"),
    ("Paper manufacturing", 123263916, "Real manufacturing sales: Paper", "Paper"),
    ("Printing and related support activities", 123263917, "Real manufacturing sales: Printing", "Printing"),
    ("Petroleum and coal products manufacturing", 123263918, "Real manufacturing sales: Petroleum & coal", "Petroleum & coal"),
    ("Chemical manufacturing", 123263919, "Real manufacturing sales: Chemicals", "Chemicals"),
    ("Plastics and rubber products manufacturing", 123263920, "Real manufacturing sales: Plastics & rubber", "Plastics & rubber"),
    ("Wood product manufacturing", 123263922, "Real manufacturing sales: Wood products", "Wood products"),
    ("Non-metallic mineral product manufacturing", 123263923, "Real manufacturing sales: Non-metallic minerals", "Non-metallic minerals"),
    ("Primary metal manufacturing", 123263924, "Real manufacturing sales: Primary metals", "Primary metals"),
    ("Fabricated metal product manufacturing", 123263925, "Real manufacturing sales: Fabricated metals", "Fabricated metals"),
    ("Machinery manufacturing", 123263926, "Real manufacturing sales: Machinery", "Machinery"),
    ("Computer and electronic product manufacturing", 123263927, "Real manufacturing sales: Computers & electronics", "Computers & electronics"),
    ("Electrical equipment, appliance and component manufacturing", 123263928, "Real manufacturing sales: Electrical equipment", "Electrical equipment"),
    ("Transportation equipment manufacturing", 123263929, "Real manufacturing sales: Transportation equipment", "Transportation equip."),
    ("Furniture and related product manufacturing", 123263937, "Real manufacturing sales: Furniture", "Furniture"),
    ("Miscellaneous manufacturing", 123263938, "Real manufacturing sales: Miscellaneous", "Miscellaneous"),
]
for d1v, vec, fl, sl in mfg_real_rows:
    rows_to_append.append(("Business and industry", "monthly", "mfg_real", "Real manufacturing sales (2017$, SA)", "16-10-0013-01", "Industry", d1v, "", "", str(vec), fl, sl, FILL_WHITE))

# mfg_nom: 44 rows (2 adjustments × 22 industries)
mfg_nom_industries = [
    ("Manufacturing [31-33]", "Total"),
    ("Food manufacturing [311]", "Food"),
    ("Beverage and tobacco product manufacturing [312]", "Beverage & tobacco"),
    ("Textile mills [313]", "Textile mills"),
    ("Textile product mills [314]", "Textile products"),
    ("Apparel manufacturing [315]", "Apparel"),
    ("Leather and allied product manufacturing [316]", "Leather products"),
    ("Paper manufacturing [322]", "Paper"),
    ("Printing and related support activities [323]", "Printing"),
    ("Petroleum and coal product manufacturing [324]", "Petroleum & coal"),
    ("Chemical manufacturing [325]", "Chemicals"),
    ("Plastics and rubber products manufacturing [326]", "Plastics & rubber"),
    ("Wood product manufacturing [321]", "Wood products"),
    ("Non-metallic mineral product manufacturing [327]", "Non-metallic minerals"),
    ("Primary metal manufacturing [331]", "Primary metals"),
    ("Fabricated metal product manufacturing [332]", "Fabricated metals"),
    ("Machinery manufacturing [333]", "Machinery"),
    ("Computer and electronic product manufacturing [334]", "Computers & electronics"),
    ("Electrical equipment, appliance and component manufacturing [335]", "Electrical equipment"),
    ("Transportation equipment manufacturing [336]", "Transportation equip."),
    ("Furniture and related product manufacturing [337]", "Furniture"),
    ("Miscellaneous manufacturing [339]", "Miscellaneous"),
]
mfg_unadj_vecs = [800025, 800027, 800078, 800087, 800098, 1745902, 800129, 800136, 800155, 800164, 800171, 800202, 800223, 800245, 800264, 800287, 800324, 800353, 800370, 800389, 800417, 800418]
mfg_sa_vecs = [800450, 800452, 800453, 800454, 800455, 1745922, 800457, 800458, 800459, 800460, 800461, 800462, 800464, 800465, 800466, 800467, 800468, 800469, 800470, 800471, 800479, 800480]

for i, ((ind_name, ind_short), uv) in enumerate(zip(mfg_nom_industries, mfg_unadj_vecs)):
    fl = f"Unadjusted - {ind_name}"
    rows_to_append.append(("Business and industry", "monthly", "mfg_nom", "Manufacturing sales by industry", "16-10-0047-01", "Adjustment", "Unadjusted", "Industry", ind_name, str(uv), fl, ind_short, FILL_WHITE))

for i, ((ind_name, ind_short), sv) in enumerate(zip(mfg_nom_industries, mfg_sa_vecs)):
    fl = f"Seasonally adjusted - {ind_name}"
    rows_to_append.append(("Business and industry", "monthly", "mfg_nom", "Manufacturing sales by industry", "16-10-0047-01", "Adjustment", "Seasonally adjusted", "Industry", ind_name, str(sv), fl, ind_short, FILL_WHITE))

# mv_vehicle: 3 rows
mv_vehicle_rows = [
    ("Total, all vehicle types", 1617802379, "Motor vehicle sales: Total", "Total"),
    ("Passenger cars", 1617802382, "Motor vehicle sales: Passenger cars", "Passenger cars"),
    ("Trucks", 1617802385, "Motor vehicle sales: Trucks", "Trucks"),
]
for d1v, vec, fl, sl in mv_vehicle_rows:
    rows_to_append.append(("Sales", "monthly", "mv_vehicle", "Motor vehicle sales by vehicle type", "20-10-0085-01", "Vehicle type", d1v, "", "", str(vec), fl, sl, FILL_WHITE))

# mv_fuel: 3 rows
mv_fuel_rows = [
    ("Total, all fuel types", 1617802379, "Motor vehicle sales: Total", "Total"),
    ("Zero-emission vehicles", 1645353696, "Motor vehicle sales: Zero-emission", "Zero-emission"),
    ("All other fuel types", 1645353729, "Motor vehicle sales: All other fuel types", "Other fuel types"),
]
for d1v, vec, fl, sl in mv_fuel_rows:
    rows_to_append.append(("Sales", "monthly", "mv_fuel", "Motor vehicle sales by fuel type", "20-10-0085-01", "Fuel type", d1v, "", "", str(vec), fl, sl, FILL_WHITE))

# ── CATEGORY: Trade (grey EEEEEE, monthly) ────────────────────────────────────

# ── 1. Goods trade by trading partners (12-10-0011-01) ──────────────────────
# pid=1210001101 | dim1=Flow & Adjustment, dim2=Country
# 5 flow/adj × 8 countries = 40 rows

PARTNER_VECTORS = {
    ("Import (NSA)",  "All countries"):  87008810,
    ("Import (NSA)",  "United States"):  87008811,
    ("Import (NSA)",  "European Union"): 87008812,
    ("Import (NSA)",  "China"):          87008820,
    ("Import (NSA)",  "United Kingdom"): 87008813,
    ("Import (NSA)",  "Mexico"):         87008821,
    ("Import (NSA)",  "Japan"):          87008822,
    ("Import (NSA)",  "South Korea"):    87008823,
    ("Import (SA)",   "All countries"):  87008839,
    ("Import (SA)",   "United States"):  87008840,
    ("Import (SA)",   "European Union"): 87008841,
    ("Import (SA)",   "China"):          87008849,
    ("Import (SA)",   "United Kingdom"): 87008842,
    ("Import (SA)",   "Mexico"):         87008850,
    ("Import (SA)",   "Japan"):          87008851,
    ("Import (SA)",   "South Korea"):    87008852,
    ("Export (NSA)",  "All countries"):  87008926,
    ("Export (NSA)",  "United States"):  87008927,
    ("Export (NSA)",  "European Union"): 87008928,
    ("Export (NSA)",  "China"):          87008936,
    ("Export (NSA)",  "United Kingdom"): 87008929,
    ("Export (NSA)",  "Mexico"):         87008937,
    ("Export (NSA)",  "Japan"):          87008938,
    ("Export (NSA)",  "South Korea"):    87008939,
    ("Export (SA)",   "All countries"):  87008955,
    ("Export (SA)",   "United States"):  87008956,
    ("Export (SA)",   "European Union"): 87008957,
    ("Export (SA)",   "China"):          87008965,
    ("Export (SA)",   "United Kingdom"): 87008958,
    ("Export (SA)",   "Mexico"):         87008966,
    ("Export (SA)",   "Japan"):          87008967,
    ("Export (SA)",   "South Korea"):    87008968,
    ("Balance (SA)",  "All countries"):  87008984,
    ("Balance (SA)",  "United States"):  87008985,
    ("Balance (SA)",  "European Union"): 87008986,
    ("Balance (SA)",  "China"):          87008994,
    ("Balance (SA)",  "United Kingdom"): 87008987,
    ("Balance (SA)",  "Mexico"):         87008995,
    ("Balance (SA)",  "Japan"):          87008996,
    ("Balance (SA)",  "South Korea"):    87008997,
}
PARTNER_FLOWS = [
    "Import (NSA)", "Import (SA)",
    "Export (NSA)", "Export (SA)",
    "Balance (SA)",
]
PARTNER_COUNTRIES = [
    "All countries", "United States", "European Union", "China",
    "United Kingdom", "Mexico", "Japan", "South Korea",
]
COUNTRY_SHORT = {
    "All countries": "All countries", "United States": "US",
    "European Union": "EU",          "China": "China",
    "United Kingdom": "UK",          "Mexico": "Mexico",
    "Japan": "Japan",                "South Korea": "South Korea",
}
for flow in PARTNER_FLOWS:
    for country in PARTNER_COUNTRIES:
        vec = PARTNER_VECTORS[(flow, country)]
        fl  = f"{flow} — {country}"
        sl  = COUNTRY_SHORT[country]
        rows_to_append.append(("Trade", "monthly", "goods_trading_partners",
                               "Goods trade by trading partners", "12-10-0011-01",
                               "Flow & Adjustment", flow, "Country", country,
                               str(vec), fl, sl, FILL_GREY))


# ── 2. Trade in services (12-10-0144-01) ────────────────────────────────────
# pid=1210014401 | dim1=Flow & Adjustment, dim2=Service type
# 6 flow/adj × 5 service types = 30 rows
# Vectors are sequential pairs (NSA base, SA base+1) then +2 per service type

SVC_FLOWS = [
    ("Export (NSA)",  1105277794),
    ("Export (SA)",   1105277795),
    ("Import (NSA)",  1105277804),
    ("Import (SA)",   1105277805),
    ("Balance (NSA)", 1105277814),
    ("Balance (SA)",  1105277815),
]
SVC_TYPES = [
    ("Total services",          "Total"),
    ("Commercial services",     "Commercial"),
    ("Travel services",         "Travel"),
    ("Transportation services", "Transportation"),
    ("Government services",     "Government"),
]
for flow_label, base_vec in SVC_FLOWS:
    for i, (svc_full, svc_short) in enumerate(SVC_TYPES):
        vec = base_vec + 2 * i
        fl  = f"{flow_label} — {svc_full}"
        rows_to_append.append(("Trade", "monthly", "trade_services",
                               "Trade in services", "12-10-0144-01",
                               "Flow & Adjustment", flow_label,
                               "Service type", svc_full,
                               str(vec), fl, svc_short, FILL_GREY))


# ── 3 & 4. Goods by commodity, nominal (12-10-0163-01) ────────────────────
# pid=1210016301 | unified series: dims=5
#   dim1 = Trade direction       (Import / Export)
#   dim2 = Seasonal adj.         (Unadjusted only for now — SA vectors TBD)
#                                 → single option so UI auto-skips this step
#   dim3 = Commodity group       (L2)
#   dim4 = Sub-commodity         (L3; same as dim3 for L2-total rows)
#   dim5 = Sub-sub-commodity     (L4; same as dim4 when no drill-down → auto-add)
#
# Vector formula: import_nsa_vec = COMM_BASE + member_id - 1
# COMM_BASE = 1566910579  (member 1 = Total of all merchandise)
# Full 153-member NAPCS 2022 hierarchy from getCubeMetadata API (2026-03-22)
# EXP_NSA_OFFSET = 606  (Export NSA = Import NSA + 606; confirmed)

COMM_BASE = 1566910579

def cvec(member_id):
    return COMM_BASE + member_id - 1

# (dim3=L2, dim4=L3, dim5=L4, import_nsa_vec)
# dim3==dim4==dim5 → L2 total (auto-add if only dim4 option)
# dim4==dim5       → L3 item with no L4 drill-down (auto-add at dim5 step)
# dim4!=dim5       → specific L4 item (shown in dim5 dropdown)

COMMODITY_ROWS = [
    # ── Total ───────────────────────────────────────────────────────────────
    ("Total of all merchandise", "Total of all merchandise", "Total of all merchandise", cvec(1)),

    # ── Farm, fishing and intermediate food products ─────────────────────────
    ("Farm, fishing and intermediate food products", "Farm, fishing and intermediate food products", "Farm, fishing and intermediate food products", cvec(2)),
    # L3: Farm and fishing products → 8 L4 sub-items
    ("Farm, fishing and intermediate food products", "Farm and fishing products", "Farm and fishing products",                                    cvec(3)),
    ("Farm, fishing and intermediate food products", "Farm and fishing products", "Live animals",                                                 cvec(4)),
    ("Farm, fishing and intermediate food products", "Farm and fishing products", "Wheat",                                                        cvec(5)),
    ("Farm, fishing and intermediate food products", "Farm and fishing products", "Canola (including rapeseed)",                                  cvec(6)),
    ("Farm, fishing and intermediate food products", "Farm and fishing products", "Fresh fruit, nuts and vegetables, and pulse crops",            cvec(7)),
    ("Farm, fishing and intermediate food products", "Farm and fishing products", "Other crop products",                                          cvec(8)),
    ("Farm, fishing and intermediate food products", "Farm and fishing products", "Other animal products",                                        cvec(9)),
    ("Farm, fishing and intermediate food products", "Farm and fishing products", "Fish, crustaceans, shellfish and other fishery products",      cvec(10)),
    ("Farm, fishing and intermediate food products", "Farm and fishing products", "Animal feed",                                                  cvec(11)),
    # L3: Intermediate food products → same-name L4 → auto-add
    ("Farm, fishing and intermediate food products", "Intermediate food products", "Intermediate food products",                                  cvec(12)),

    # ── Energy products ──────────────────────────────────────────────────────
    ("Energy products", "Energy products", "Energy products", cvec(14)),
    # L3: Crude oil → 1 L4 (different name)
    ("Energy products", "Crude oil and crude bitumen", "Crude oil and crude bitumen",                                                            cvec(15)),
    ("Energy products", "Crude oil and crude bitumen", "Crude oil and bitumen",                                                                  cvec(16)),
    # L3: Natural gas → 2 L4 sub-items
    ("Energy products", "Natural gas, natural gas liquids and related products", "Natural gas, natural gas liquids and related products",         cvec(17)),
    ("Energy products", "Natural gas, natural gas liquids and related products", "Natural gas",                                                   cvec(18)),
    ("Energy products", "Natural gas, natural gas liquids and related products", "Natural gas liquids (including condensate) and related products", cvec(19)),
    # L3: same-name L4 → auto-add each
    ("Energy products", "Coal", "Coal", cvec(20)),
    ("Energy products", "Nuclear fuel and other energy products", "Nuclear fuel and other energy products", cvec(22)),
    ("Energy products", "Electricity", "Electricity", cvec(24)),
    # L3: Refined petroleum → 1 L4 (slightly different name)
    ("Energy products", "Refined petroleum energy products", "Refined petroleum energy products",                                                cvec(26)),
    ("Energy products", "Refined petroleum energy products", "Refined petroleum energy products (including liquid biofuels)",                    cvec(27)),

    # ── Metal ores and non-metallic minerals ─────────────────────────────────
    ("Metal ores and non-metallic minerals", "Metal ores and non-metallic minerals", "Metal ores and non-metallic minerals", cvec(28)),
    # L3: Metal ores and concentrates → 5 L4 sub-items
    ("Metal ores and non-metallic minerals", "Metal ores and concentrates", "Metal ores and concentrates",                                       cvec(29)),
    ("Metal ores and non-metallic minerals", "Metal ores and concentrates", "Iron ores and concentrates",                                        cvec(30)),
    ("Metal ores and non-metallic minerals", "Metal ores and concentrates", "Copper ores and concentrates",                                      cvec(31)),
    ("Metal ores and non-metallic minerals", "Metal ores and concentrates", "Nickel ores and concentrates",                                      cvec(32)),
    ("Metal ores and non-metallic minerals", "Metal ores and concentrates", "Radioactive ores and concentrates",                                 cvec(33)),
    ("Metal ores and non-metallic minerals", "Metal ores and concentrates", "Other metal ores and concentrates",                                 cvec(34)),
    # L3: Non-metallic minerals → 3 L4 sub-items
    ("Metal ores and non-metallic minerals", "Non-metallic minerals", "Non-metallic minerals",                                                   cvec(35)),
    ("Metal ores and non-metallic minerals", "Non-metallic minerals", "Potash",                                                                  cvec(36)),
    ("Metal ores and non-metallic minerals", "Non-metallic minerals", "Stone, sand, gravel, clay, and refractory minerals",                     cvec(37)),
    ("Metal ores and non-metallic minerals", "Non-metallic minerals", "Diamonds and other non-metallic minerals (except cut gemstones)",         cvec(38)),

    # ── Metal and non-metallic mineral products ───────────────────────────────
    ("Metal and non-metallic mineral products", "Metal and non-metallic mineral products", "Metal and non-metallic mineral products", cvec(39)),
    # L3: Intermediate metal products → 9 L4 sub-items (confirmed v1566910619–627)
    ("Metal and non-metallic mineral products", "Intermediate metal products", "Intermediate metal products",                                    cvec(40)),
    ("Metal and non-metallic mineral products", "Intermediate metal products", "Unwrought iron, steel and ferro-alloys",                         cvec(41)),
    ("Metal and non-metallic mineral products", "Intermediate metal products", "Basic and semi-finished iron or steel products",                 cvec(42)),
    ("Metal and non-metallic mineral products", "Intermediate metal products", "Unwrought aluminum and aluminum alloys",                         cvec(43)),
    ("Metal and non-metallic mineral products", "Intermediate metal products", "Unwrought copper and copper alloys",                             cvec(44)),
    ("Metal and non-metallic mineral products", "Intermediate metal products", "Unwrought nickel and nickel alloys",                             cvec(45)),
    ("Metal and non-metallic mineral products", "Intermediate metal products", "Unwrought gold, silver, and platinum group metals, and their alloys", cvec(46)),
    ("Metal and non-metallic mineral products", "Intermediate metal products", "Other unwrought non-ferrous metals and non-ferrous metal alloys", cvec(47)),
    ("Metal and non-metallic mineral products", "Intermediate metal products", "Basic and semi-finished products of non-ferrous metals and non-ferrous metal alloys (except aluminum)", cvec(48)),
    ("Metal and non-metallic mineral products", "Intermediate metal products", "Basic and semi-finished products of aluminum and aluminum alloys", cvec(49)),
    # L3: same-name L4 → auto-add
    ("Metal and non-metallic mineral products", "Fabricated metal products", "Fabricated metal products",                                        cvec(50)),
    ("Metal and non-metallic mineral products", "Non-metallic mineral products", "Non-metallic mineral products",                                cvec(52)),
    # L3: Waste and scrap → 2 L4 sub-items
    ("Metal and non-metallic mineral products", "Waste and scrap of metal and glass", "Waste and scrap of metal and glass",                      cvec(54)),
    ("Metal and non-metallic mineral products", "Waste and scrap of metal and glass", "Waste and scrap of metal",                                cvec(55)),
    ("Metal and non-metallic mineral products", "Waste and scrap of metal and glass", "Waste and scrap of glass",                                cvec(56)),

    # ── Basic and industrial chemical, plastic and rubber products ────────────
    ("Basic and industrial chemical, plastic and rubber products", "Basic and industrial chemical, plastic and rubber products", "Basic and industrial chemical, plastic and rubber products", cvec(57)),
    # L3: Basic chemicals → 4 L4 sub-items
    ("Basic and industrial chemical, plastic and rubber products", "Basic chemicals and industrial chemical products", "Basic chemicals and industrial chemical products",   cvec(58)),
    ("Basic and industrial chemical, plastic and rubber products", "Basic chemicals and industrial chemical products", "Dyes and pigments, and petrochemicals",             cvec(59)),
    ("Basic and industrial chemical, plastic and rubber products", "Basic chemicals and industrial chemical products", "Lubricants and other petroleum refinery products", cvec(60)),
    ("Basic and industrial chemical, plastic and rubber products", "Basic chemicals and industrial chemical products", "Basic chemicals",                                   cvec(61)),
    ("Basic and industrial chemical, plastic and rubber products", "Basic chemicals and industrial chemical products", "Fertilizers, pesticides and other chemical products", cvec(62)),
    # L3: Plastic and rubber → 3 L4 sub-items
    ("Basic and industrial chemical, plastic and rubber products", "Plastic and rubber products", "Plastic and rubber products",                                            cvec(63)),
    ("Basic and industrial chemical, plastic and rubber products", "Plastic and rubber products", "Plastic resins",                                                        cvec(64)),
    ("Basic and industrial chemical, plastic and rubber products", "Plastic and rubber products", "Plastic and rubber basic products not for packaging use (except plastic resins)", cvec(65)),
    ("Basic and industrial chemical, plastic and rubber products", "Plastic and rubber products", "Plastic and rubber finished products",                                  cvec(66)),
    # L3: same-name L4 → auto-add
    ("Basic and industrial chemical, plastic and rubber products", "Waste and scrap of plastic and rubber", "Waste and scrap of plastic and rubber",                       cvec(67)),

    # ── Forestry products and building and packaging materials ────────────────
    ("Forestry products and building and packaging materials", "Forestry products and building and packaging materials", "Forestry products and building and packaging materials", cvec(69)),
    # L3: same-name L4 → auto-add
    ("Forestry products and building and packaging materials", "Logs, pulpwood and other forestry products", "Logs, pulpwood and other forestry products",                  cvec(70)),
    ("Forestry products and building and packaging materials", "Pulp and paper", "Pulp and paper",                                                                          cvec(72)),
    # L3: Building and packaging → 9 L4 sub-items
    ("Forestry products and building and packaging materials", "Building and packaging materials", "Building and packaging materials",                                       cvec(74)),
    ("Forestry products and building and packaging materials", "Building and packaging materials", "Lumber and other sawmill products",                                     cvec(75)),
    ("Forestry products and building and packaging materials", "Building and packaging materials", "Asphalt (except natural) and asphalt products",                         cvec(76)),
    ("Forestry products and building and packaging materials", "Building and packaging materials", "Wood millwork, and wood products not elsewhere classified",              cvec(77)),
    ("Forestry products and building and packaging materials", "Building and packaging materials", "Paints, coatings, and adhesive products",                               cvec(78)),
    ("Forestry products and building and packaging materials", "Building and packaging materials", "Plastic and foam building and construction materials",                   cvec(79)),
    ("Forestry products and building and packaging materials", "Building and packaging materials", "Cement, lime and gypsum products",                                      cvec(80)),
    ("Forestry products and building and packaging materials", "Building and packaging materials", "Metal building and construction materials",                              cvec(81)),
    ("Forestry products and building and packaging materials", "Building and packaging materials", "Prefabricated buildings and components thereof",                         cvec(82)),
    ("Forestry products and building and packaging materials", "Building and packaging materials", "Packaging materials",                                                   cvec(83)),
    # L3: same-name L4 → auto-add
    ("Forestry products and building and packaging materials", "Waste and scrap of wood, wood by-products, paper and paperboard", "Waste and scrap of wood, wood by-products, paper and paperboard", cvec(84)),

    # ── Industrial machinery, equipment and parts ─────────────────────────────
    # L2 (86) and L3 (87) have same name → treat L4 items (88-95) as dim4 level
    ("Industrial machinery, equipment and parts", "Industrial machinery, equipment and parts", "Industrial machinery, equipment and parts",                                  cvec(86)),
    ("Industrial machinery, equipment and parts", "Agricultural, lawn and garden machinery and equipment", "Agricultural, lawn and garden machinery and equipment",          cvec(88)),
    ("Industrial machinery, equipment and parts", "Logging, construction, mining, and oil and gas field machinery and equipment", "Logging, construction, mining, and oil and gas field machinery and equipment", cvec(89)),
    ("Industrial machinery, equipment and parts", "Metalworking machinery", "Metalworking machinery",                                                                        cvec(90)),
    ("Industrial machinery, equipment and parts", "Commercial and service industry machinery and equipment", "Commercial and service industry machinery and equipment",       cvec(91)),
    ("Industrial machinery, equipment and parts", "Other industry-specific manufacturing machinery, not elsewhere classified", "Other industry-specific manufacturing machinery, not elsewhere classified", cvec(92)),
    ("Industrial machinery, equipment and parts", "Heating, cooling and air purification equipment", "Heating, cooling and air purification equipment",                      cvec(93)),
    ("Industrial machinery, equipment and parts", "Other general-purpose machinery and equipment, not elsewhere classified", "Other general-purpose machinery and equipment, not elsewhere classified", cvec(94)),
    ("Industrial machinery, equipment and parts", "Parts of industrial machinery and equipment", "Parts of industrial machinery and equipment",                              cvec(95)),

    # ── Electronic and electrical equipment and parts ─────────────────────────
    ("Electronic and electrical equipment and parts", "Electronic and electrical equipment and parts", "Electronic and electrical equipment and parts", cvec(96)),
    # L3: same-name L4 → auto-add
    ("Electronic and electrical equipment and parts", "Computers and computer peripherals", "Computers and computer peripherals",                                            cvec(97)),
    ("Electronic and electrical equipment and parts", "Communication, and audio and video equipment", "Communication, and audio and video equipment",                        cvec(99)),
    # L3: Other electronic → 3 L4 sub-items
    ("Electronic and electrical equipment and parts", "Other electronic and electrical machinery, equipment and parts", "Other electronic and electrical machinery, equipment and parts", cvec(101)),
    ("Electronic and electrical equipment and parts", "Other electronic and electrical machinery, equipment and parts", "Medical, measuring, and other electronic and electrical machinery and equipment", cvec(102)),
    ("Electronic and electrical equipment and parts", "Other electronic and electrical machinery, equipment and parts", "Electronic and electrical parts",                   cvec(103)),
    ("Electronic and electrical equipment and parts", "Other electronic and electrical machinery, equipment and parts", "Electrical components",                              cvec(104)),

    # ── Motor vehicles and parts ──────────────────────────────────────────────
    ("Motor vehicles and parts", "Motor vehicles and parts", "Motor vehicles and parts", cvec(105)),
    # L3: same-name L4 → auto-add
    ("Motor vehicles and parts", "Passenger cars and light trucks", "Passenger cars and light trucks",                                                                       cvec(106)),
    ("Motor vehicles and parts", "Medium and heavy trucks, buses, and other motor vehicles", "Medium and heavy trucks, buses, and other motor vehicles",                     cvec(108)),
    # L3: Tires → 2 L4 sub-items
    ("Motor vehicles and parts", "Tires; motor vehicle engines and motor vehicle parts", "Tires; motor vehicle engines and motor vehicle parts",                              cvec(110)),
    ("Motor vehicles and parts", "Tires; motor vehicle engines and motor vehicle parts", "Tires",                                                                            cvec(111)),
    ("Motor vehicles and parts", "Tires; motor vehicle engines and motor vehicle parts", "Motor vehicle engines and motor vehicle parts",                                    cvec(112)),

    # ── Aircraft and other transportation equipment and parts ─────────────────
    ("Aircraft and other transportation equipment and parts", "Aircraft and other transportation equipment and parts", "Aircraft and other transportation equipment and parts", cvec(113)),
    # L3: Aircraft engines and parts → 2 L4 sub-items
    ("Aircraft and other transportation equipment and parts", "Aircraft, aircraft engines and aircraft parts", "Aircraft, aircraft engines and aircraft parts",               cvec(114)),
    ("Aircraft and other transportation equipment and parts", "Aircraft, aircraft engines and aircraft parts", "Aircraft",                                                    cvec(115)),
    ("Aircraft and other transportation equipment and parts", "Aircraft, aircraft engines and aircraft parts", "Aircraft engines, aircraft parts and other aerospace equipment", cvec(116)),
    # L3: Other transportation → 3 L4 sub-items
    ("Aircraft and other transportation equipment and parts", "Other transportation equipment and parts", "Other transportation equipment and parts",                          cvec(117)),
    ("Aircraft and other transportation equipment and parts", "Other transportation equipment and parts", "Ships, locomotives, railway rolling stock, and rapid transit equipment", cvec(118)),
    ("Aircraft and other transportation equipment and parts", "Other transportation equipment and parts", "Boats and other transportation equipment",                          cvec(119)),
    ("Aircraft and other transportation equipment and parts", "Other transportation equipment and parts", "Parts of railway rolling stock and of other transportation equipment", cvec(120)),

    # ── Consumer goods ────────────────────────────────────────────────────────
    ("Consumer goods", "Consumer goods", "Consumer goods", cvec(121)),
    # L3: Food, beverage and tobacco → 9 L4 sub-items (note: member IDs 123/124 swapped)
    ("Consumer goods", "Food, beverage and tobacco products", "Food, beverage and tobacco products",                                                                          cvec(122)),
    ("Consumer goods", "Food, beverage and tobacco products", "Meat products",                                                                                                cvec(123)),
    ("Consumer goods", "Food, beverage and tobacco products", "Prepared and packaged seafood products",                                                                       cvec(124)),
    ("Consumer goods", "Food, beverage and tobacco products", "Dairy products",                                                                                               cvec(125)),
    ("Consumer goods", "Food, beverage and tobacco products", "Other food products",                                                                                          cvec(126)),
    ("Consumer goods", "Food, beverage and tobacco products", "Coffee and tea",                                                                                               cvec(127)),
    ("Consumer goods", "Food, beverage and tobacco products", "Frozen, fresh and canned fruit and vegetable juices",                                                          cvec(128)),
    ("Consumer goods", "Food, beverage and tobacco products", "Carbonated and non-carbonated drinks (including low alcohol fermented drinks), bottled water and ice",         cvec(129)),
    ("Consumer goods", "Food, beverage and tobacco products", "Alcoholic beverages",                                                                                          cvec(130)),
    ("Consumer goods", "Food, beverage and tobacco products", "Tobacco products (including electronic cigarettes)",                                                           cvec(131)),
    # L3: Clothing, footwear → 3 L4 sub-items
    ("Consumer goods", "Clothing, footwear and textile products", "Clothing, footwear and textile products",                                                                  cvec(132)),
    ("Consumer goods", "Clothing, footwear and textile products", "Fabric, fibre and yarn, and leather and dressed furs",                                                     cvec(133)),
    ("Consumer goods", "Clothing, footwear and textile products", "Clothing, footwear and accessories",                                                                       cvec(134)),
    ("Consumer goods", "Clothing, footwear and textile products", "Carpets, textile furnishings and other textile products",                                                   cvec(135)),
    # L3: Paper and published → 3 L4 sub-items
    ("Consumer goods", "Paper and published products", "Paper and published products",                                                                                        cvec(136)),
    ("Consumer goods", "Paper and published products", "Converted paper products (except for packaging)",                                                                     cvec(137)),
    ("Consumer goods", "Paper and published products", "Published products and recorded media (except software)",                                                             cvec(138)),
    ("Consumer goods", "Paper and published products", "Software and software licensing",                                                                                     cvec(139)),
    # L3: same-name L4 → auto-add
    ("Consumer goods", "Pharmaceutical and medicinal products", "Pharmaceutical and medicinal products",                                                                      cvec(140)),
    ("Consumer goods", "Furniture and fixtures", "Furniture and fixtures",                                                                                                    cvec(142)),
    # L3: Cleaning products etc. → 3 L4 sub-items
    ("Consumer goods", "Cleaning products, appliances, and miscellaneous goods and supplies", "Cleaning products, appliances, and miscellaneous goods and supplies",          cvec(144)),
    ("Consumer goods", "Cleaning products, appliances, and miscellaneous goods and supplies", "Cleaning products and toiletries",                                              cvec(145)),
    ("Consumer goods", "Cleaning products, appliances, and miscellaneous goods and supplies", "Appliances",                                                                   cvec(146)),
    ("Consumer goods", "Cleaning products, appliances, and miscellaneous goods and supplies", "Miscellaneous goods and supplies",                                              cvec(147)),

    # ── Special transactions trade ─────────────────────────────────────────────
    ("Special transactions trade", "Special transactions trade", "Special transactions trade", cvec(148)),

    # ── Other balance of payments adjustments ─────────────────────────────────
    ("Other balance of payments adjustments", "Other balance of payments adjustments", "Other balance of payments adjustments", cvec(151)),
]

DIM_SHORT = {
    # L2 totals
    "Total of all merchandise":                                                         "Total",
    "Farm, fishing and intermediate food products":                                     "Farm & fishing (total)",
    "Energy products":                                                                  "Energy (total)",
    "Metal ores and non-metallic minerals":                                             "Metal ores (total)",
    "Metal and non-metallic mineral products":                                          "Metal products (total)",
    "Basic and industrial chemical, plastic and rubber products":                       "Chemical & rubber (total)",
    "Forestry products and building and packaging materials":                           "Forestry & building (total)",
    "Industrial machinery, equipment and parts":                                        "Industrial mach. (total)",
    "Electronic and electrical equipment and parts":                                    "Electronic & elec. (total)",
    "Motor vehicles and parts":                                                         "Motor vehicles (total)",
    "Aircraft and other transportation equipment and parts":                            "Aircraft & transport (total)",
    "Consumer goods":                                                                   "Consumer goods (total)",
    "Special transactions trade":                                                       "Special transactions",
    "Other balance of payments adjustments":                                            "Other BOP adj.",
    # L3 items
    "Farm and fishing products":                                                        "Farm & fishing",
    "Intermediate food products":                                                       "Intermediate food",
    "Crude oil and crude bitumen":                                                      "Crude oil",
    "Natural gas, natural gas liquids and related products":                            "Natural gas (total)",
    "Coal":                                                                             "Coal",
    "Nuclear fuel and other energy products":                                           "Nuclear fuel",
    "Electricity":                                                                      "Electricity",
    "Refined petroleum energy products":                                                "Refined petroleum",
    "Metal ores and concentrates":                                                      "Metal ores & concentrates",
    "Non-metallic minerals":                                                            "Non-metallic minerals",
    "Intermediate metal products":                                                      "Intermediate metal",
    "Fabricated metal products":                                                        "Fabricated metal",
    "Non-metallic mineral products":                                                    "Non-metallic mineral products",
    "Waste and scrap of metal and glass":                                               "Metal/glass scrap",
    "Basic chemicals and industrial chemical products":                                 "Basic chemicals (total)",
    "Plastic and rubber products":                                                      "Plastic & rubber",
    "Waste and scrap of plastic and rubber":                                            "Plastic/rubber scrap",
    "Logs, pulpwood and other forestry products":                                       "Logs & pulpwood",
    "Pulp and paper":                                                                   "Pulp & paper",
    "Building and packaging materials":                                                 "Building & packaging",
    "Waste and scrap of wood, wood by-products, paper and paperboard":                  "Wood/paper scrap",
    "Agricultural, lawn and garden machinery and equipment":                            "Agri. machinery",
    "Logging, construction, mining, and oil and gas field machinery and equipment":     "Mining & construction mach.",
    "Metalworking machinery":                                                           "Metalworking mach.",
    "Commercial and service industry machinery and equipment":                          "Commercial & service mach.",
    "Other industry-specific manufacturing machinery, not elsewhere classified":        "Other industry mach.",
    "Heating, cooling and air purification equipment":                                  "HVAC equipment",
    "Other general-purpose machinery and equipment, not elsewhere classified":          "Other general mach.",
    "Parts of industrial machinery and equipment":                                      "Industrial mach. parts",
    "Computers and computer peripherals":                                               "Computers",
    "Communication, and audio and video equipment":                                     "Communication & AV",
    "Other electronic and electrical machinery, equipment and parts":                   "Other electronic",
    "Passenger cars and light trucks":                                                  "Passenger cars & light trucks",
    "Medium and heavy trucks, buses, and other motor vehicles":                         "Medium/heavy trucks & buses",
    "Tires; motor vehicle engines and motor vehicle parts":                             "Tires & MV parts",
    "Aircraft, aircraft engines and aircraft parts":                                    "Aircraft & engines",
    "Other transportation equipment and parts":                                         "Other transport",
    "Food, beverage and tobacco products":                                              "Food, bev. & tobacco",
    "Clothing, footwear and textile products":                                          "Clothing & textiles",
    "Paper and published products":                                                     "Paper & published",
    "Pharmaceutical and medicinal products":                                            "Pharmaceuticals",
    "Furniture and fixtures":                                                           "Furniture",
    "Cleaning products, appliances, and miscellaneous goods and supplies":              "Cleaning & misc.",
    # L4 items
    "Live animals":                                                                     "Live animals",
    "Wheat":                                                                            "Wheat",
    "Canola (including rapeseed)":                                                      "Canola",
    "Fresh fruit, nuts and vegetables, and pulse crops":                                "Fruit, veg & pulses",
    "Other crop products":                                                              "Other crops",
    "Other animal products":                                                            "Other animal products",
    "Fish, crustaceans, shellfish and other fishery products":                          "Fish & seafood",
    "Animal feed":                                                                      "Animal feed",
    "Crude oil and bitumen":                                                            "Crude oil & bitumen",
    "Natural gas":                                                                      "Natural gas",
    "Natural gas liquids (including condensate) and related products":                  "NGL & related",
    "Refined petroleum energy products (including liquid biofuels)":                    "Refined petroleum (incl. biofuels)",
    "Iron ores and concentrates":                                                       "Iron ores",
    "Copper ores and concentrates":                                                     "Copper ores",
    "Nickel ores and concentrates":                                                     "Nickel ores",
    "Radioactive ores and concentrates":                                                "Radioactive ores",
    "Other metal ores and concentrates":                                                "Other metal ores",
    "Potash":                                                                           "Potash",
    "Stone, sand, gravel, clay, and refractory minerals":                               "Stone, sand & clay",
    "Diamonds and other non-metallic minerals (except cut gemstones)":                  "Diamonds & other",
    "Unwrought iron, steel and ferro-alloys":                                           "Unwrought iron & steel",
    "Basic and semi-finished iron or steel products":                                   "Semi-finished iron/steel",
    "Unwrought aluminum and aluminum alloys":                                           "Unwrought aluminum",
    "Unwrought copper and copper alloys":                                               "Unwrought copper",
    "Unwrought nickel and nickel alloys":                                               "Unwrought nickel",
    "Unwrought gold, silver, and platinum group metals, and their alloys":              "Unwrought gold/silver/Pt",
    "Other unwrought non-ferrous metals and non-ferrous metal alloys":                  "Other unwrought non-ferrous",
    "Basic and semi-finished products of non-ferrous metals and non-ferrous metal alloys (except aluminum)": "Semi-finished non-ferrous",
    "Basic and semi-finished products of aluminum and aluminum alloys":                 "Semi-finished aluminum",
    "Waste and scrap of metal":                                                         "Metal scrap",
    "Waste and scrap of glass":                                                         "Glass scrap",
    "Dyes and pigments, and petrochemicals":                                            "Dyes & petrochemicals",
    "Lubricants and other petroleum refinery products":                                 "Lubricants & petroleum",
    "Basic chemicals":                                                                  "Basic chemicals",
    "Fertilizers, pesticides and other chemical products":                              "Fertilizers & pesticides",
    "Plastic resins":                                                                   "Plastic resins",
    "Plastic and rubber basic products not for packaging use (except plastic resins)":  "Plastic/rubber basic",
    "Plastic and rubber finished products":                                             "Plastic/rubber finished",
    "Lumber and other sawmill products":                                                "Lumber",
    "Asphalt (except natural) and asphalt products":                                    "Asphalt",
    "Wood millwork, and wood products not elsewhere classified":                         "Wood millwork",
    "Paints, coatings, and adhesive products":                                          "Paints & adhesives",
    "Plastic and foam building and construction materials":                              "Plastic & foam materials",
    "Cement, lime and gypsum products":                                                 "Cement & gypsum",
    "Metal building and construction materials":                                        "Metal building materials",
    "Prefabricated buildings and components thereof":                                   "Prefab buildings",
    "Packaging materials":                                                              "Packaging materials",
    "Medical, measuring, and other electronic and electrical machinery and equipment":  "Medical & measuring equip.",
    "Electronic and electrical parts":                                                  "Electronic parts",
    "Electrical components":                                                            "Electrical components",
    "Tires":                                                                            "Tires",
    "Motor vehicle engines and motor vehicle parts":                                    "MV engines & parts",
    "Aircraft":                                                                         "Aircraft",
    "Aircraft engines, aircraft parts and other aerospace equipment":                   "Aircraft engines & parts",
    "Ships, locomotives, railway rolling stock, and rapid transit equipment":           "Ships & locomotives",
    "Boats and other transportation equipment":                                         "Boats & other transport",
    "Parts of railway rolling stock and of other transportation equipment":             "Rolling stock parts",
    "Meat products":                                                                    "Meat products",
    "Prepared and packaged seafood products":                                           "Seafood products",
    "Dairy products":                                                                   "Dairy products",
    "Other food products":                                                              "Other food",
    "Coffee and tea":                                                                   "Coffee & tea",
    "Frozen, fresh and canned fruit and vegetable juices":                              "Fruit & veg. juices",
    "Carbonated and non-carbonated drinks (including low alcohol fermented drinks), bottled water and ice": "Beverages & water",
    "Alcoholic beverages":                                                              "Alcoholic beverages",
    "Tobacco products (including electronic cigarettes)":                               "Tobacco",
    "Fabric, fibre and yarn, and leather and dressed furs":                             "Fabrics & leather",
    "Clothing, footwear and accessories":                                               "Clothing & footwear",
    "Carpets, textile furnishings and other textile products":                          "Carpets & textiles",
    "Converted paper products (except for packaging)":                                  "Paper products",
    "Published products and recorded media (except software)":                          "Published media",
    "Software and software licensing":                                                  "Software",
    "Cleaning products and toiletries":                                                 "Cleaning & toiletries",
    "Appliances":                                                                       "Appliances",
    "Miscellaneous goods and supplies":                                                 "Misc. goods",
}

EXP_NSA_OFFSET = 606  # Export NSA vector = Import NSA vector + 606
IMP_SA_OFFSET  = 153  # Import SA  vector = Import NSA vector + 153 (confirmed 2026-03-22)
EXP_SA_OFFSET  = 759  # Export SA  vector = Import NSA vector + 759 (= 606 + 153)

COMM_FLOWS_NSA = [
    ("Import", "Unadjusted", 0),
    ("Export", "Unadjusted", EXP_NSA_OFFSET),
]

COMM_FLOWS_SA = [
    ("Import", "Seasonally adjusted", IMP_SA_OFFSET),
    ("Export", "Seasonally adjusted", EXP_SA_OFFSET),
]

for trade_dir, sa_label, offset in COMM_FLOWS_NSA:
    for d3, d4, d5, imp_nsa_vec in COMMODITY_ROWS:
        vec = imp_nsa_vec + offset
        sl  = DIM_SHORT.get(d5, d5)
        if d3 == d4 == d5:
            fl = f"{trade_dir} ({sa_label[:3]}) — {d3}"
        elif d4 == d5:
            fl = f"{trade_dir} ({sa_label[:3]}) — {d4}"
        else:
            fl = f"{trade_dir} ({sa_label[:3]}) — {d4} / {d5}"
        # 19-element 5-dim tuple
        rows_to_append.append(("Trade", "monthly", "goods_commodity",
                               "Goods by commodity", "12-10-0163-01",
                               "Trade direction", trade_dir,
                               "Seasonal adjustment", sa_label,
                               "Commodity group", d3,
                               "Sub-commodity", d4,
                               "Sub-sub-commodity", d5,
                               str(vec), fl, sl, FILL_GREY))

for trade_dir, sa_label, offset in COMM_FLOWS_SA:
    for d3, d4, d5, imp_nsa_vec in COMMODITY_ROWS:
        vec = imp_nsa_vec + offset
        sl  = DIM_SHORT.get(d5, d5)
        if d3 == d4 == d5:
            fl = f"{trade_dir} (SA) — {d3}"
        elif d4 == d5:
            fl = f"{trade_dir} (SA) — {d4}"
        else:
            fl = f"{trade_dir} (SA) — {d4} / {d5}"
        # 19-element 5-dim tuple
        rows_to_append.append(("Trade", "monthly", "goods_commodity_sa",
                               "Goods by commodity, SA", "12-10-0163-01",
                               "Trade direction", trade_dir,
                               "Seasonal adjustment", sa_label,
                               "Commodity group", d3,
                               "Sub-commodity", d4,
                               "Sub-sub-commodity", d5,
                               str(vec), fl, sl, FILL_GREY))


# ── 5. Goods by commodity, chained 2017$ (12-10-0166-01) ──────────────────
# pid=1210016601 | dims=4, identical structure to goods_commodity
#   dim1 = Trade direction   (Import / Export)
#   dim2 = Seasonal adj.     (always "Seasonally adjusted" → auto-skips)
#   dim3 = Commodity group   (L2 only — no L3 sub-items in this table)
#   dim4 = Sub-commodity     (same as dim3 → auto-adds at L2 step)
# 2 trade directions × 14 commodities = 28 rows

CHAINED_COMMODITIES = [
    # (commodity_name,                                               imp_vec,    short_label)
    ("Total of all merchandise",                                   1566912703, "Total"),
    ("Farm, fishing and intermediate food products",               1566912704, "Farm & fishing"),
    ("Energy products",                                            1566912705, "Energy"),
    ("Metal ores and non-metallic minerals",                       1566912706, "Metal ores"),
    ("Metal and non-metallic mineral products",                    1566912707, "Metal products"),
    ("Basic and industrial chemical, plastic and rubber products", 1566912708, "Chemical, plastic & rubber"),
    ("Forestry products and building and packaging materials",     1566912709, "Forestry & building"),
    ("Industrial machinery, equipment and parts",                  1566912710, "Industrial machinery"),
    ("Electronic and electrical equipment and parts",              1566912711, "Electronic & electrical"),
    ("Motor vehicles and parts",                                   1566912712, "Motor vehicles"),
    ("Aircraft and other transportation equipment and parts",      1566912713, "Aircraft & transportation"),
    ("Consumer goods",                                             1566912714, "Consumer goods"),
    ("Special transactions trade",                                 1566912715, "Special transactions"),
    ("Other balance of payments adjustments",                      1566912716, "Other BoP adjustments"),
]
CHAINED_EXP_OFFSET = 14  # Export vectors immediately follow Import (14 commodities)

for i, (comm, imp_vec, sl) in enumerate(CHAINED_COMMODITIES):
    for trade_dir, offset in [("Import", 0), ("Export", CHAINED_EXP_OFFSET)]:
        vec = imp_vec + offset
        fl  = f"{trade_dir} (SA) — {comm}"
        # dim3==dim4 (same commodity name) → UI auto-adds at the commodity step
        rows_to_append.append(("Trade", "monthly", "goods_commodity_chained",
                               "Goods by commodity, chained 2017$", "12-10-0166-01",
                               "Trade direction", trade_dir,
                               "Seasonal adjustment", "Seasonally adjusted",
                               "Commodity group", comm,
                               "Sub-commodity", comm,
                               str(vec), fl, sl, FILL_GREY))


# ── 6. Goods trade by province, commodity and partner (12-10-0175-01) ─────
# pid=1210017501 | Import, Total of all merchandise, all provinces + Canada
# dim1 = Province (Canada + 10 provinces = 11 values)
# dim2 = Partner (All countries + 6 specific = 7 values)
# Fixed: Total of all merchandise commodity (comm_idx=0)
#
# Vector formula (derived from 3 confirmed data points + structural analysis):
#   PROV_BASE = 1567082903  (Canada Import Total All countries)
#   NL        = PROV_BASE + 1131  (= base + 39×29)  ← confirmed v1567084034
#   Ontario   = PROV_BASE + 4901  (= base + 169×29) ← confirmed v1567087804
#
#   Canada block = 39 groups-of-29  (3 flows × 13 commodities)
#   Province blocks = 26 groups-of-29 each  (2 flows × 13 commodities)
#   Province n (1-based, NL=1…BC=10) offset = 1131 + (n-1) × 754
#
#   Within each block, flow is the outer dimension:
#     flow_offset = flow_idx × 13 × 29  (= flow_idx × 377)
#     Import:          flow_offset = 0
#     Domestic export: flow_offset = 377  (13 commodities × 29 partners)
#     Re-export:       flow_offset = 754  (Canada only; provinces have only 2 flows)
#
#   vector = PROV_BASE + prov_offset + flow_offset + comm_idx×29 + partner_idx
#   (for Total commodity: comm_idx = 0)

PROV_BASE = 1567082903
# (geo_name, short_label, prov_offset)  where Canada offset=0, provinces = 1131+(n-1)*754
PROV_GEOS = [
    ("Canada",                          "Canada", 0),
    ("Newfoundland and Labrador",        "NL",     1131),
    ("Prince Edward Island",             "PEI",    1885),
    ("Nova Scotia",                      "NS",     2639),
    ("New Brunswick",                    "NB",     3393),
    ("Quebec",                           "QC",     4147),
    ("Ontario",                          "ON",     4901),
    ("Manitoba",                         "MB",     5655),
    ("Saskatchewan",                     "SK",     6409),
    ("Alberta",                          "AB",     7163),
    ("British Columbia",                 "BC",     7917),
]
# Province blocks have 2 flows only (Import + Domestic export).
# Canada has 3 (adds Re-export), but using a uniform 2-flow list keeps the
# Province→Flow→Partner cascade consistent across all geographies.
PROV_FLOWS = [
    ("Import",          0,   "Import"),
    ("Domestic export", 377, "Dom. export"),   # 13 commodities × 29 partners
]
PROV_PARTNERS = [
    ("All countries",  0, "All countries"),
    ("United States",  1, "US"),
    ("China",          2, "China"),
    ("Mexico",         3, "Mexico"),
    ("United Kingdom", 4, "UK"),
    ("Japan",          5, "Japan"),
    ("Germany",        6, "Germany"),
]
for geo_full, geo_sl, prov_offset in PROV_GEOS:
    for flow_full, flow_offset, flow_sl in PROV_FLOWS:
        for d3_full, ptnr_idx, d3_sl in PROV_PARTNERS:
            vec = PROV_BASE + prov_offset + flow_offset + ptnr_idx  # commodity_idx=0 (Total)
            fl  = f"{flow_full} — {geo_full} — {d3_full}"
            # 15-element tuple: includes dim3_name and dim3_value
            rows_to_append.append(("Trade", "monthly", "goods_prov_comm_partner",
                                   "Goods trade by province, commodity and partner",
                                   "12-10-0175-01",
                                   "Province", geo_full,
                                   "Flow", flow_full,
                                   "Partner", d3_full,
                                   str(vec), fl, d3_sl, FILL_GREY))


# ── 7. Goods trade price and volume indexes (12-10-0170-01) ───────────────
# pid=1210017001 | dim1=Flow & Adjustment, dim2=Index type
# 4 flow/adj × 2 index types = 8 rows

PRICE_VOL_VECTORS = {
    ("Import (NSA)", "Price index"):  1566921084,
    ("Import (NSA)", "Volume index"): 1566921237,
    ("Import (SA)",  "Price index"):  1566921543,
    ("Import (SA)",  "Volume index"): 1566921696,
    ("Export (NSA)", "Price index"):  1566922902,
    ("Export (NSA)", "Volume index"): 1566923055,
    ("Export (SA)",  "Price index"):  1566923361,
    ("Export (SA)",  "Volume index"): 1566923514,
}
for (flow, idx_type), vec in PRICE_VOL_VECTORS.items():
    sl = idx_type.replace(" index", "")
    rows_to_append.append(("Trade", "monthly", "goods_price_vol_idx",
                           "Goods trade price and volume indexes", "12-10-0170-01",
                           "Flow & Adjustment", flow, "Index type", idx_type,
                           str(vec), f"{flow} — {idx_type}", sl, FILL_GREY))

# ── CATEGORY: Population (white FFFFFF, quarterly) ───────────────────────────

# pop_est: 11 rows
pop_est_rows = [
    ("Canada", 1, "Population estimates: Canada", "Canada"),
    ("Newfoundland and Labrador", 2, "Population estimates: NL", "NL"),
    ("Prince Edward Island", 8, "Population estimates: PEI", "PEI"),
    ("Nova Scotia", 9, "Population estimates: NS", "NS"),
    ("New Brunswick", 10, "Population estimates: NB", "NB"),
    ("Quebec", 11, "Population estimates: QC", "QC"),
    ("Ontario", 12, "Population estimates: ON", "ON"),
    ("Manitoba", 13, "Population estimates: MB", "MB"),
    ("Saskatchewan", 14, "Population estimates: SK", "SK"),
    ("Alberta", 15, "Population estimates: AB", "AB"),
    ("British Columbia", 3, "Population estimates: BC", "BC"),
]
for geo, vec, fl, sl in pop_est_rows:
    rows_to_append.append(("Population", "quarterly", "pop_est", "Population estimates", "17-10-0009-01", "Geography", geo, "", "", str(vec), fl, sl, FILL_WHITE))

# pop_npr: 15 rows
pop_npr_rows = [
    ("Canada", 1566927590, "Non-permanent residents: Canada", "Canada"),
    ("Newfoundland and Labrador", 1566927601, "NPR: NL", "NL"),
    ("Prince Edward Island", 1566927612, "NPR: PEI", "PEI"),
    ("Nova Scotia", 1566927623, "NPR: NS", "NS"),
    ("New Brunswick", 1566927634, "NPR: NB", "NB"),
    ("Quebec", 1566927645, "NPR: QC", "QC"),
    ("Ontario", 1566927656, "NPR: ON", "ON"),
    ("Manitoba", 1566927667, "NPR: MB", "MB"),
    ("Saskatchewan", 1566927678, "NPR: SK", "SK"),
    ("Alberta", 1566927689, "NPR: AB", "AB"),
    ("British Columbia", 1566927700, "NPR: BC", "BC"),
    ("Asylum claimants, protected persons etc.", 1566927591, "NPR: Asylum claimants", "Asylum claimants"),
    ("Work permit holders only", 1566927597, "NPR: Work permit holders", "Work permits"),
    ("Study permit holders only", 1566927598, "NPR: Study permit holders", "Study permits"),
    ("Work and study permit holders", 1566927599, "NPR: Work and study permit holders", "Work & study"),
]
for d1v, vec, fl, sl in pop_npr_rows:
    rows_to_append.append(("Population", "quarterly", "pop_npr", "Non-permanent residents", "17-10-0121-01", "Breakdown", d1v, "", "", str(vec), fl, sl, FILL_WHITE))

# pop_births: 1 row, no dims
rows_to_append.append(("Population", "quarterly", "pop_births", "Births, Canada", "17-10-0059-01", "", "", "", "", "62", "Births, Canada", "Births", FILL_WHITE))

# pop_deaths: 1 row, no dims
rows_to_append.append(("Population", "quarterly", "pop_deaths", "Deaths, Canada", "17-10-0059-01", "", "", "", "", "77", "Deaths, Canada", "Deaths", FILL_WHITE))

# pop_immigrants: 1 row, no dims
rows_to_append.append(("Population", "quarterly", "pop_immigrants", "Immigrants, Canada", "17-10-0040-01", "", "", "", "", "29850342", "Immigrants, Canada", "Immigrants", FILL_WHITE))

# pop_net_npr: 1 row, no dims
rows_to_append.append(("Population", "quarterly", "pop_net_npr", "Net non-permanent residents, Canada", "17-10-0040-01", "", "", "", "", "29850346", "Net non-permanent residents, Canada", "Net NPR", FILL_WHITE))

# pop_npr_in: 1 row, no dims
rows_to_append.append(("Population", "quarterly", "pop_npr_in", "Non-permanent resident inflows, Canada", "17-10-0040-01", "", "", "", "", "1566834758", "Non-permanent resident inflows, Canada", "NPR inflows", FILL_WHITE))

# pop_npr_out: 1 row, no dims
rows_to_append.append(("Population", "quarterly", "pop_npr_out", "Non-permanent resident outflows, Canada", "17-10-0040-01", "", "", "", "", "1566834773", "Non-permanent resident outflows, Canada", "NPR outflows", FILL_WHITE))

# pop_inmig: 11 rows
pop_inmig_rows = [
    ("Canada", 509037, "Interprovincial in-migrants: Canada", "Canada"),
    ("Newfoundland and Labrador", 509038, "Interprovincial in-migrants: NL", "NL"),
    ("Prince Edward Island", 509044, "Interprovincial in-migrants: PEI", "PEI"),
    ("Nova Scotia", 509045, "Interprovincial in-migrants: NS", "NS"),
    ("New Brunswick", 509046, "Interprovincial in-migrants: NB", "NB"),
    ("Quebec", 509047, "Interprovincial in-migrants: QC", "QC"),
    ("Ontario", 509048, "Interprovincial in-migrants: ON", "ON"),
    ("Manitoba", 509049, "Interprovincial in-migrants: MB", "MB"),
    ("Saskatchewan", 509050, "Interprovincial in-migrants: SK", "SK"),
    ("Alberta", 509051, "Interprovincial in-migrants: AB", "AB"),
    ("British Columbia", 509039, "Interprovincial in-migrants: BC", "BC"),
]
for d1v, vec, fl, sl in pop_inmig_rows:
    rows_to_append.append(("Population", "quarterly", "pop_inmig", "Interprovincial in-migrants", "17-10-0020-01", "Geography", d1v, "", "", str(vec), fl, sl, FILL_WHITE))

# pop_outmig: 11 rows
pop_outmig_rows = [
    ("Canada", 509052, "Interprovincial out-migrants: Canada", "Canada"),
    ("Newfoundland and Labrador", 509053, "Interprovincial out-migrants: NL", "NL"),
    ("Prince Edward Island", 509059, "Interprovincial out-migrants: PEI", "PEI"),
    ("Nova Scotia", 509060, "Interprovincial out-migrants: NS", "NS"),
    ("New Brunswick", 509061, "Interprovincial out-migrants: NB", "NB"),
    ("Quebec", 509062, "Interprovincial out-migrants: QC", "QC"),
    ("Ontario", 509063, "Interprovincial out-migrants: ON", "ON"),
    ("Manitoba", 509064, "Interprovincial out-migrants: MB", "MB"),
    ("Saskatchewan", 509065, "Interprovincial out-migrants: SK", "SK"),
    ("Alberta", 509066, "Interprovincial out-migrants: AB", "AB"),
    ("British Columbia", 509054, "Interprovincial out-migrants: BC", "BC"),
]
for d1v, vec, fl, sl in pop_outmig_rows:
    rows_to_append.append(("Population", "quarterly", "pop_outmig", "Interprovincial out-migrants", "17-10-0020-01", "Geography", d1v, "", "", str(vec), fl, sl, FILL_WHITE))

# ── CATEGORY: Travel (grey EEEEEE) ────────────────────────────────────────────
travel_rows = [
    ("daily", "trv_car_lead", "Canadian residents returning from US by car (leading indicator)", "24-10-0054-01", "", "", "", "", "1545910054", "Canadian residents returning from the U.S. by car (leading indicator)", "Returning by car (leading)"),
    ("daily", "trv_air_lead", "Canadian residents returning from US by air (leading indicator)", "24-10-0054-01", "", "", "", "", "1324883057", "Canadian residents returning from the U.S. by air (leading indicator)", "Returning by air (leading)"),
    ("monthly", "trv_us_final", "Canadian residents returning from US (final)", "24-10-0054-01", "", "", "", "", "1296956586", "Canadian residents returning from the U.S. (final)", "Returning from US (final)"),
    ("monthly", "trv_nonus_final", "Canadian residents returning from non-US countries (final)", "24-10-0054-01", "", "", "", "", "1296956631", "Canadian residents returning from non-U.S. countries (final)", "Returning from non-US (final)"),
    ("monthly", "trv_us_in", "US residents entering Canada (final)", "24-10-0054-01", "", "", "", "", "1296956469", "U.S. residents entering Canada (final)", "US residents entering (final)"),
    ("monthly", "trv_nonus_in", "Residents from non-US countries entering Canada (final)", "24-10-0054-01", "", "", "", "", "1296956514", "Residents from non-U.S. countries entering Canada (final)", "Non-US residents entering (final)"),
]
for freq, sid, sname, tid, d1n, d1v, d2n, d2v, vec, fl, sl in travel_rows:
    rows_to_append.append(("Travel", freq, sid, sname, tid, d1n, d1v, d2n, d2v, vec, fl, sl, FILL_GREY))

# ── CATEGORY: National balance sheets (white FFFFFF, quarterly) ───────────────

# hh_fin_ind: 4 rows
hh_fin_rows = [
    ("Debt to gross domestic product (GDP)", 62698062, "Household debt to GDP", "Debt to GDP"),
    ("Credit market debt to disposable income", 62698064, "Household credit market debt to disposable income", "Debt to income"),
    ("Debt to total assets", 62698067, "Household debt to total assets", "Debt to assets"),
    ("Debt to net worth", 62698068, "Household debt to net worth", "Debt to net worth"),
]
for d1v, vec, fl, sl in hh_fin_rows:
    rows_to_append.append(("National balance sheets", "quarterly", "hh_fin_ind", "Financial indicators of households", "38-10-0235-01", "Indicator", d1v, "", "", str(vec), fl, sl, FILL_WHITE))

# net_worth: 4 rows
net_worth_rows = [
    ("Total", 62693792, "Net worth: Total, national balance sheets", "Total"),
    ("Households and non-profit institutions serving households", 62693897, "Net worth: Households and non-profit institutions", "Households"),
    ("Corporations", 62694212, "Net worth: Corporations", "Corporations"),
    ("General governments", 62694737, "Net worth: General governments", "Governments"),
]
for d1v, vec, fl, sl in net_worth_rows:
    rows_to_append.append(("National balance sheets", "quarterly", "net_worth", "Net worth by sector", "36-10-0580-01", "Sector", d1v, "", "", str(vec), fl, sl, FILL_WHITE))

# hh_debt_svc: Debt service indicators of households (11-10-0065-01)
for label, vec, short in [
    ("Household income",                          1001696792, "Household income"),
    ("Compensation of employees",                 1001696793, "Compensation of employees"),
    ("Total debt payments",                       1001696804, "Total debt payments"),
    ("Total obligated payments of principal",     1001696807, "Obligated principal pmts"),
    ("Total interest paid",                       1001696810, "Total interest paid"),
    ("Debt service ratio",                        1001696813, "Debt service ratio"),
    ("Debt service ratio, interest only",         1001696816, "DSR, interest only"),
]:
    rows_to_append.append(("National balance sheets", "quarterly", "hh_debt_svc",
                           "Debt service indicators of households", "11-10-0065-01",
                           "Indicator", label, "", "", str(vec), label, short, FILL_WHITE))

# ── CATEGORY: GDP (expenditure based) additions ───────────────────────────────

# hh_accounts_sa: Current and capital accounts – Households, SA (36-10-0112-01)
for label, vec, short in [
    ("Compensation of employees", 62305952, "Compensation of employees"),
    ("Household disposable income", 62305981, "Disposable income"),
    ("Household saving rate",       62305984, "Saving rate"),
]:
    rows_to_append.append(("GDP (expenditure based)", "quarterly", "hh_accounts_sa",
                           "Current and capital accounts \u2013 Households, SA", "36-10-0112-01",
                           "Account", label, "", "", str(vec), label, short, FILL_WHITE))

# ── CATEGORY: Housing and households ─────────────────────────────────────────

# bldg_permits: Building permits, SA (34-10-0292-01)
for label, vec, short in [
    ("Total (residential and non-residential)", "1675119645", "Total"),
    ("Residential",                             "1675119646", "Residential"),
]:
    rows_to_append.append(("Housing and households", "monthly", "bldg_permits",
                           "Building permits, SA", "34-10-0292-01",
                           "Type", label, "", "", vec, label, short, FILL_WHITE))

# hh_credit: Credit liabilities of households (36-10-0639-01)
for seasonality, credit_type, vec, short in [
    ("Raw data",            "Non-mortgage loans",                          "v1231415568", "Non-mortgage loans"),
    ("Raw data",            "Personal loans",                              "v1231415570", "Personal loans"),
    ("Raw data",            "Credit cards",                                "v1231415571", "Credit cards"),
    ("Raw data",            "Lines of credit",                             "v1231415572", "Lines of credit"),
    ("Raw data",            "Other personal loans",                        "v1231415574", "Other personal loans"),
    ("Raw data",            "Unincorporated business",                     "v1231415575", "Unincorporated business"),
    ("Raw data",            "Mortgage loans",                              "v1231415577", "Mortgage loans"),
    ("Raw data",            "Residential mortgages",                       "v1231415578", "Residential mortgages"),
    ("Raw data",            "Non-residential mortgages",                   "v1231415581", "Non-residential mortgages"),
    ("Raw data",            "Total credit liabilities",                    "v1231415582", "Total credit liabilities"),
    ("Raw data",            "Real estate secured lending",                 "v1231415583", "Real estate secured lending"),
    ("Raw data",            "Real estate: Residential mortgages",          "v1231415584", "RE: Residential mortgages"),
    ("Raw data",            "Real estate: Non-residential mortgages",      "v1231415587", "RE: Non-residential mortgages"),
    ("Raw data",            "Home equity lines of credit",                 "v1231415590", "Home equity lines of credit"),
    ("Seasonally adjusted", "Non-mortgage loans",                          "v1231415611", "Non-mortgage loans"),
    ("Seasonally adjusted", "Personal loans",                              "v1231415613", "Personal loans"),
    ("Seasonally adjusted", "Credit cards",                                "v1231415614", "Credit cards"),
    ("Seasonally adjusted", "Lines of credit",                             "v1231415615", "Lines of credit"),
    ("Seasonally adjusted", "Other personal loans",                        "v1231415617", "Other personal loans"),
    ("Seasonally adjusted", "Unincorporated business",                     "v1231415618", "Unincorporated business"),
    ("Seasonally adjusted", "Mortgage loans",                              "v1231415620", "Mortgage loans"),
    ("Seasonally adjusted", "Residential mortgages",                       "v1231415621", "Residential mortgages"),
    ("Seasonally adjusted", "Non-residential mortgages",                   "v1231415624", "Non-residential mortgages"),
    ("Seasonally adjusted", "Total credit liabilities",                    "v1231415625", "Total credit liabilities"),
    ("Seasonally adjusted", "Real estate secured lending",                 "v1231415626", "Real estate secured lending"),
    ("Seasonally adjusted", "Real estate: Residential mortgages",          "v1231415627", "RE: Residential mortgages"),
    ("Seasonally adjusted", "Real estate: Non-residential mortgages",      "v1231415630", "RE: Non-residential mortgages"),
    ("Seasonally adjusted", "Home equity lines of credit",                 "v1231415633", "Home equity lines of credit"),
]:
    fl = f"{seasonality} \u2014 {credit_type}"
    rows_to_append.append(("Housing and households", "monthly", "hh_credit",
                           "Credit liabilities of households", "36-10-0639-01",
                           "Seasonality", seasonality, "Credit type", credit_type,
                           vec, fl, short, FILL_WHITE))

# housing_cmhc: Housing starts, under construction & completions (34-10-0135-01)
for geo, measure, vec, short in [
    ("Canada",                       "Housing starts - Total units",              730416,  "Starts (total)"),
    ("Canada",                       "Housing starts - Single-detached",          730442,  "Starts (single)"),
    ("Canada",                       "Housing starts - Multiples",                730486,  "Starts (multiples)"),
    ("Canada",                       "Housing under construction - Total units",  731381,  "UC (total)"),
    ("Canada",                       "Housing under construction - Single-detached", 731392, "UC (single)"),
    ("Canada",                       "Housing under construction - Multiples",    731436,  "UC (multiples)"),
    ("Canada",                       "Housing completions - Total units",         732318,  "Completions (total)"),
    ("Canada",                       "Housing completions - Single-detached",     732329,  "Completions (single)"),
    ("Canada",                       "Housing completions - Multiples",           732373,  "Completions (multiples)"),
    ("Newfoundland and Labrador",     "Housing starts - Total units",              730417,  "Starts (total)"),
    ("Newfoundland and Labrador",     "Housing starts - Single-detached",          730443,  "Starts (single)"),
    ("Newfoundland and Labrador",     "Housing starts - Multiples",                730487,  "Starts (multiples)"),
    ("Newfoundland and Labrador",     "Housing under construction - Total units",  731382,  "UC (total)"),
    ("Newfoundland and Labrador",     "Housing under construction - Single-detached", 731393, "UC (single)"),
    ("Newfoundland and Labrador",     "Housing under construction - Multiples",    731437,  "UC (multiples)"),
    ("Newfoundland and Labrador",     "Housing completions - Total units",         732319,  "Completions (total)"),
    ("Newfoundland and Labrador",     "Housing completions - Single-detached",     732330,  "Completions (single)"),
    ("Newfoundland and Labrador",     "Housing completions - Multiples",           732374,  "Completions (multiples)"),
    ("Prince Edward Island",          "Housing starts - Total units",              730419,  "Starts (total)"),
    ("Prince Edward Island",          "Housing starts - Single-detached",          730445,  "Starts (single)"),
    ("Prince Edward Island",          "Housing starts - Multiples",                730489,  "Starts (multiples)"),
    ("Prince Edward Island",          "Housing under construction - Total units",  731384,  "UC (total)"),
    ("Prince Edward Island",          "Housing under construction - Single-detached", 731395, "UC (single)"),
    ("Prince Edward Island",          "Housing under construction - Multiples",    731439,  "UC (multiples)"),
    ("Prince Edward Island",          "Housing completions - Total units",         732321,  "Completions (total)"),
    ("Prince Edward Island",          "Housing completions - Single-detached",     732332,  "Completions (single)"),
    ("Prince Edward Island",          "Housing completions - Multiples",           732376,  "Completions (multiples)"),
    ("Nova Scotia",                   "Housing starts - Total units",              730420,  "Starts (total)"),
    ("Nova Scotia",                   "Housing starts - Single-detached",          730446,  "Starts (single)"),
    ("Nova Scotia",                   "Housing starts - Multiples",                730490,  "Starts (multiples)"),
    ("Nova Scotia",                   "Housing under construction - Total units",  731385,  "UC (total)"),
    ("Nova Scotia",                   "Housing under construction - Single-detached", 731396, "UC (single)"),
    ("Nova Scotia",                   "Housing under construction - Multiples",    731440,  "UC (multiples)"),
    ("Nova Scotia",                   "Housing completions - Total units",         732322,  "Completions (total)"),
    ("Nova Scotia",                   "Housing completions - Single-detached",     732333,  "Completions (single)"),
    ("Nova Scotia",                   "Housing completions - Multiples",           732377,  "Completions (multiples)"),
    ("New Brunswick",                 "Housing starts - Total units",              730421,  "Starts (total)"),
    ("New Brunswick",                 "Housing starts - Single-detached",          730447,  "Starts (single)"),
    ("New Brunswick",                 "Housing starts - Multiples",                730491,  "Starts (multiples)"),
    ("New Brunswick",                 "Housing under construction - Total units",  731386,  "UC (total)"),
    ("New Brunswick",                 "Housing under construction - Single-detached", 731397, "UC (single)"),
    ("New Brunswick",                 "Housing under construction - Multiples",    731441,  "UC (multiples)"),
    ("New Brunswick",                 "Housing completions - Total units",         732323,  "Completions (total)"),
    ("New Brunswick",                 "Housing completions - Single-detached",     732334,  "Completions (single)"),
    ("New Brunswick",                 "Housing completions - Multiples",           732378,  "Completions (multiples)"),
    ("Quebec",                        "Housing starts - Total units",              730422,  "Starts (total)"),
    ("Quebec",                        "Housing starts - Single-detached",          730448,  "Starts (single)"),
    ("Quebec",                        "Housing starts - Multiples",                730492,  "Starts (multiples)"),
    ("Quebec",                        "Housing under construction - Total units",  731387,  "UC (total)"),
    ("Quebec",                        "Housing under construction - Single-detached", 731398, "UC (single)"),
    ("Quebec",                        "Housing under construction - Multiples",    731442,  "UC (multiples)"),
    ("Quebec",                        "Housing completions - Total units",         732324,  "Completions (total)"),
    ("Quebec",                        "Housing completions - Single-detached",     732335,  "Completions (single)"),
    ("Quebec",                        "Housing completions - Multiples",           732379,  "Completions (multiples)"),
    ("Ontario",                       "Housing starts - Total units",              730423,  "Starts (total)"),
    ("Ontario",                       "Housing starts - Single-detached",          730449,  "Starts (single)"),
    ("Ontario",                       "Housing starts - Multiples",                730493,  "Starts (multiples)"),
    ("Ontario",                       "Housing under construction - Total units",  731388,  "UC (total)"),
    ("Ontario",                       "Housing under construction - Single-detached", 731399, "UC (single)"),
    ("Ontario",                       "Housing under construction - Multiples",    731443,  "UC (multiples)"),
    ("Ontario",                       "Housing completions - Total units",         732325,  "Completions (total)"),
    ("Ontario",                       "Housing completions - Single-detached",     732336,  "Completions (single)"),
    ("Ontario",                       "Housing completions - Multiples",           732380,  "Completions (multiples)"),
    ("Manitoba",                      "Housing starts - Total units",              730424,  "Starts (total)"),
    ("Manitoba",                      "Housing starts - Single-detached",          730450,  "Starts (single)"),
    ("Manitoba",                      "Housing starts - Multiples",                730494,  "Starts (multiples)"),
    ("Manitoba",                      "Housing under construction - Total units",  731389,  "UC (total)"),
    ("Manitoba",                      "Housing under construction - Single-detached", 731400, "UC (single)"),
    ("Manitoba",                      "Housing under construction - Multiples",    731444,  "UC (multiples)"),
    ("Manitoba",                      "Housing completions - Total units",         732326,  "Completions (total)"),
    ("Manitoba",                      "Housing completions - Single-detached",     732337,  "Completions (single)"),
    ("Manitoba",                      "Housing completions - Multiples",           732381,  "Completions (multiples)"),
    ("Saskatchewan",                  "Housing starts - Total units",              730425,  "Starts (total)"),
    ("Saskatchewan",                  "Housing starts - Single-detached",          730451,  "Starts (single)"),
    ("Saskatchewan",                  "Housing starts - Multiples",                730495,  "Starts (multiples)"),
    ("Saskatchewan",                  "Housing under construction - Total units",  731390,  "UC (total)"),
    ("Saskatchewan",                  "Housing under construction - Single-detached", 731401, "UC (single)"),
    ("Saskatchewan",                  "Housing under construction - Multiples",    731445,  "UC (multiples)"),
    ("Saskatchewan",                  "Housing completions - Total units",         732327,  "Completions (total)"),
    ("Saskatchewan",                  "Housing completions - Single-detached",     732338,  "Completions (single)"),
    ("Saskatchewan",                  "Housing completions - Multiples",           732382,  "Completions (multiples)"),
    ("Alberta",                       "Housing starts - Total units",              730426,  "Starts (total)"),
    ("Alberta",                       "Housing starts - Single-detached",          730452,  "Starts (single)"),
    ("Alberta",                       "Housing starts - Multiples",                730496,  "Starts (multiples)"),
    ("Alberta",                       "Housing under construction - Total units",  731391,  "UC (total)"),
    ("Alberta",                       "Housing under construction - Single-detached", 731402, "UC (single)"),
    ("Alberta",                       "Housing under construction - Multiples",    731446,  "UC (multiples)"),
    ("Alberta",                       "Housing completions - Total units",         732328,  "Completions (total)"),
    ("Alberta",                       "Housing completions - Single-detached",     732339,  "Completions (single)"),
    ("Alberta",                       "Housing completions - Multiples",           732383,  "Completions (multiples)"),
    ("British Columbia",              "Housing starts - Total units",              730418,  "Starts (total)"),
    ("British Columbia",              "Housing starts - Single-detached",          730444,  "Starts (single)"),
    ("British Columbia",              "Housing starts - Multiples",                730488,  "Starts (multiples)"),
    ("British Columbia",              "Housing under construction - Total units",  731383,  "UC (total)"),
    ("British Columbia",              "Housing under construction - Single-detached", 731394, "UC (single)"),
    ("British Columbia",              "Housing under construction - Multiples",    731438,  "UC (multiples)"),
    ("British Columbia",              "Housing completions - Total units",         732320,  "Completions (total)"),
    ("British Columbia",              "Housing completions - Single-detached",     732331,  "Completions (single)"),
    ("British Columbia",              "Housing completions - Multiples",           732375,  "Completions (multiples)"),
]:
    fl = f"{geo} \u2014 {measure}"
    rows_to_append.append(("Housing and households", "quarterly", "housing_cmhc",
                           "Housing starts, under construction & completions", "34-10-0135-01",
                           "Geography", geo, "Measure", measure,
                           str(vec), fl, short, FILL_WHITE))

# ── CATEGORY: International securities ───────────────────────────────────────

# int_sec: International transactions in securities (36-10-0028-01)
for instrument, transaction, vec, short in [
    ("Canadian securities",                                                        "Net flows",   "v61915649",  "Net flows"),
    ("Canadian securities",                                                        "Sales",       "v61915650",  "Sales"),
    ("Canadian securities",                                                        "Purchases",   "v61915651",  "Purchases"),
    ("Canadian debt securities",                                                   "Net flows",   "v61915652",  "Net flows"),
    ("Canadian debt securities",                                                   "Sales",       "v61915653",  "Sales"),
    ("Canadian debt securities",                                                   "Purchases",   "v61915654",  "Purchases"),
    ("Canadian money market instruments",                                          "Net flows",   "v61915655",  "Net flows"),
    ("Canadian money market instruments",                                          "Sales",       "v61915656",  "Sales"),
    ("Canadian money market instruments",                                          "Purchases",   "v61915657",  "Purchases"),
    ("Canadian money market instruments, governments",                             "Net flows",   "v61915658",  "Net flows"),
    ("Canadian money market instruments, governments",                             "Sales",       "v61915659",  "Sales"),
    ("Canadian money market instruments, governments",                             "Purchases",   "v61915660",  "Purchases"),
    ("Canadian money market instruments, federal government",                      "Net flows",   "v61915661",  "Net flows"),
    ("Canadian money market instruments, federal government",                      "Sales",       "v61915662",  "Sales"),
    ("Canadian money market instruments, federal government",                      "Purchases",   "v61915663",  "Purchases"),
    ("Canadian money market instruments, provincial governments",                  "Net flows",   "v61915664",  "Net flows"),
    ("Canadian money market instruments, provincial governments",                  "Sales",       "v61915665",  "Sales"),
    ("Canadian money market instruments, provincial governments",                  "Purchases",   "v61915666",  "Purchases"),
    ("Canadian money market instruments, corporations",                            "Net flows",   "v61915667",  "Net flows"),
    ("Canadian money market instruments, corporations",                            "Sales",       "v61915668",  "Sales"),
    ("Canadian money market instruments, corporations",                            "Purchases",   "v61915669",  "Purchases"),
    ("Canadian money market instruments, government business enterprises",         "Net flows",   "v61915670",  "Net flows"),
    ("Canadian money market instruments, government business enterprises",         "Sales",       "v61915671",  "Sales"),
    ("Canadian money market instruments, government business enterprises",         "Purchases",   "v61915672",  "Purchases"),
    ("Canadian money market instruments, federal government enterprises",          "Net flows",   "v61915673",  "Net flows"),
    ("Canadian money market instruments, federal government enterprises",          "Sales",       "v61915674",  "Sales"),
    ("Canadian money market instruments, federal government enterprises",          "Purchases",   "v61915675",  "Purchases"),
    ("Canadian money market instruments, provincial government enterprises",       "Net flows",   "v61915676",  "Net flows"),
    ("Canadian money market instruments, provincial government enterprises",       "Sales",       "v61915677",  "Sales"),
    ("Canadian money market instruments, provincial government enterprises",       "Purchases",   "v61915678",  "Purchases"),
    ("Canadian money market instruments, private corporations",                    "Net flows",   "v61915679",  "Net flows"),
    ("Canadian money market instruments, private corporations",                    "Sales",       "v61915680",  "Sales"),
    ("Canadian money market instruments, private corporations",                    "Purchases",   "v61915681",  "Purchases"),
    ("Canadian bonds",                                                             "Net flows",   "v61915682",  "Net flows"),
    ("Canadian bonds",                                                             "Sales",       "v61915683",  "Sales"),
    ("Canadian bonds",                                                             "Purchases",   "v61915684",  "Purchases"),
    ("Canadian bonds, governments",                                                "Net flows",   "v61915685",  "Net flows"),
    ("Canadian bonds, governments",                                                "Sales",       "v61915686",  "Sales"),
    ("Canadian bonds, governments",                                                "Purchases",   "v61915687",  "Purchases"),
    ("Canadian bonds, federal government",                                         "Net flows",   "v61915688",  "Net flows"),
    ("Canadian bonds, federal government",                                         "Sales",       "v61915689",  "Sales"),
    ("Canadian bonds, federal government",                                         "Purchases",   "v61915690",  "Purchases"),
    ("Canadian bonds, provincial governments",                                     "Net flows",   "v61915691",  "Net flows"),
    ("Canadian bonds, provincial governments",                                     "Sales",       "v61915692",  "Sales"),
    ("Canadian bonds, provincial governments",                                     "Purchases",   "v61915693",  "Purchases"),
    ("Canadian bonds, municipal governments",                                      "Net flows",   "v61915694",  "Net flows"),
    ("Canadian bonds, municipal governments",                                      "Sales",       "v61915695",  "Sales"),
    ("Canadian bonds, municipal governments",                                      "Purchases",   "v61915696",  "Purchases"),
    ("Canadian bonds, corporations",                                               "Net flows",   "v61915697",  "Net flows"),
    ("Canadian bonds, corporations",                                               "Sales",       "v61915698",  "Sales"),
    ("Canadian bonds, corporations",                                               "Purchases",   "v61915699",  "Purchases"),
    ("Canadian bonds, government business enterprises",                            "Net flows",   "v61915700",  "Net flows"),
    ("Canadian bonds, government business enterprises",                            "Sales",       "v61915701",  "Sales"),
    ("Canadian bonds, government business enterprises",                            "Purchases",   "v61915702",  "Purchases"),
    ("Canadian bonds, federal government enterprises",                             "Net flows",   "v61915703",  "Net flows"),
    ("Canadian bonds, federal government enterprises",                             "Sales",       "v61915704",  "Sales"),
    ("Canadian bonds, federal government enterprises",                             "Purchases",   "v61915705",  "Purchases"),
    ("Canadian bonds, provincial government enterprises",                          "Net flows",   "v61915706",  "Net flows"),
    ("Canadian bonds, provincial government enterprises",                          "Sales",       "v61915707",  "Sales"),
    ("Canadian bonds, provincial government enterprises",                          "Purchases",   "v61915708",  "Purchases"),
    ("Canadian bonds, private corporations",                                       "Net flows",   "v61915709",  "Net flows"),
    ("Canadian bonds, private corporations",                                       "Sales",       "v61915710",  "Sales"),
    ("Canadian bonds, private corporations",                                       "Purchases",   "v61915711",  "Purchases"),
    ("Canadian equity and investment fund shares",                                 "Net flows",   "v61915712",  "Net flows"),
    ("Canadian equity and investment fund shares",                                 "Sales",       "v61915713",  "Sales"),
    ("Canadian equity and investment fund shares",                                 "Purchases",   "v61915714",  "Purchases"),
    ("Foreign securities",                                                         "Net flows",   "v61915715",  "Net flows"),
    ("Foreign securities",                                                         "Sales",       "v61915716",  "Sales"),
    ("Foreign securities",                                                         "Purchases",   "v61915717",  "Purchases"),
    ("Foreign debt securities",                                                    "Net flows",   "v61915718",  "Net flows"),
    ("Foreign debt securities",                                                    "Sales",       "v61915719",  "Sales"),
    ("Foreign debt securities",                                                    "Purchases",   "v61915720",  "Purchases"),
    ("Foreign money market instruments",                                           "Net flows",   "v61915721",  "Net flows"),
    ("Foreign money market instruments",                                           "Sales",       "v61915722",  "Sales"),
    ("Foreign money market instruments",                                           "Purchases",   "v61915723",  "Purchases"),
    ("Foreign money market instruments, United States government",                 "Net flows",   "v61915724",  "Net flows"),
    ("Foreign money market instruments, United States government",                 "Sales",       "v61915725",  "Sales"),
    ("Foreign money market instruments, United States government",                 "Purchases",   "v61915726",  "Purchases"),
    ("Foreign money market instruments, all other United States issuers",          "Net flows",   "v61915727",  "Net flows"),
    ("Foreign money market instruments, all other United States issuers",          "Sales",       "v61915728",  "Sales"),
    ("Foreign money market instruments, all other United States issuers",          "Purchases",   "v61915729",  "Purchases"),
    ("Foreign money market instruments, non-United States foreign issuers",        "Net flows",   "v61915730",  "Net flows"),
    ("Foreign money market instruments, non-United States foreign issuers",        "Sales",       "v61915731",  "Sales"),
    ("Foreign money market instruments, non-United States foreign issuers",        "Purchases",   "v61915732",  "Purchases"),
    ("Foreign bonds",                                                              "Net flows",   "v61915733",  "Net flows"),
    ("Foreign bonds",                                                              "Sales",       "v61915734",  "Sales"),
    ("Foreign bonds",                                                              "Purchases",   "v61915735",  "Purchases"),
    ("Foreign bonds, United States government",                                    "Net flows",   "v61915736",  "Net flows"),
    ("Foreign bonds, United States government",                                    "Sales",       "v61915737",  "Sales"),
    ("Foreign bonds, United States government",                                    "Purchases",   "v61915738",  "Purchases"),
    ("Foreign bonds, all other United States issuers",                             "Net flows",   "v61915739",  "Net flows"),
    ("Foreign bonds, all other United States issuers",                             "Sales",       "v61915740",  "Sales"),
    ("Foreign bonds, all other United States issuers",                             "Purchases",   "v61915741",  "Purchases"),
    ("Foreign bonds, non-United States foreign issuers",                           "Net flows",   "v61915742",  "Net flows"),
    ("Foreign bonds, non-United States foreign issuers",                           "Sales",       "v61915743",  "Sales"),
    ("Foreign bonds, non-United States foreign issuers",                           "Purchases",   "v61915744",  "Purchases"),
    ("Foreign equity and investment fund shares",                                  "Net flows",   "v61915745",  "Net flows"),
    ("Foreign equity and investment fund shares",                                  "Sales",       "v61915746",  "Sales"),
    ("Foreign equity and investment fund shares",                                  "Purchases",   "v61915747",  "Purchases"),
    ("Foreign equity and investment fund shares, United States issuers",           "Net flows",   "v61915748",  "Net flows"),
    ("Foreign equity and investment fund shares, United States issuers",           "Sales",       "v61915749",  "Sales"),
    ("Foreign equity and investment fund shares, United States issuers",           "Purchases",   "v61915750",  "Purchases"),
    ("Foreign equity and investment fund shares, non-United States foreign issuers", "Net flows", "v61915751",  "Net flows"),
    ("Foreign equity and investment fund shares, non-United States foreign issuers", "Sales",     "v61915752",  "Sales"),
    ("Foreign equity and investment fund shares, non-United States foreign issuers", "Purchases", "v61915753",  "Purchases"),
]:
    fl = f"{instrument} \u2014 {transaction}"
    rows_to_append.append(("International securities", "monthly", "int_sec",
                           "International transactions in securities", "36-10-0028-01",
                           "Instrument & issuer", instrument, "Transaction", transaction,
                           vec, fl, short, FILL_WHITE))

# ── CATEGORY: Business and industry ──────────────────────────────────────────

# biz_openclose: Canadian business counts (33-10-0270-01)
for industry, measure, vec, short in [
    ("Business sector industries",                                          "Active businesses",   "v1203704156", "Active businesses"),
    ("Business sector industries",                                          "Opening businesses",  "v1203704157", "Opening businesses"),
    ("Business sector industries",                                          "Closing businesses",  "v1203704159", "Closing businesses"),
    ("Forestry, fishing and hunting",                                       "Active businesses",   "v1203704160", "Active businesses"),
    ("Forestry, fishing and hunting",                                       "Opening businesses",  "v1203704161", "Opening businesses"),
    ("Forestry, fishing and hunting",                                       "Closing businesses",  "v1203704163", "Closing businesses"),
    ("Mining, quarrying, and oil and gas extraction",                       "Active businesses",   "v1203704164", "Active businesses"),
    ("Mining, quarrying, and oil and gas extraction",                       "Opening businesses",  "v1203704165", "Opening businesses"),
    ("Mining, quarrying, and oil and gas extraction",                       "Closing businesses",  "v1203704167", "Closing businesses"),
    ("Utilities",                                                           "Active businesses",   "v1203704168", "Active businesses"),
    ("Utilities",                                                           "Opening businesses",  "v1203704169", "Opening businesses"),
    ("Utilities",                                                           "Closing businesses",  "v1203704171", "Closing businesses"),
    ("Construction",                                                        "Active businesses",   "v1203704172", "Active businesses"),
    ("Construction",                                                        "Opening businesses",  "v1203704173", "Opening businesses"),
    ("Construction",                                                        "Closing businesses",  "v1203704175", "Closing businesses"),
    ("Manufacturing",                                                       "Active businesses",   "v1203704176", "Active businesses"),
    ("Manufacturing",                                                       "Opening businesses",  "v1203704177", "Opening businesses"),
    ("Manufacturing",                                                       "Closing businesses",  "v1203704179", "Closing businesses"),
    ("Food manufacturing",                                                  "Active businesses",   "v1247780270", "Active businesses"),
    ("Food manufacturing",                                                  "Opening businesses",  "v1247780271", "Opening businesses"),
    ("Food manufacturing",                                                  "Closing businesses",  "v1247780273", "Closing businesses"),
    ("Beverage and tobacco product manufacturing",                          "Active businesses",   "v1247780466", "Active businesses"),
    ("Beverage and tobacco product manufacturing",                          "Opening businesses",  "v1247780467", "Opening businesses"),
    ("Beverage and tobacco product manufacturing",                          "Closing businesses",  "v1247780469", "Closing businesses"),
    ("Wholesale trade",                                                     "Active businesses",   "v1203704180", "Active businesses"),
    ("Wholesale trade",                                                     "Opening businesses",  "v1203704181", "Opening businesses"),
    ("Wholesale trade",                                                     "Closing businesses",  "v1203704183", "Closing businesses"),
    ("Retail trade",                                                        "Active businesses",   "v1203704184", "Active businesses"),
    ("Retail trade",                                                        "Opening businesses",  "v1203704185", "Opening businesses"),
    ("Retail trade",                                                        "Closing businesses",  "v1203704187", "Closing businesses"),
    ("Transportation and warehousing",                                      "Active businesses",   "v1203704188", "Active businesses"),
    ("Transportation and warehousing",                                      "Opening businesses",  "v1203704189", "Opening businesses"),
    ("Transportation and warehousing",                                      "Closing businesses",  "v1203704191", "Closing businesses"),
    ("Information and cultural industries",                                 "Active businesses",   "v1203704192", "Active businesses"),
    ("Information and cultural industries",                                 "Opening businesses",  "v1203704193", "Opening businesses"),
    ("Information and cultural industries",                                 "Closing businesses",  "v1203704195", "Closing businesses"),
    ("Finance, insurance and management of companies",                      "Active businesses",   "v1203704196", "Active businesses"),
    ("Finance, insurance and management of companies",                      "Opening businesses",  "v1203704197", "Opening businesses"),
    ("Finance, insurance and management of companies",                      "Closing businesses",  "v1203704199", "Closing businesses"),
    ("Real estate and rental and leasing",                                  "Active businesses",   "v1203704200", "Active businesses"),
    ("Real estate and rental and leasing",                                  "Opening businesses",  "v1203704201", "Opening businesses"),
    ("Real estate and rental and leasing",                                  "Closing businesses",  "v1203704203", "Closing businesses"),
    ("Professional, scientific and technical services",                     "Active businesses",   "v1203704204", "Active businesses"),
    ("Professional, scientific and technical services",                     "Opening businesses",  "v1203704205", "Opening businesses"),
    ("Professional, scientific and technical services",                     "Closing businesses",  "v1203704207", "Closing businesses"),
    ("Administrative and support, waste management and remediation services", "Active businesses", "v1203704208", "Active businesses"),
    ("Administrative and support, waste management and remediation services", "Opening businesses","v1203704209", "Opening businesses"),
    ("Administrative and support, waste management and remediation services", "Closing businesses","v1203704211", "Closing businesses"),
    ("Educational services",                                                "Active businesses",   "v1203704212", "Active businesses"),
    ("Educational services",                                                "Opening businesses",  "v1203704213", "Opening businesses"),
    ("Educational services",                                                "Closing businesses",  "v1203704215", "Closing businesses"),
    ("Health care and social assistance",                                   "Active businesses",   "v1203704216", "Active businesses"),
    ("Health care and social assistance",                                   "Opening businesses",  "v1203704217", "Opening businesses"),
    ("Health care and social assistance",                                   "Closing businesses",  "v1203704219", "Closing businesses"),
    ("Arts, entertainment and recreation",                                  "Active businesses",   "v1203704220", "Active businesses"),
    ("Arts, entertainment and recreation",                                  "Opening businesses",  "v1203704221", "Opening businesses"),
    ("Arts, entertainment and recreation",                                  "Closing businesses",  "v1203704223", "Closing businesses"),
    ("Accommodation and food services",                                     "Active businesses",   "v1203704224", "Active businesses"),
    ("Accommodation and food services",                                     "Opening businesses",  "v1203704225", "Opening businesses"),
    ("Accommodation and food services",                                     "Closing businesses",  "v1203704227", "Closing businesses"),
    ("Other services (except public administration)",                       "Active businesses",   "v1203704228", "Active businesses"),
    ("Other services (except public administration)",                       "Opening businesses",  "v1203704229", "Opening businesses"),
    ("Other services (except public administration)",                       "Closing businesses",  "v1203704231", "Closing businesses"),
    ("Tourism industry",                                                    "Active businesses",   "v1231413605", "Active businesses"),
    ("Tourism industry",                                                    "Opening businesses",  "v1231413606", "Opening businesses"),
    ("Tourism industry",                                                    "Closing businesses",  "v1231413608", "Closing businesses"),
    ("Business sector (excl. educational services and health care)",        "Active businesses",   "v1247788765", "Active businesses"),
    ("Business sector (excl. educational services and health care)",        "Opening businesses",  "v1247788766", "Opening businesses"),
    ("Business sector (excl. educational services and health care)",        "Closing businesses",  "v1247788768", "Closing businesses"),
]:
    fl = f"{industry} \u2014 {measure}"
    rows_to_append.append(("Business and industry", "monthly", "biz_openclose",
                           "Canadian business counts", "33-10-0270-01",
                           "Industry", industry, "Measure", measure,
                           vec, fl, short, FILL_WHITE))

# ── CATEGORY: Bank of Canada (BoC Valet API, table_id='BOC') ─────────────────

# boc_policy_rate: BoC overnight rate target (dims=0)
rows_to_append.append(("Bank of Canada", "monthly", "boc_policy_rate",
                       "Overnight rate target", "BOC",
                       "", "", "", "", "V39079", "Overnight rate target", "Policy rate", FILL_WHITE))

# boc_bank_rate: BoC bank rate (dims=0)
rows_to_append.append(("Bank of Canada", "monthly", "boc_bank_rate",
                       "Bank rate", "BOC",
                       "", "", "", "", "V122530", "Bank rate", "Bank rate", FILL_WHITE))

# boc_fx_usd: CAD/USD exchange rate (dims=0)
rows_to_append.append(("Bank of Canada", "monthly", "boc_fx_usd",
                       "CAD/USD exchange rate", "BOC",
                       "", "", "", "", "FXUSDCAD", "CAD/USD exchange rate", "CAD/USD", FILL_WHITE))

# boc_bond_yields: Government of Canada benchmark bond yields (dims=1)
for term, vec, short in [
    ("2-year",     "BD.CDN.2YR.DQ.YLD",   "2-yr yield"),
    ("5-year",     "BD.CDN.5YR.DQ.YLD",   "5-yr yield"),
    ("10-year",    "BD.CDN.10YR.DQ.YLD",  "10-yr yield"),
    ("Long-term",  "BD.CDN.LONG.DQ.YLD",  "Long-term yield"),
]:
    rows_to_append.append(("Bank of Canada", "monthly", "boc_bond_yields",
                           "GoC benchmark bond yields", "BOC",
                           "Term", term, "", "", vec, f"GoC bond yield \u2014 {term}", short, FILL_WHITE))

# boc_mortgage_new: New mortgage rates (dims=1)
for mtype, vec, short in [
    ("Total \u2014 insured",          "V122667775", "Total insured"),
    ("Variable rate \u2014 insured",  "V122667776", "Variable insured"),
    ("Fixed 5yr+ \u2014 insured",     "V122667780", "Fixed 5yr+ insured"),
    ("Total \u2014 uninsured",        "V122667781", "Total uninsured"),
    ("Variable rate \u2014 uninsured","V122667782", "Variable uninsured"),
    ("Fixed 5yr+ \u2014 uninsured",   "V122667786", "Fixed 5yr+ uninsured"),
]:
    rows_to_append.append(("Bank of Canada", "monthly", "boc_mortgage_new",
                           "New mortgage rates", "BOC",
                           "Type", mtype, "", "", vec, f"New mortgage \u2014 {mtype}", short, FILL_WHITE))

# boc_mortgage_existing: Existing mortgage rates (dims=1)
for mtype, vec, short in [
    ("Total \u2014 insured",          "V122667787", "Total insured"),
    ("Variable rate \u2014 insured",  "V122667788", "Variable insured"),
    ("Fixed 5yr+ \u2014 insured",     "V122667792", "Fixed 5yr+ insured"),
    ("Total \u2014 uninsured",        "V122667793", "Total uninsured"),
    ("Variable rate \u2014 uninsured","V122667794", "Variable uninsured"),
    ("Fixed 5yr+ \u2014 uninsured",   "V122667798", "Fixed 5yr+ uninsured"),
]:
    rows_to_append.append(("Bank of Canada", "monthly", "boc_mortgage_existing",
                           "Existing mortgage rates", "BOC",
                           "Type", mtype, "", "", vec, f"Existing mortgage \u2014 {mtype}", short, FILL_WHITE))

# bcpi_monthly: Bank of Canada commodity price index, monthly (dims=1)
for component, vec, short in [
    ("Total",            "M.BCPI", "Total"),
    ("Excluding energy", "M.BCNE", "Ex-energy"),
    ("Energy",           "M.ENER", "Energy"),
    ("Metals & minerals","M.MTLS", "Metals & min."),
    ("Agriculture",      "M.AGRI", "Agriculture"),
    ("Fish",             "M.FISH", "Fish"),
    ("Forestry",         "M.FOPR", "Forestry"),
]:
    rows_to_append.append(("Bank of Canada", "monthly", "bcpi_monthly",
                           "Commodity price index (monthly)", "BOC",
                           "Component", component, "", "", vec, f"BCPI (monthly) \u2014 {component}", short, FILL_WHITE))

# bcpi_weekly: Bank of Canada commodity price index, weekly (dims=1)
for component, vec, short in [
    ("Total",            "W.BCPI", "Total"),
    ("Excluding energy", "W.BCNE", "Ex-energy"),
    ("Energy",           "W.ENER", "Energy"),
    ("Metals & minerals","W.MTLS", "Metals & min."),
    ("Agriculture",      "W.AGRI", "Agriculture"),
    ("Fish",             "W.FISH", "Fish"),
    ("Forestry",         "W.FOPR", "Forestry"),
]:
    rows_to_append.append(("Bank of Canada", "monthly", "bcpi_weekly",
                           "Commodity price index (weekly)", "BOC",
                           "Component", component, "", "", vec, f"BCPI (weekly) \u2014 {component}", short, FILL_WHITE))

# ── CATEGORY: Labour ──────────────────────────────────────────────────────────

# lfs_ind: LFS Employment by industry (14-10-0355-02) — Canada, SA and NSA
LFS_IND_SERIES = [
    ("Total employed, all industries",                       "Total",           2057603, 2057812),
    ("Goods-producing sector",                               "Goods-producing", 2057604, 2057813),
    ("  Forestry, fishing, mining, quarrying, oil and gas",  "Forestry/mining", 2057606, 2057815),
    ("  Construction",                                       "Construction",    2057608, 2057817),
    ("  Manufacturing",                                      "Manufacturing",   2057609, 2057818),
    ("Services-producing sector",                            "Services",        2057610, 2057819),
    ("  Wholesale and retail trade",                         "Wholesale/retail",2057611, 2057820),
    ("  Transportation and warehousing",                     "Transportation",  2057612, 2057821),
    ("  Finance, insurance, real estate, rental and leasing","Finance/ins/RE",  2057613, 2057822),
    ("  Professional, scientific and technical services",    "Prof/sci/tech",   2057614, 2057823),
    ("  Educational services",                               "Education",       2057616, 2057825),
    ("  Health care and social assistance",                  "Health care",     2057617, 2057826),
    ("  Accommodation and food services",                    "Accomm/food",     2057619, 2057828),
    ("  Public administration",                              "Public admin",    2057621, 2057830),
]
for ind_name, short, sa_vec, nsa_vec in LFS_IND_SERIES:
    for adj, vec in [("Seasonally adjusted", sa_vec), ("Unadjusted", nsa_vec)]:
        fl = f"LFS — {ind_name.strip()} ({adj[:3]})"
        rows_to_append.append(("Labour", "monthly", "lfs_ind",
                               "LFS employment by industry", "14-10-0355-02",
                               "Seasonal adjustment", adj, "Industry", ind_name.strip(),
                               str(vec), fl, short, FILL_WHITE))

# lfs_class: LFS Employment by class of worker (14-10-0288-01) — Canada, SA and NSA
LFS_CLASS_SERIES = [
    ("Total employed, all classes of workers", "Total employed",   2066967, 2067132),
    ("Employees",                              "Employees",         2066968, 2067133),
    ("  Public sector employees",              "Public sector",     2066969, 2067134),
    ("  Private sector employees",             "Private sector",    2066970, 2067135),
    ("Self-employed",                          "Self-employed",     2066971, 2067136),
]
for cls_name, short, sa_vec, nsa_vec in LFS_CLASS_SERIES:
    for adj, vec in [("Seasonally adjusted", sa_vec), ("Unadjusted", nsa_vec)]:
        fl = f"LFS — {cls_name.strip()} ({adj[:3]})"
        rows_to_append.append(("Labour", "monthly", "lfs_class",
                               "LFS employment by class of worker", "14-10-0288-01",
                               "Seasonal adjustment", adj, "Class of worker", cls_name.strip(),
                               str(vec), fl, short, FILL_WHITE))

# lfs_wages: LFS average hourly wages by industry (14-10-0063-01) — Canada, NSA
LFS_WAGES_SERIES = [
    ("Total employees, all industries",                      "Total",           2132579),
    ("Goods-producing sector",                               "Goods-producing", 2132654),
    ("  Forestry, fishing, mining, quarrying, oil and gas",  "Forestry/mining", 2132584),
    ("  Construction",                                       "Construction",    2132599),
    ("  Manufacturing",                                      "Manufacturing",   2132604),
    ("Services-producing sector",                            "Services",        2132594),
    ("  Wholesale and retail trade",                         "Wholesale/retail",2132609),
    ("  Transportation and warehousing",                     "Transportation",  2132614),
    ("  Finance, insurance, real estate, rental and leasing","Finance/ins/RE",  2132619),
    ("  Professional, scientific and technical services",    "Prof/sci/tech",   2132664),
    ("  Educational services",                               "Education",       2132624),
    ("  Health care and social assistance",                  "Health care",     2132629),
    ("  Information, culture and recreation",                "Info/culture",    2132634),
    ("  Accommodation and food services",                    "Accomm/food",     2132644),
    ("  Public administration",                              "Public admin",    2132669),
]
for ind_name, short, vec in LFS_WAGES_SERIES:
    fl = f"LFS avg hourly wage — {ind_name.strip()}"
    rows_to_append.append(("Labour", "monthly", "lfs_wages",
                           "LFS avg hourly wage by industry (NSA)", "14-10-0063-01",
                           "Industry", ind_name.strip(), "", "",
                           str(vec), fl, short, FILL_WHITE))

# seph_earnings: SEPH average weekly earnings by industry (14-10-0223-01) — Canada, SA
SEPH_EARNINGS_SERIES = [
    ("Industrial aggregate (incl. unclassified)",            "IA incl. uncl.",  "v1544290282"),
    ("Industrial aggregate (excl. unclassified)",            "IA excl. uncl.",  "v79311153"),
    ("Goods producing industries",                           "Goods-producing", "v79311152"),
    ("  Construction",                                       "Construction",    "v79311156"),
    ("  Manufacturing",                                      "Manufacturing",   "v79311157"),
    ("Service producing industries",                         "Services",        "v79311162"),
    ("  Wholesale trade",                                    "Wholesale",       "v79311160"),
    ("  Retail trade",                                       "Retail",          "v79311163"),
    ("  Transportation and warehousing",                     "Transportation",  "v79311164"),
    ("  Finance and insurance",                              "Finance/ins.",    "v79311166"),
    ("  Real estate and rental and leasing",                 "Real estate",     "v79311167"),
    ("  Professional, scientific and technical services",    "Prof/sci/tech",   "v79311168"),
    ("  Educational services",                               "Education",       "v79311171"),
    ("  Health care and social assistance",                  "Health care",     "v79311172"),
    ("  Arts, entertainment and recreation",                 "Arts/entertain.", "v79311173"),
    ("  Accommodation and food services",                    "Accomm/food",     "v79311174"),
]
for ind_name, short, vec in SEPH_EARNINGS_SERIES:
    fl = f"SEPH avg weekly earnings — {ind_name.strip()}"
    rows_to_append.append(("Labour", "monthly", "seph_earnings",
                           "SEPH avg weekly earnings by industry (SA)", "14-10-0223-01",
                           "Industry", ind_name.strip(), "", "",
                           vec, fl, short, FILL_WHITE))

# ── CATEGORY: Housing additions ────────────────────────────────────────────────

# nhpi: New housing price index (18-10-0205-01) — Canada
for price_type, vec, short in [
    ("Total (house and land)", "v111955442", "Total"),
    ("House only",             "v111955443", "House only"),
    ("Land only",              "v111955444", "Land only"),
]:
    rows_to_append.append(("Housing and households", "monthly", "nhpi",
                           "New housing price index", "18-10-0205-01",
                           "Price type", price_type, "", "",
                           vec, f"NHPI — {price_type}", short, FILL_GREY))

# ── CATEGORY: Prices additions ─────────────────────────────────────────────────

# ippi: Industrial Product Price Index (18-10-0265-01)
IPPI_SERIES = [
    ("Total IPPI",                                                          "Total IPPI",       "v1230995983"),
    ("Total IPPI, excl. energy and petroleum products",                     "Excl. energy",     "v1230995984"),
    ("Meat, fish and dairy products [P11]",                                 "Meat/fish/dairy",  "v1230995985"),
    ("Fruit, vegetables, feed and other food products [P12]",               "Fruit/veg/feed",   "v1230995986"),
    ("Chemicals and chemical products [P31]",                               "Chemicals",        "v1230995992"),
    ("Lumber and other wood products [P41]",                                "Lumber/wood",      "v1230995994"),
    ("Pulp and paper products [P42]",                                       "Pulp & paper",     "v1230995995"),
    ("Energy and petroleum products [P51]",                                 "Energy/petroleum", "v1230995996"),
    ("Primary ferrous metal products [P61]",                                "Ferrous metals",   "v1230995997"),
    ("Primary non-ferrous metal products [P62]",                            "Non-ferrous metals","v1230995998"),
    ("Fabricated metal products and construction materials [P63]",          "Fab metals/constr","v1230995999"),
    ("Motorized and recreational vehicles [P71]",                           "Motor vehicles",   "v1230996000"),
    ("Machinery and equipment [P72]",                                       "Machinery/equip.", "v1230996001"),
    ("Electrical, electronic and telecom products [P73]",                   "Electrical/elec.", "v1230996002"),
    ("Cement, glass, and other non-metallic mineral products [P81]",        "Cement/glass",     "v1230996004"),
]
for prod_name, short, vec in IPPI_SERIES:
    rows_to_append.append(("Prices", "monthly", "ippi",
                           "IPPI (industrial product prices)", "18-10-0265-01",
                           "Product group", prod_name, "", "",
                           vec, f"IPPI — {prod_name}", short, FILL_GREY))

# rmpi: Raw Materials Price Index (18-10-0268-01)
RMPI_SERIES = [
    ("Total RMPI",                                  "Total RMPI",       "v1230998135"),
    ("Crude energy products [M51]",                 "Crude energy",     "v1230998136"),
    ("  Crude oil and bitumen",                     "Crude oil/bitumen","v1230998137"),
    ("  Natural gas",                               "Natural gas",      "v1230998141"),
    ("Total, excluding crude energy products",      "Excl. crude energy","v1230998148"),
    ("Crop products [M11]",                         "Crop products",    "v1230998149"),
    ("  Wheat",                                     "Wheat",            "v1230998150"),
    ("  Canola",                                    "Canola",           "v1230998151"),
    ("Animals and animal products [M21]",           "Animals/products", "v1230998165"),
    ("Non-metallic minerals [M31]",                 "Non-metallic min.","v1230998177"),
    ("Forestry products [M41]",                     "Forestry products","v1230998186"),
    ("Metal ores, concentrates and scrap [M61]",    "Metal ores/scrap", "v1230998193"),
]
for mat_name, short, vec in RMPI_SERIES:
    rows_to_append.append(("Prices", "monthly", "rmpi",
                           "RMPI (raw materials prices)", "18-10-0268-01",
                           "Material group", mat_name.strip(), "", "",
                           vec, f"RMPI — {mat_name.strip()}", short, FILL_GREY))

# ── CATEGORY: Government finance ───────────────────────────────────────────────

GOVT_FIN_SERIES = [
    ("Revenue",                      "Revenue",         "v52531053"),
    ("Expense",                      "Expense",         "v52531064"),
    ("Gross operating balance",      "Gross op. bal.",  "v52531073"),
    ("Net operating balance",        "Net op. bal.",    "v52531074"),
    ("Net lending or borrowing",     "Net lend/borrow", "v52531076"),
    ("Total expenditure",            "Total expend.",   "v52531092"),
]
for item_name, short, vec in GOVT_FIN_SERIES:
    rows_to_append.append(("Government finance", "quarterly", "fed_govt_ops",
                           "Federal government operations", "10-10-0015-01",
                           "Item", item_name, "", "",
                           vec, f"Federal govt — {item_name}", short, FILL_WHITE))

# ── CATEGORY: Energy ───────────────────────────────────────────────────────────

# crude_oil: Supply and disposition of crude oil (25-10-0063-01)
for measure, vec, short in [
    ("Crude oil production (cubic metres)", "v107757044", "Production (m³)"),
    ("Crude oil production (barrels)",      "v107757045", "Production (bbl)"),
    ("Synthetic crude oil production (barrels)", "v107757061", "Synth. crude (bbl)"),
]:
    rows_to_append.append(("Energy", "monthly", "crude_oil",
                           "Crude oil supply & disposition", "25-10-0063-01",
                           "Measure", measure, "", "",
                           vec, f"Crude oil — {measure}", short, FILL_GREY))

# nat_gas: Supply and disposition of natural gas (25-10-0055-01)
for measure, vec, short in [
    ("Gross withdrawals (GJ)",       "v107638455", "Gross withdrawals"),
    ("Marketable production (GJ)",   "v107638457", "Mktable production"),
    ("Exports (GJ)",                 "v107638467", "Exports"),
]:
    rows_to_append.append(("Energy", "monthly", "nat_gas",
                           "Natural gas supply & disposition", "25-10-0055-01",
                           "Measure", measure, "", "",
                           vec, f"Natural gas — {measure}", short, FILL_GREY))

# ── CATEGORY: Agriculture ──────────────────────────────────────────────────────

# farm_receipts: Farm cash receipts (32-10-0046-01) — Canada, quarterly
FARM_RECEIPTS_SERIES = [
    ("Total farm cash receipts",               "Total",              "v170328"),
    ("Total crop receipts",                    "Total crops",        "v170329"),
    ("  Wheat (except durum)",                 "Wheat",              "v170330"),
    ("  Durum wheat",                          "Durum wheat",        "v170352"),
    ("  Canola (including rapeseed)",           "Canola",             "v170334"),
    ("  Soybeans",                             "Soybeans",           "v170335"),
    ("  Corn for grain",                       "Corn",               "v170336"),
    ("  Oats",                                 "Oats",               "v170363"),
    ("  Barley",                               "Barley",             "v170365"),
    ("  Lentils",                              "Lentils",            "v170349"),
    ("  Dry peas",                             "Dry peas",           "v170353"),
    ("  Fresh potatoes",                       "Potatoes",           "v170338"),
    ("Total livestock and livestock product receipts", "Total livestock", "v170368"),
    ("  Cattle",                               "Cattle",             "v170369"),
    ("  Hogs",                                 "Hogs",               "v170380"),
    ("  Unprocessed milk from bovine",         "Dairy milk",         "v170383"),
    ("  Chickens for meat",                    "Chickens",           "v170384"),
    ("  Eggs in shell",                        "Eggs",               "v170386"),
    ("Total receipts from direct payments",    "Direct payments",    "v170387"),
]
for item_name, short, vec in FARM_RECEIPTS_SERIES:
    rows_to_append.append(("Agriculture", "quarterly", "farm_receipts",
                           "Farm cash receipts", "32-10-0046-01",
                           "Type", item_name.strip(), "", "",
                           vec, f"Farm cash receipts — {item_name.strip()}", short, FILL_WHITE))

# crop_prod: Crop production, metric tonnes (32-10-0359-01) — Canada, annual
CROP_PROD_SERIES = [
    ("Wheat, all",      "Wheat (all)",   "v114995749"),
    ("Wheat, spring",   "Wheat (spring)","v114995750"),
    ("Wheat, durum",    "Wheat (durum)", "v115115187"),
    ("Barley",          "Barley",        "v114995720"),
    ("Oats",            "Oats",          "v114995738"),
    ("Canola (rapeseed)","Canola",       "v114995727"),
    ("Corn for grain",  "Corn",          "v114995731"),
    ("Soybeans",        "Soybeans",      "v114995744"),
    ("Lentils",         "Lentils",       "v114995735"),
    ("Peas, dry",       "Dry peas",      "v114995739"),
    ("Flaxseed",        "Flaxseed",      "v114995734"),
    ("Mustard seed",    "Mustard seed",  "v114995737"),
    ("Chick peas",      "Chick peas",    "v114995729"),
]
for crop_name, short, vec in CROP_PROD_SERIES:
    rows_to_append.append(("Agriculture", "annual", "crop_prod",
                           "Crop production (metric tonnes)", "32-10-0359-01",
                           "Crop", crop_name, "", "",
                           vec, f"Crop production — {crop_name}", short, FILL_WHITE))

# ── CATEGORY: International additions ─────────────────────────────────────────

# curr_acct: Current account balance, SA (36-10-0018-01) — quarterly
CURR_ACCT_SERIES = [
    ("Total current account",  "Balance", "Total CA",       "v61915304"),
    ("Goods and services",     "Balance", "Goods & svcs",   "v61915305"),
    ("  Goods",                "Balance", "Goods",          "v61915306"),
    ("  Services",             "Balance", "Services",       "v61915308"),
    ("Primary income",         "Balance", "Primary income", "v61915313"),
    ("Secondary income",       "Balance", "Secondary inc.", "v61915327"),
    ("Total current account",  "Receipts","Receipts",       "v61915244"),
    ("Total current account",  "Payments","Payments",       "v61915273"),
]
for comp, flow, short, vec in CURR_ACCT_SERIES:
    fl = f"Current account (SA) — {comp.strip()}, {flow}"
    rows_to_append.append(("International securities", "quarterly", "curr_acct",
                           "Current account balance (SA)", "36-10-0018-01",
                           "Component", comp.strip(), "Flow", flow,
                           vec, fl, short, FILL_GREY))

# fdi: Foreign direct investment flows (36-10-0025-01) — quarterly
FDI_SERIES = [
    ("Canadian direct investment abroad", "All countries",    "CDI abroad (all)",  "v61913911"),
    ("Canadian direct investment abroad", "United States",    "CDI abroad (US)",   "v61913915"),
    ("Foreign direct investment in Canada","All countries",   "FDI in Canada (all)","v61913923"),
    ("Foreign direct investment in Canada","United States",   "FDI in Canada (US)","v61913927"),
]
for direction, country, short, vec in FDI_SERIES:
    fl = f"FDI — {direction}, {country}"
    rows_to_append.append(("International securities", "quarterly", "fdi_flows",
                           "Foreign direct investment flows", "36-10-0025-01",
                           "Direction", direction, "Countries", country,
                           vec, fl, short, FILL_GREY))

# ── Write all rows to the worksheet ──────────────────────────────────────────
current_row = start_row
for row_data in rows_to_append:
    if len(row_data) == 19:
        # 5-dim row: ..., d3n, d3v, d4n, d4v, d5n, d5v, vec, fl, sl, fill
        category, freq, series_id, series_name, table_id, d1n, d1v, d2n, d2v, d3n, d3v, d4n, d4v, d5n, d5v, vector, full_label, short_label, fill = row_data
    elif len(row_data) == 17:
        # 4-dim row
        category, freq, series_id, series_name, table_id, d1n, d1v, d2n, d2v, d3n, d3v, d4n, d4v, vector, full_label, short_label, fill = row_data
        d5n, d5v = "", ""
    elif len(row_data) == 15:
        # 3-dim row
        category, freq, series_id, series_name, table_id, d1n, d1v, d2n, d2v, d3n, d3v, vector, full_label, short_label, fill = row_data
        d4n, d4v, d5n, d5v = "", "", "", ""
    else:
        # Standard ≤2-dim row (13 elements)
        category, freq, series_id, series_name, table_id, d1n, d1v, d2n, d2v, vector, full_label, short_label, fill = row_data
        d3n, d3v, d4n, d4v, d5n, d5v = "", "", "", "", "", ""
    values = [category, freq, series_id, series_name, table_id, d1n, d1v, d2n, d2v, d3n, d3v, d4n, d4v, d5n, d5v, vector, full_label, short_label]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = val
        cell.fill = fill
        cell.font = FONT
        # Vector column (col 16) must be text format
        if col_idx == 16:
            cell.number_format = "@"
            cell.data_type = "s"
    current_row += 1

wb.save("/Users/jasonkirby/Desktop/StatCanApp/vectors-template.xlsx")

# ── Summary ───────────────────────────────────────────────────────────────────
total_rows = ws.max_row
print(f"File saved successfully.")
print(f"Total rows (header + data): {total_rows}")
print(f"New rows appended: {len(rows_to_append)}")
print()

# Count per category
cat_counts = {}
for row_data in rows_to_append:
    cat = row_data[0]
    cat_counts[cat] = cat_counts.get(cat, 0) + 1

print("Row count per NEW category:")
for cat, count in cat_counts.items():
    print(f"  {cat}: {count}")

# Also count existing
existing_cats = {}
for r in range(2, start_row):
    cat = ws.cell(r, 1).value
    existing_cats[cat] = existing_cats.get(cat, 0) + 1
print()
print("Existing categories (unchanged):")
for cat, count in existing_cats.items():
    print(f"  {cat}: {count}")
