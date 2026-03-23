#!/usr/bin/env python3
"""
Build vectors-template.xlsx from scratch using openpyxl.
All rows exactly as specified.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/Users/jasonkirby/Desktop/StatCanApp/vectors-template.xlsx"

COLUMNS = [
    "category", "freq", "series_id", "series_name", "table_id",
    "dim1_name", "dim1_value", "dim2_name", "dim2_value",
    "dim3_name", "dim3_value", "dim4_name", "dim4_value",
    "dim5_name", "dim5_value",
    "vector", "full_label", "short_label"
]

def make_row(category, freq, series_id, series_name, table_id,
             dim1_name, dim1_value, dim2_name, dim2_value,
             vector, full_label, short_label):
    return [category, freq, series_id, series_name, table_id,
            dim1_name, dim1_value, dim2_name, dim2_value,
            "", "", "", "",  # dim3/dim4 name+value (empty for ≤2-dim series)
            "", "",          # dim5 name+value (empty for ≤2-dim series)
            str(vector), full_label, short_label]

rows = []

# ══════════════════════════════════════════════════════════════════
# CATEGORY: GDP (expenditure based) | quarterly | fill: WHITE
# ══════════════════════════════════════════════════════════════════
CAT_GDP_EXP = "GDP (expenditure based)"
FREQ_Q      = "quarterly"
TABLE_EXP   = "36-10-0104-01"

# gdp_real_exp — GDP at market prices FIRST
gdp_real = [
    (62305752, "Gross domestic product at market prices",                             "GDP"),
    (62305724, "Household final consumption expenditure",                             "Household consumption"),
    (62305731, "General governments final consumption expenditure",                   "Gov't consumption"),
    (62305733, "Business gross fixed capital formation",                              "Business investment"),
    (62305734, "Residential structures",                                              "Residential"),
    (62305735, "Non-residential structures, machinery and equipment",                 "Non-res. structures & M&E"),
    (62305736, "Non-residential structures",                                          "Non-res. structures"),
    (62305737, "Machinery and equipment",                                             "Machinery & equipment"),
    (62305738, "Intellectual property products",                                      "Intellectual property"),
    (62305740, "General governments gross fixed capital formation",                   "Gov't investment"),
    (62305741, "Investment in inventories",                                           "Inventory investment"),
    (62305742, "Of which: business investment in inventories",                        "Business inventories"),
    (62305745, "Exports of goods and services",                                       "Exports"),
    (62305748, "Less: imports of goods and services",                                 "Imports"),
]
for vec, dim1, short in gdp_real:
    rows.append(make_row(CAT_GDP_EXP, FREQ_Q, "gdp_real_exp", "Real GDP (2017$, SAAR)",
                         TABLE_EXP, "Component", dim1, "", "", vec, dim1, short))

# gdp_nom_exp — GDP at market prices FIRST
gdp_nom = [
    (62305783, "Gross domestic product at market prices",                             "GDP"),
    (62305755, "Household final consumption expenditure",                             "Household consumption"),
    (62305762, "General governments final consumption expenditure",                   "Gov't consumption"),
    (62305764, "Business gross fixed capital formation",                              "Business investment"),
    (62305765, "Residential structures",                                              "Residential"),
    (62305766, "Non-residential structures, machinery and equipment",                 "Non-res. structures & M&E"),
    (62305767, "Non-residential structures",                                          "Non-res. structures"),
    (62305768, "Machinery and equipment",                                             "Machinery & equipment"),
    (62305769, "Intellectual property products",                                      "Intellectual property"),
    (62305771, "General governments gross fixed capital formation",                   "Gov't investment"),
    (62305772, "Investment in inventories",                                           "Inventory investment"),
    (62305773, "Of which: business investment in inventories",                        "Business inventories"),
    (62305776, "Exports of goods and services",                                       "Exports"),
    (62305779, "Less: imports of goods and services",                                 "Imports"),
]
for vec, dim1, short in gdp_nom:
    rows.append(make_row(CAT_GDP_EXP, FREQ_Q, "gdp_nom_exp", "Nominal GDP (SAAR)",
                         TABLE_EXP, "Component", dim1, "", "", vec, dim1, short))

# ══════════════════════════════════════════════════════════════════
# CATEGORY: GDP (by industry) | monthly | fill: EEEEEE
# ══════════════════════════════════════════════════════════════════
CAT_GDP_IND = "GDP (by industry)"
FREQ_M      = "monthly"
TABLE_IND   = "36-10-0434-01"

gdp_ind = [
    (65201210, "All industries [T001]",                                                       "All industries"),
    (65201211, "Goods-producing industries [T002]",                                           "Goods-producing"),
    (65201212, "Services-producing industries [T003]",                                        "Services-producing"),
    (65201213, "Business sector industries [T004]",                                           "Business sector"),
    (65201216, "Non-business sector industries [T007]",                                       "Non-business sector"),
    (65201222, "Information and communication technology sector [T013]",                      "ICT sector"),
    (65201225, "Energy sector [T016]",                                                        "Energy sector"),
    (65201227, "Public sector [T018]",                                                        "Public sector"),
    (65201229, "Agriculture, forestry, fishing and hunting [11]",                             "Agriculture"),
    (65201236, "Mining, quarrying, and oil and gas extraction [21]",                          "Mining & oil/gas"),
    (65201254, "Utilities [22]",                                                              "Utilities"),
    (65201258, "Construction [23]",                                                           "Construction"),
    (65201263, "Manufacturing [31-33]",                                                       "Manufacturing"),
    (65201358, "Wholesale trade [41]",                                                        "Wholesale"),
    (65201368, "Retail trade [44-45]",                                                        "Retail"),
    (65201381, "Transportation and warehousing [48-49]",                                      "Transportation"),
    (65201398, "Information and cultural industries [51]",                                    "Information"),
    (65201407, "Finance and insurance [52]",                                                  "Finance & insurance"),
    (65201419, "Real estate and rental and leasing [53]",                                     "Real estate"),
    (65201429, "Professional, scientific and technical services [54]",                        "Professional services"),
    (65201441, "Management of companies and enterprises [55]",                                "Management"),
    (65201442, "Administrative and support, waste management and remediation services [56]",  "Admin. services"),
    (65201452, "Educational services [61]",                                                   "Education"),
    (65201457, "Health care and social assistance [62]",                                      "Health care"),
    (65201463, "Arts, entertainment and recreation [71]",                                     "Arts & recreation"),
    (65201468, "Accommodation and food services [72]",                                        "Accommodation & food"),
    (65201471, "Other services (except public administration) [81]",                          "Other services"),
    (65201476, "Public administration [91]",                                                  "Public administration"),
]
for vec, dim1, short in gdp_ind:
    rows.append(make_row(CAT_GDP_IND, FREQ_M, "gdp_industry", "GDP by industry (2017$)",
                         TABLE_IND, "Industry", dim1, "", "", vec, dim1, short))

# ══════════════════════════════════════════════════════════════════
# CATEGORY: Labour | monthly | fill: WHITE
# ══════════════════════════════════════════════════════════════════
CAT_LAB   = "Labour"
TABLE_LFS = "14-10-0287-01"

GEOS = [
    "Canada",
    "Newfoundland and Labrador",
    "Prince Edward Island",
    "Nova Scotia",
    "New Brunswick",
    "Quebec",
    "Ontario",
    "Manitoba",
    "Saskatchewan",
    "Alberta",
    "British Columbia",
]
CHARS = [
    "Population",
    "Labour force",
    "Employment",
    "Full-time employment",
    "Part-time employment",
    "Unemployment",
    "Unemployment rate",
    "Participation rate",
    "Employment rate",
]
SHORT_CHAR = {
    "Population":           "Population",
    "Labour force":         "Labour force",
    "Employment":           "Employment",
    "Full-time employment": "Full-time",
    "Part-time employment": "Part-time",
    "Unemployment":         "Unemployment",
    "Unemployment rate":    "Unemp. rate",
    "Participation rate":   "Participation rate",
    "Employment rate":      "Employment rate",
}

# ── LFS SA ─────────────────────────────────────────────────────────
LFS_SA_VEC = {
    "Canada":                    [2062809, 2062810, 2062811, 2062812, 2062813, 2062814, 2062815, 2062816, 2062817],
    "Newfoundland and Labrador": [2062998, 2062999, 2063000, 2063001, 2063002, 2063003, 2063004, 2063005, 2063006],
    "Prince Edward Island":      [2063187, 2063188, 2063189, 2063190, 2063191, 2063192, 2063193, 2063194, 2063195],
    "Nova Scotia":               [2063376, 2063377, 2063378, 2063379, 2063380, 2063381, 2063382, 2063383, 2063384],
    "New Brunswick":             [2063565, 2063566, 2063567, 2063568, 2063569, 2063570, 2063571, 2063572, 2063573],
    "Quebec":                    [2063754, 2063755, 2063756, 2063757, 2063758, 2063759, 2063760, 2063761, 2063762],
    "Ontario":                   [2063943, 2063944, 2063945, 2063946, 2063947, 2063948, 2063949, 2063950, 2063951],
    "Manitoba":                  [2064132, 2064133, 2064134, 2064135, 2064136, 2064137, 2064138, 2064139, 2064140],
    "Saskatchewan":              [2064321, 2064322, 2064323, 2064324, 2064325, 2064326, 2064327, 2064328, 2064329],
    "Alberta":                   [2064510, 2064511, 2064512, 2064513, 2064514, 2064515, 2064516, 2064517, 2064518],
    "British Columbia":          [2064699, 2064700, 2064701, 2064702, 2064703, 2064704, 2064705, 2064706, 2064707],
}
for geo in GEOS:
    for i, char in enumerate(CHARS):
        rows.append(make_row(CAT_LAB, FREQ_M, "lfs_sa", "LFS (SA)", TABLE_LFS,
                             "Geography", geo, "Characteristic", char,
                             LFS_SA_VEC[geo][i],
                             f"{geo} - {char}",
                             SHORT_CHAR[char]))

# ── LFS NSA ────────────────────────────────────────────────────────
LFS_NSA_VEC = {
    "Canada":                    [2064888, 2064889, 2064890, 2064891, 2064892, 2064893, 2064894, 2064895, 2064896],
    "Newfoundland and Labrador": [2065077, 2065078, 2065079, 2065080, 2065081, 2065082, 2065083, 2065084, 2065085],
    "Prince Edward Island":      [2065266, 2065267, 2065268, 2065269, 2065270, 2065271, 2065272, 2065273, 2065274],
    "Nova Scotia":               [2065455, 2065456, 2065457, 2065458, 2065459, 2065460, 2065461, 2065462, 2065463],
    "New Brunswick":             [2065644, 2065645, 2065646, 2065647, 2065648, 2065649, 2065650, 2065651, 2065652],
    "Quebec":                    [2065833, 2065834, 2065835, 2065836, 2065837, 2065838, 2065839, 2065840, 2065841],
    "Ontario":                   [2066022, 2066023, 2066024, 2066025, 2066026, 2066027, 2066028, 2066029, 2066030],
    "Manitoba":                  [2066211, 2066212, 2066213, 2066214, 2066215, 2066216, 2066217, 2066218, 2066219],
    "Saskatchewan":              [2066400, 2066401, 2066402, 2066403, 2066404, 2066405, 2066406, 2066407, 2066408],
    "Alberta":                   [2066589, 2066590, 2066591, 2066592, 2066593, 2066594, 2066595, 2066596, 2066597],
    "British Columbia":          [2066778, 2066779, 2066780, 2066781, 2066782, 2066783, 2066784, 2066785, 2066786],
}
for geo in GEOS:
    for i, char in enumerate(CHARS):
        rows.append(make_row(CAT_LAB, FREQ_M, "lfs_nsa", "LFS (NSA)", TABLE_LFS,
                             "Geography", geo, "Characteristic", char,
                             LFS_NSA_VEC[geo][i],
                             f"{geo} - {char}",
                             SHORT_CHAR[char]))

# ── SEPH SA ────────────────────────────────────────────────────────
TABLE_SEPH_SA = "14-10-0220-01"
SEPH_SA_ROWS = [
    (54026324, "Industrial aggregate including unclassified businesses [00-91N]",            "Industrial aggregate"),
    (54026340, "Forestry, logging and support [11N]",                                        "Forestry"),
    (54026342, "Mining, quarrying, and oil and gas extraction [21]",                         "Mining & oil/gas"),
    (54026344, "Oil and gas extraction [211,2111]",                                          "Oil & gas"),
    (54026346, "Mining and quarrying (except oil and gas) [212]",                            "Mining (ex. oil/gas)"),
    (54026354, "Support activities for mining, and oil and gas extraction [213,2131]",       "Mining support"),
    (54026356, "Utilities [22,221]",                                                         "Utilities"),
    (54026364, "Construction [23]",                                                          "Construction"),
    (54026392, "Manufacturing [31-33]",                                                      "Manufacturing"),
    (54026610, "Trade [41-45N]",                                                             "Trade"),
    (54026752, "Transportation and warehousing [48-49]",                                     "Transportation"),
    (54026820, "Information and cultural industries [51]",                                   "Information"),
    (54026854, "Finance and insurance [52]",                                                 "Finance & insurance"),
    (54026878, "Real estate and rental and leasing [53]",                                    "Real estate"),
    (54026918, "Professional, scientific and technical services [54,541]",                   "Professional services"),
    (54026920, "Management of companies and enterprises [55,551,5511]",                      "Management"),
    (54026922, "Administrative and support, waste management and remediation services [56]", "Admin. services"),
    (54026948, "Educational services [61,611]",                                              "Education"),
    (54026966, "Health care and social assistance [62]",                                     "Health care"),
    (54027012, "Arts, entertainment and recreation [71]",                                    "Arts & recreation"),
    (54027036, "Accommodation and food services [72]",                                       "Accommodation & food"),
    (54027056, "Other services (except public administration) [81]",                         "Other services"),
    (54027088, "Public administration [91]",                                                 "Public administration"),
]
for vec, dim1, short in SEPH_SA_ROWS:
    rows.append(make_row(CAT_LAB, FREQ_M, "seph_sa", "SEPH (SA)",
                         TABLE_SEPH_SA, "Industry", dim1, "", "", vec, dim1, short))

# ── SEPH NSA ───────────────────────────────────────────────────────
TABLE_SEPH_NSA = "14-10-0201-01"

# 20 industries for SEPH NSA (full names as dim2_value)
SEPH_NSA_IND = [
    "Industrial aggregate including unclassified businesses",
    "Forestry, logging and support",
    "Mining, quarrying, and oil and gas extraction",
    "Utilities",
    "Construction",
    "Manufacturing",
    "Trade",
    "Transportation and warehousing",
    "Information and cultural industries",
    "Finance and insurance",
    "Real estate and rental and leasing",
    "Professional, scientific and technical services",
    "Management of companies and enterprises",
    "Administrative and support, waste management and remediation services",
    "Educational services",
    "Health care and social assistance",
    "Arts, entertainment and recreation",
    "Accommodation and food services",
    "Other services (except public administration)",
    "Public administration",
]

SEPH_NSA_SHORT = {
    "Industrial aggregate including unclassified businesses":                    "Industrial aggregate",
    "Forestry, logging and support":                                             "Forestry",
    "Mining, quarrying, and oil and gas extraction":                             "Mining & oil/gas",
    "Utilities":                                                                 "Utilities",
    "Construction":                                                              "Construction",
    "Manufacturing":                                                             "Manufacturing",
    "Trade":                                                                     "Trade",
    "Transportation and warehousing":                                            "Transportation",
    "Information and cultural industries":                                       "Information",
    "Finance and insurance":                                                     "Finance & insurance",
    "Real estate and rental and leasing":                                        "Real estate",
    "Professional, scientific and technical services":                           "Professional services",
    "Management of companies and enterprises":                                   "Management",
    "Administrative and support, waste management and remediation services":     "Admin. services",
    "Educational services":                                                      "Education",
    "Health care and social assistance":                                         "Health care",
    "Arts, entertainment and recreation":                                        "Arts & recreation",
    "Accommodation and food services":                                           "Accommodation & food",
    "Other services (except public administration)":                             "Other services",
    "Public administration":                                                     "Public administration",
}

# 11 geographies × 20 industries = 220 rows
SEPH_NSA_GEO_VEC = {
    "Canada":                    [1556538,  1556541,    1556544,    1556551,    1556555,    1556564,    1556634,    1556740,    1556764,    1556777,    1556786,    1556795,    1556801,    1556802,    1556811,    1556690,    1556707,    1556716,    1556722,    1556736],
    "Newfoundland and Labrador": [1556817,  1544329221, 1544329225, 1544329231, 1556821,    1556827,    1556836,    1556890,    1556903,    13920919,   1544329393, 13920933,   44308727,   13920965,   1556918,    1556857,    44308734,   1556872,    1556878,    1556887],
    "Prince Edward Island":      [1556922,  1544329975, 1544329981, 1544329988, 1556925,    1556929,    1556934,    1556976,    1544330159, 13920920,   13920929,   13920934,   44308763,   44308764,   1556991,    1556947,    44308770,   1556960,    1556965,    1556973],
    "Nova Scotia":               [1556994,  1544330839, 1544330844, 1544330849, 1557000,    1557008,    1557023,    1557084,    13920909,   1557093,    1544330972, 1557103,    44308819,   13920966,   1557111,    1557051,    13920839,   1557065,    1557071,    1557081],
    "New Brunswick":             [1557115,  13920509,   13920514,   13920519,   1557121,    1557129,    1557148,    1557219,    13920910,   1557230,    1544331606, 1557244,    44308887,   1557248,    1557255,    1557182,    44308893,   1557201,    1557207,    1557216],
    "Quebec":                    [1557259,  1557262,    13920515,   1557268,    1557271,    1557280,    1557350,    1557454,    1557475,    1557488,    1557496,    1557505,    1557509,    1557510,    1557520,    1557403,    1557419,    1557428,    1557436,    1557450],
    "Ontario":                   [1557526,  13920510,   13920516,   1557536,    1557538,    1557546,    1557614,    1557718,    1557740,    1557752,    1557762,    1557771,    1557775,    1557776,    1557784,    1557668,    1557685,    1557693,    1557699,    1557714],
    "Manitoba":                  [1557790,  1544332566, 1557794,    1557796,    1557798,    1557807,    1557837,    1557912,    1557926,    1557933,    1557940,    1557948,    1557953,    1557954,    1557961,    1557875,    1557888,    1557891,    1557897,    1557909],
    "Saskatchewan":              [1557965,  13920511,   1557970,    1557975,    1557976,    1557982,    1557999,    1558068,    1558078,    1558083,    1558089,    1558096,    1558099,    1558100,    1558104,    1558034,    1558045,    1558050,    1558056,    1558065],
    "Alberta":                   [1558107,  13920512,   1558113,    1558116,    1558119,    1558127,    1558162,    1558246,    1558259,    1558268,    1558277,    1558286,    1558289,    1558290,    1558297,    1558203,    1558217,    1558225,    1558231,    1558242],
    "British Columbia":          [1558300,  1558303,    1558308,    1558313,    1558315,    1558325,    1558374,    1558483,    1558507,    1558519,    1558528,    1558537,    1558543,    1558544,    1558554,    1558433,    1558449,    1558459,    1558465,    1558479],
}

SEPH_NSA_GEOS = [
    "Canada", "Newfoundland and Labrador", "Prince Edward Island",
    "Nova Scotia", "New Brunswick", "Quebec", "Ontario",
    "Manitoba", "Saskatchewan", "Alberta", "British Columbia",
]

for geo in SEPH_NSA_GEOS:
    vecs = SEPH_NSA_GEO_VEC[geo]
    for i, ind in enumerate(SEPH_NSA_IND):
        rows.append(make_row(CAT_LAB, FREQ_M, "seph_nsa", "SEPH (NSA)",
                             TABLE_SEPH_NSA,
                             "Geography", geo,
                             "Industry", ind,
                             vecs[i],
                             f"{geo} - {ind}",
                             SEPH_NSA_SHORT[ind]))

# ── Job Vacancies SA ───────────────────────────────────────────────
TABLE_JOB = "14-10-0406-01"

JOB_IND = [
    ("Total all industries",                                       1446283287, 1446283289),
    ("Agriculture, forestry, fishing and hunting",                 1446283290, 1446283292),
    ("Mining, quarrying, and oil and gas extraction",              1446283293, 1446283295),
    ("Utilities",                                                  1446283296, 1446283298),
    ("Construction",                                               1446283299, 1446283301),
    ("Manufacturing",                                              1446283302, 1446283304),
    ("Wholesale trade",                                            1446283305, 1446283307),
    ("Retail trade",                                               1446283308, 1446283310),
    ("Transportation and warehousing",                             1446283311, 1446283313),
    ("Information and cultural industries",                        1446283314, 1446283316),
    ("Finance and insurance",                                      1446283317, 1446283319),
    ("Real estate and rental and leasing",                         1446283320, 1446283322),
    ("Professional, scientific and technical services",            1446283323, 1446283325),
    ("Management of companies and enterprises",                    1446283326, 1446283328),
    ("Administrative and support services",                        1446283329, 1446283331),
    ("Educational services",                                       1446283332, 1446283334),
    ("Health care and social assistance",                          1446283335, 1446283337),
    ("Arts, entertainment and recreation",                         1446283338, 1446283340),
    ("Accommodation and food services",                            1446283341, 1446283343),
    ("Other services (except public administration)",              1446283344, 1446283346),
    ("Public administration",                                      1446283347, 1446283349),
]
for ind, vac_vec, rate_vec in JOB_IND:
    rows.append(make_row(CAT_LAB, FREQ_M, "job_vac_sa", "Job vacancies (SA)",
                         TABLE_JOB,
                         "Measure", "Job vacancies",
                         "Industry", ind,
                         vac_vec,
                         f"Job vacancies, {ind}",
                         "Vacancies"))
    rows.append(make_row(CAT_LAB, FREQ_M, "job_vac_sa", "Job vacancies (SA)",
                         TABLE_JOB,
                         "Measure", "Job vacancy rate",
                         "Industry", ind,
                         rate_vec,
                         f"Job vacancy rate, {ind}",
                         "Vacancy rate"))

# ══════════════════════════════════════════════════════════════════
# Build workbook
# ══════════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "series"

FONT_HDR  = Font(name="Arial", size=10, bold=True)
FONT_DATA = Font(name="Arial", size=10)
FILL_HDR  = PatternFill("solid", fgColor="DDEEFF")
FILL_WHT  = PatternFill("solid", fgColor="FFFFFF")
FILL_GRY  = PatternFill("solid", fgColor="EEEEEE")
ALIGN_L   = Alignment(horizontal="left", vertical="center")

CAT_FILL = {
    CAT_GDP_EXP: FILL_WHT,
    CAT_GDP_IND: FILL_GRY,
    CAT_LAB:     FILL_WHT,
}

# Header
for ci, col in enumerate(COLUMNS, 1):
    c = ws.cell(row=1, column=ci, value=col)
    c.font = FONT_HDR
    c.fill = FILL_HDR
    c.alignment = ALIGN_L

# Data
VEC_COL = COLUMNS.index("vector") + 1  # 1-based (col 10)
for ri, data in enumerate(rows, 2):
    fill = CAT_FILL.get(data[0], FILL_WHT)
    for ci, val in enumerate(data, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = FONT_DATA
        c.fill = fill
        c.alignment = ALIGN_L
        if ci == VEC_COL:
            c.number_format = "@"

# Freeze pane
ws.freeze_panes = "A2"

# Auto-size columns (min 10, max 60)
for ci, col in enumerate(COLUMNS, 1):
    cl = get_column_letter(ci)
    max_len = len(col)
    for ri in range(2, len(rows) + 2):
        v = ws.cell(row=ri, column=ci).value
        if v is not None:
            max_len = max(max_len, len(str(v)))
    ws.column_dimensions[cl].width = min(max(max_len + 2, 10), 60)

wb.save(OUTPUT_PATH)
print(f"Saved: {OUTPUT_PATH}")

# ── Verify ────────────────────────────────────────────────────────
import os
assert os.path.exists(OUTPUT_PATH), "File not found!"
wb2   = openpyxl.load_workbook(OUTPUT_PATH)
ws2   = wb2.active
total = ws2.max_row
data  = total - 1

# Component counts
c_gdp_real   = len(gdp_real)
c_gdp_nom    = len(gdp_nom)
c_gdp_ind    = len(gdp_ind)
c_lfs_sa     = len(GEOS) * len(CHARS)
c_lfs_nsa    = len(GEOS) * len(CHARS)
c_seph_sa    = len(SEPH_SA_ROWS)
c_seph_nsa   = len(SEPH_NSA_GEOS) * len(SEPH_NSA_IND)
c_job_vac    = len(JOB_IND) * 2

print(f"\nRow breakdown:")
print(f"  gdp_real_exp   : {c_gdp_real:4d}")
print(f"  gdp_nom_exp    : {c_gdp_nom:4d}")
print(f"  gdp_industry   : {c_gdp_ind:4d}")
print(f"  lfs_sa         : {c_lfs_sa:4d}  ({len(GEOS)} geos × {len(CHARS)} chars)")
print(f"  lfs_nsa        : {c_lfs_nsa:4d}  ({len(GEOS)} geos × {len(CHARS)} chars)")
print(f"  seph_sa        : {c_seph_sa:4d}")
print(f"  seph_nsa       : {c_seph_nsa:4d}  ({len(SEPH_NSA_GEOS)} geos × {len(SEPH_NSA_IND)} industries)")
print(f"  job_vac_sa     : {c_job_vac:4d}  ({len(JOB_IND)} industries × 2 measures)")
total_expected = c_gdp_real + c_gdp_nom + c_gdp_ind + c_lfs_sa + c_lfs_nsa + c_seph_sa + c_seph_nsa + c_job_vac
print(f"  TOTAL expected : {total_expected}")
print(f"\nFile total rows (incl. header): {total}")
print(f"File data rows               : {data}")

assert total_expected == 539, f"Expected 539 data rows in logic, got {total_expected}"
assert data == 539,           f"Expected 539 data rows in file, got {data}"
print("\nAll checks passed. 539 data rows + 1 header = 540 total.")
