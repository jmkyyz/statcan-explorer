#!/usr/bin/env python3
"""
update_vectors.py — Updates Vectors.xlsx with:
  1. Trade series: replaces only the old 3-dim goods_prov_comm_partner with
     a new 4-dim (Province/Flow/Commodity/Partner) version. All other trade
     series are preserved.
  2. CPI: removes 8 old flat CPI series and replaces with CPI NSA (85 items)
     and CPI SA (11 items) with dim1_group optgroup support. The other 4
     Prices series (median, trimmed mean, IPPI, RMPI) are preserved.
  3. Category order: Prices and Trade appear just before Business and industry.

Usage:  python3 update_vectors.py
Output: Vectors.xlsx (overwritten in-place)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INPUT_PATH  = "/Users/jasonkirby/Desktop/StatCanApp/vectors-template.xlsx"
OUTPUT_PATH = "/Users/jasonkirby/Desktop/StatCanApp/Vectors.xlsx"

# ─────────────────────────────────────────────────────────────────
# TRADE VECTOR FORMULA (verified in memory)
# vec = PROV_BASE + prov_offset + flow_offset + comm_idx * 29 + partner_idx
# ─────────────────────────────────────────────────────────────────
PROV_BASE = 1567082903

PROVINCE_OFFSETS = {
    "Canada":                       0,
    "Newfoundland and Labrador":    1131,
    "Prince Edward Island":         1885,
    "Nova Scotia":                  2639,
    "New Brunswick":                3393,
    "Quebec":                       4147,
    "Ontario":                      4901,
    "Manitoba":                     5655,
    "Saskatchewan":                 6409,
    "Alberta":                      7163,
    "British Columbia":             7917,
}

# Canada has 3 flows; provinces have 2
FLOW_OFFSETS_CANADA = {
    "Import":           0,
    "Domestic export":  377,
    "Re-export":        754,
}
FLOW_OFFSETS_PROVINCE = {
    "Import":           0,
    "Domestic export":  377,
}

PARTNER_IDX = {
    "All countries":  0,
    "United States":  1,
    "China":          2,
    "Mexico":         3,
    "United Kingdom": 4,
    "Japan":          5,
    "Germany":        6,
}
PARTNER_SHORT = {
    "All countries":  "All",
    "United States":  "US",
    "China":          "China",
    "Mexico":         "Mexico",
    "United Kingdom": "UK",
    "Japan":          "Japan",
    "Germany":        "Germany",
}

# 13 NAPCS level-1 commodity categories (comm_idx 0–12)
COMMODITIES = [
    (0,  "Total of all merchandise",                                    "Total"),
    (1,  "Farm, fishing and intermediate food products",                 "Farm & fishing"),
    (2,  "Energy products",                                             "Energy"),
    (3,  "Metal ores and non-metallic minerals",                        "Metal ores"),
    (4,  "Metal and non-metallic mineral products",                     "Metal products"),
    (5,  "Basic and industrial chemical, plastic and rubber products",   "Chemicals"),
    (6,  "Forestry products and building and packaging materials",       "Forestry"),
    (7,  "Industrial machinery, equipment and parts",                   "Industrial machinery"),
    (8,  "Electronic and electrical equipment and parts",               "Electronics"),
    (9,  "Motor vehicles and parts",                                    "Motor vehicles"),
    (10, "Aircraft and other transportation equipment and parts",       "Aircraft"),
    (11, "Consumer goods",                                              "Consumer goods"),
    (12, "Special transactions trade",                                  "Special transactions"),
]

def trade_vector(province, flow, comm_idx, partner):
    prov_off = PROVINCE_OFFSETS[province]
    if province == "Canada":
        flow_off = FLOW_OFFSETS_CANADA[flow]
    else:
        flow_off = FLOW_OFFSETS_PROVINCE[flow]
    partner_off = PARTNER_IDX[partner]
    return PROV_BASE + prov_off + flow_off + comm_idx * 29 + partner_off


def build_trade_rows():
    """
    Build all trade rows for the 4-dim goods_prov_comm_partner series.
    Provinces: Canada + NL through BC.
    Flows: Import, Domestic export.
    Commodities: all 13 NAPCS level-1 categories.
    Partners: All countries, US, China, Mexico, UK, Japan, Germany.
    """
    rows = []
    cat        = "Trade"
    freq       = "M"
    series_id  = "goods_prov_comm_partner"
    series_name= "Goods trade by province, commodity and partner"
    table_id   = "12-10-0175-01"

    provinces = [
        "Canada",
        "Newfoundland and Labrador",
        "Prince Edward Island",
        "Nova Scotia",
        "New Brunswick",
        "Quebec",
        "Ontario",
        "Manitoba",
        "Saskatchewan",
        "Alberta",
        "British Columbia",
    ]
    flows    = ["Import", "Domestic export"]
    partners = ["All countries", "United States", "China", "Mexico",
                "United Kingdom", "Japan", "Germany"]

    for prov in provinces:
        for flow in flows:
            for comm_idx, comm_name, comm_short in COMMODITIES:
                for partner in partners:
                    vec = trade_vector(prov, flow, comm_idx, partner)
                    p_short = PARTNER_SHORT[partner]
                    full_label  = f"{prov} — {flow} — {comm_name} — {partner}"
                    short_label = p_short
                    rows.append((
                        cat, freq, series_id, series_name, table_id,
                        "Province",   prov,       # dim1
                        "Flow",       flow,        # dim2
                        "Commodity",  comm_name,   # dim3
                        "Partner",    partner,     # dim4
                        "", "",                    # dim5 (unused)
                        vec,
                        full_label,
                        short_label,
                        "",  # dim1_group (not used for trade)
                    ))
    return rows


# ─────────────────────────────────────────────────────────────────
# CPI NSA — 85 items from table 18-10-0004-01
# Structure: (dim1_value, vector_id, dim1_group)
# ─────────────────────────────────────────────────────────────────
CPI_NSA_ITEMS = [
    # group="" (top-level, no optgroup)
    ("All-items",                                           "41690973",  ""),

    # Food group
    ("Food",                                                "41690974",  "Food"),
    ("— Food purchased from stores",                        "41690975",  "Food"),
    ("  — Meat",                                            "41690976",  "Food"),
    ("    — Fresh or frozen meat (excl. poultry)",          "41690977",  "Food"),
    ("      — Fresh or frozen beef",                        "41690978",  "Food"),
    ("      — Fresh or frozen pork",                        "41690979",  "Food"),
    ("      — Other fresh or frozen meat (excl. poultry)", "41690980",  "Food"),
    ("    — Fresh or frozen poultry",                       "41690981",  "Food"),
    ("      — Fresh or frozen chicken",                     "41690982",  "Food"),
    ("      — Other fresh or frozen poultry",               "41690983",  "Food"),
    ("    — Processed meat",                                "41690984",  "Food"),
    ("      — Ham and bacon",                               "41690985",  "Food"),
    ("      — Other processed meat",                        "41690986",  "Food"),
    ("  — Fish, seafood and other marine products",         "41690987",  "Food"),
    ("    — Fish",                                          "41690988",  "Food"),
    ("    — Seafood and other marine products",             "41690991",  "Food"),
    ("  — Dairy products and eggs",                         "41690992",  "Food"),
    ("    — Dairy products",                                "41690993",  "Food"),
    ("    — Eggs",                                          "41690999",  "Food"),
    ("  — Bakery and cereal products",                      "41691000",  "Food"),
    ("    — Bakery products",                               "41691001",  "Food"),
    ("    — Cereal products",                               "41691005",  "Food"),
    ("  — Fruit, fruit preparations and nuts",              "41691010",  "Food"),
    ("    — Fresh fruit",                                   "41691011",  "Food"),
    ("    — Preserved fruit and fruit preparations",        "41691016",  "Food"),
    ("    — Nuts and seeds",                                "41691019",  "Food"),
    ("  — Vegetables and vegetable preparations",           "41691020",  "Food"),
    ("    — Fresh vegetables",                              "41691021",  "Food"),
    ("    — Preserved vegetables and vegetable preparations","41691026", "Food"),
    ("  — Other food products and non-alcoholic beverages", "41691029",  "Food"),
    ("    — Sugar and confectionery",                       "41691030",  "Food"),
    ("    — Edible fats and oils",                          "41691033",  "Food"),
    ("    — Coffee and tea",                                "41691036",  "Food"),
    ("    — Condiments, spices and vinegars",               "41691039",  "Food"),
    ("    — Other food preparations",                       "41691040",  "Food"),
    ("    — Non-alcoholic beverages",                       "41691045",  "Food"),
    ("— Food purchased from restaurants",                   "41691046",  "Food"),

    # Shelter group
    ("Shelter",                                             "41691050",  "Shelter"),
    ("— Rented accommodation",                              "41691051",  "Shelter"),
    ("  — Rent",                                            "41691052",  "Shelter"),
    ("— Owned accommodation",                               "41691055",  "Shelter"),
    ("  — Mortgage interest cost",                          "41691056",  "Shelter"),
    ("  — Homeowners' replacement cost",                    "41691057",  "Shelter"),
    ("— Water, fuel and electricity",                       "41691062",  "Shelter"),
    ("  — Electricity",                                     "41691063",  "Shelter"),
    ("  — Water",                                           "41691064",  "Shelter"),
    ("  — Natural gas",                                     "41691065",  "Shelter"),
    ("  — Fuel oil and other fuels",                        "41691066",  "Shelter"),

    # Household operations group
    ("Household operations, furnishings and equipment",     "41691067",  "Household operations, furnishings and equipment"),
    ("— Household operations",                              "41691068",  "Household operations, furnishings and equipment"),
    ("— Household furnishings and equipment",               "41691087",  "Household operations, furnishings and equipment"),

    # Clothing group
    ("Clothing and footwear",                               "41691108",  "Clothing and footwear"),
    ("— Clothing",                                          "41691109",  "Clothing and footwear"),
    ("— Footwear",                                          "41691113",  "Clothing and footwear"),
    ("— Clothing accessories, watches and jewellery",       "41691118",  "Clothing and footwear"),
    ("— Clothing material, notions and services",           "41691123",  "Clothing and footwear"),

    # Transportation group
    ("Transportation",                                      "41691128",  "Transportation"),

    # Health group
    ("Health and personal care",                            "41691153",  "Health and personal care"),
    ("— Health care",                                       "41691154",  "Health and personal care"),
    ("— Personal care",                                     "41691163",  "Health and personal care"),

    # Recreation group
    ("Recreation, education and reading",                   "41691170",  "Recreation, education and reading"),
    ("— Recreation",                                        "41691171",  "Recreation, education and reading"),
    ("  — Recreational equipment and services (excl. recreational vehicles)", "41691172", "Recreation, education and reading"),
    ("  — Purchase and operation of recreational vehicles", "41691179",  "Recreation, education and reading"),
    ("  — Home entertainment equipment, parts and services","41691184",  "Recreation, education and reading"),
    ("  — Travel services",                                 "41691190",  "Recreation, education and reading"),
    ("  — Other cultural and recreational services",        "41691193",  "Recreation, education and reading"),
    ("— Education and reading",                             "41691197",  "Recreation, education and reading"),
    ("  — Education",                                       "41691198",  "Recreation, education and reading"),
    ("  — Reading material (excl. textbooks)",              "41691202",  "Recreation, education and reading"),

    # Alcoholic beverages group
    ("Alcoholic beverages, tobacco and cannabis",           "41691206",  "Alcoholic beverages, tobacco and cannabis"),
    ("— Alcoholic beverages",                               "41691207",  "Alcoholic beverages, tobacco and cannabis"),
    ("— Tobacco products and smokers' supplies",            "41691216",  "Alcoholic beverages, tobacco and cannabis"),
    ("— Recreational cannabis",                             "1043024072","Alcoholic beverages, tobacco and cannabis"),

    # Special aggregates group
    ("All-items excl. food",                                "41691232",  "Special aggregates"),
    ("All-items excl. food and energy",                     "41691233",  "Special aggregates"),
    ("All-items excl. mortgage interest cost",              "41691240",  "Special aggregates"),
    ("All-items excl. shelter",                             "41691234",  "Special aggregates"),
    ("All-items excl. energy",                              "41691238",  "Special aggregates"),
    ("All-items excl. gasoline",                            "41693245",  "Special aggregates"),
    ("Private transportation excl. gasoline",               "43972163",  "Special aggregates"),
    ("Food and energy",                                     "41691237",  "Special aggregates"),
    ("Fresh fruit and vegetables",                          "41691235",  "Special aggregates"),
    ("Energy",                                              "41691239",  "Special aggregates"),
]

# ─────────────────────────────────────────────────────────────────
# CPI SA — 11 items from table 18-10-0006-01
# ─────────────────────────────────────────────────────────────────
CPI_SA_ITEMS = [
    ("All-items",                                                       "41690914", ""),
    ("Food",                                                            "41690915", "Major components"),
    ("Shelter",                                                         "41690916", "Major components"),
    ("Household operations, furnishings and equipment",                 "41690917", "Major components"),
    ("Clothing and footwear",                                           "41690918", "Major components"),
    ("Transportation",                                                  "41690919", "Major components"),
    ("Health and personal care",                                        "41690920", "Major components"),
    ("Recreation, education and reading",                               "41690921", "Major components"),
    ("Alcoholic beverages, tobacco and cannabis",                       "41690922", "Major components"),
    ("All-items excl. food",                                            "41690923", "Special aggregates"),
    ("All-items excl. food and energy",                                 "41690924", "Special aggregates"),
]

# Old CPI series IDs to remove from the template (replaced by grouped versions above)
OLD_CPI_SERIES_IDS = {
    "prc_cpi_nsa", "prc_cpi_food_nsa", "prc_cpi_foodstores",
    "prc_cpi_exfood_nsa", "prc_cpi_core_nsa",
    "prc_cpi_sa", "prc_cpi_exfood_sa", "prc_cpi_core_sa",
}

# Old trade series ID to remove (the old 3-dim version, replaced by 4-dim)
OLD_TRADE_SERIES_ID = "goods_prov_comm_partner"


def build_cpi_rows(items, series_id, series_name, table_id):
    rows = []
    cat  = "Prices"
    freq = "M"
    for label, vector, group in items:
        clean_label = label.lstrip("— ").strip()
        full_label  = f"CPI — {clean_label}"
        short_label = clean_label
        rows.append((
            cat, freq, series_id, series_name, table_id,
            "Category", label,  # dim1_name, dim1_value
            "", "",              # dim2
            "", "",              # dim3
            "", "",              # dim4
            "", "",              # dim5
            vector,
            full_label,
            short_label,
            group,              # dim1_group
        ))
    return rows


# Desired final category order — Trade and Prices just before Business and industry
CAT_ORDER = [
    "GDP (expenditure based)",
    "GDP (by industry)",
    "Labour",
    "Prices",
    "Trade",
    "Sales",
    "Business and industry",
    "Population",
    "Travel",
    "National balance sheets",
    "Housing and households",
    "International securities",
    "Bank of Canada",
    "Federal government finances",
    "Energy",
    "Agriculture",
]

NEEDED_HEADERS = [
    "category", "freq", "series_id", "series_name", "table_id",
    "dim1_name", "dim1_value", "dim2_name", "dim2_value",
    "dim3_name", "dim3_value", "dim4_name", "dim4_value",
    "dim5_name", "dim5_value",
    "vector", "full_label", "short_label", "dim1_group",
]


def main():
    print(f"Reading {INPUT_PATH}...")
    wb = openpyxl.load_workbook(INPUT_PATH)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    print(f"Existing headers: {headers}")

    h = {v: i for i, v in enumerate(headers) if v}

    def get(row, col_name):
        idx = h.get(col_name)
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return str(v).strip() if v is not None else ""

    def remap(old_row):
        """Map an existing row (old header layout) to NEEDED_HEADERS layout."""
        return [get(old_row, col) for col in NEEDED_HEADERS]

    # Read all existing data rows
    existing_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            existing_rows.append(list(row))

    # ── Group existing rows by category, filtering out only the replaced series
    rows_by_cat = {}  # category -> list of remapped rows (in original order)
    removed = 0
    for row in existing_rows:
        cat_val = get(row, "category")
        sid     = get(row, "series_id")

        # Drop old 3-dim trade series (replaced by 4-dim version)
        if cat_val == "Trade" and sid == OLD_TRADE_SERIES_ID:
            removed += 1
            continue

        # Drop old flat CPI series (replaced by grouped versions)
        if cat_val == "Prices" and sid in OLD_CPI_SERIES_IDS:
            removed += 1
            continue

        if cat_val not in rows_by_cat:
            rows_by_cat[cat_val] = []
        rows_by_cat[cat_val].append(remap(row))

    print(f"Dropped {removed} replaced rows from template.")

    # ── Build new rows ─────────────────────────────────────────────
    new_trade = [list(r) for r in build_trade_rows()]
    new_cpi_nsa = [list(r) for r in build_cpi_rows(
        CPI_NSA_ITEMS, "cpi_nsa", "CPI, NSA", "18-10-0004-01")]
    new_cpi_sa  = [list(r) for r in build_cpi_rows(
        CPI_SA_ITEMS,  "cpi_sa",  "CPI, SA",  "18-10-0006-01")]

    print(f"New trade rows: {len(new_trade)}")
    print(f"New CPI NSA rows: {len(new_cpi_nsa)}")
    print(f"New CPI SA rows: {len(new_cpi_sa)}")

    # ── Assemble all rows in desired category order ────────────────
    all_rows = []
    for cat in CAT_ORDER:
        kept = rows_by_cat.get(cat, [])
        if cat == "Trade":
            # Kept other trade series + new 4-dim province series
            all_rows.extend(kept + new_trade)
        elif cat == "Prices":
            # New CPI rows first, then kept prices series (median, trimmed, IPPI, RMPI)
            all_rows.extend(new_cpi_nsa + new_cpi_sa + kept)
        else:
            all_rows.extend(kept)

    # Any categories not in CAT_ORDER (future-proofing) go at the end
    for cat, kept in rows_by_cat.items():
        if cat not in CAT_ORDER:
            print(f"  Warning: category '{cat}' not in CAT_ORDER, appending at end")
            all_rows.extend(kept)

    # ── Rebuild workbook ──────────────────────────────────────────
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "series"

    header_font = Font(bold=True)
    for ci, h_name in enumerate(NEEDED_HEADERS, start=1):
        cell = ws2.cell(row=1, column=ci, value=h_name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    fill_grey  = PatternFill(fill_type="solid", fgColor="E8E8E8")
    fill_white = PatternFill(fill_type="solid", fgColor="FFFFFF")
    toggle = False
    prev_cat = None

    for ri, row_data in enumerate(all_rows, start=2):
        cat_val = row_data[0] if row_data else ""
        if cat_val != prev_cat:
            toggle = not toggle
            prev_cat = cat_val
        fill = fill_grey if toggle else fill_white
        for ci, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.fill = fill

    ws2.freeze_panes = "A2"
    for ci, col_cells in enumerate(ws2.iter_cols(min_row=1, max_row=ws2.max_row), start=1):
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=8)
        ws2.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 60)

    wb2.save(OUTPUT_PATH)
    print(f"\nSaved {OUTPUT_PATH}")
    print(f"  Total data rows: {len(all_rows)}")

    # Print category summary
    cats_seen = []
    cat_counts = {}
    for row in all_rows:
        c = row[0]
        if c not in cat_counts:
            cats_seen.append(c)
            cat_counts[c] = 0
        cat_counts[c] += 1
    print("\nCategory row counts:")
    for c in cats_seen:
        print(f"  {c}: {cat_counts[c]}")


if __name__ == "__main__":
    main()
